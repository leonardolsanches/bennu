# -*- coding: utf-8 -*-
"""
Rotas para relatórios financeiros
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text, or_
from app.database import get_db
from app.auth.oauth import get_current_user
from app.models.transacoes import TransacaoFinanceira
from app.models.categorias import CategoriaContabil, CategoriaGerencial, CentroCusto
from app.models.empresas import Empresa
from app.models.clientes import Cliente
from app.models.fornecedores import Fornecedor # Import Fornecedor model
from app.models.auxiliares import Imposto, ProdutoServico
from app.models.categorias import Projeto
from app.models.planejamento import LinhaOrcamentaria, PlanejamentoVersao
from typing import Optional, Dict, List, Any
import calendar
from datetime import datetime, timedelta
import unicodedata
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

router = APIRouter()


def normalizar_string(texto: str) -> str:
    """Normaliza string removendo acentos e caracteres especiais"""
    if not texto:
        return ""
    # Remover acentos
    texto_normalizado = unicodedata.normalize('NFD', texto)
    texto_sem_acentos = ''.join(c for c in texto_normalizado if unicodedata.category(c) != 'Mn')
    # Converter para minúsculas e remover espaços extras
    return re.sub(r'\s+', ' ', texto_sem_acentos.lower().strip())


def eh_categoria_dividendo(nome_categoria: str) -> bool:
    """Verifica se categoria representa dividendos/distribuições que não devem aparecer no gerencial"""
    if not nome_categoria:
        return False

    nome_normalizado = normalizar_string(nome_categoria)

    termos_dividendos = [
        "dividendo", "dividendos", 
        "distribuicao", "distribuicoes",
        "lucro distribuido", "lucros distribuidos",
        "pro labore", "prolabore", "pro-labore",
        "retirada socio", "retiradas socio", "retirada de socio",
        "distribuicao lucro", "distribuicao de lucro"
    ]

    return any(termo in nome_normalizado for termo in termos_dividendos)


def obter_aliquotas_impostos(db: Session, empresa_id: int) -> Dict[str, float]:
    """
    Busca alíquotas de impostos cadastradas para uma empresa

    Args:
        db: Sessão do banco de dados
        empresa_id: ID da empresa

    Returns:
        Dicionário com alíquotas dos impostos (PIS, COFINS, ISS, IRPJ, CSLL)
    """
    # Valores default caso não encontre cadastro
    aliquotas_default = {
        "PIS": 0.65,      # 0,65%
        "COFINS": 3.0,    # 3%
        "ISS": 5.0,       # 5% (default)
        "IRPJ": 7.93,     # 7,93%
        "CSLL": 2.88      # 2,88%
    }

    # Buscar impostos cadastrados para a empresa
    impostos = db.query(Imposto).filter(
        Imposto.empresa_id == empresa_id,
        Imposto.ativo == True
    ).all()

    # Mapear impostos por código
    aliquotas = {}
    for imposto in impostos:
        codigo = imposto.codigo.upper() if imposto.codigo else imposto.nome.upper()
        # Extrair código limpo (PIS, COFINS, ISS, IRPJ, CSLL)
        for codigo_padrao in ["PIS", "COFINS", "ISS", "IRPJ", "CSLL"]:
            if codigo_padrao in codigo:
                aliquotas[codigo_padrao] = float(imposto.valor)
                break

    # Preencher valores default para impostos não cadastrados
    for codigo, valor_default in aliquotas_default.items():
        if codigo not in aliquotas:
            aliquotas[codigo] = valor_default

    return aliquotas


def calcular_impostos_sobre_receita(valor_receita: float, aliquotas: Dict[str, float]) -> Dict[str, float]:
    """
    Calcula impostos sobre receita conforme regime tributário usando alíquotas cadastradas

    Args:
        valor_receita: Valor da receita bruta
        aliquotas: Dicionário com alíquotas cadastradas (PIS, COFINS, ISS, IRPJ, CSLL) em %

    Returns:
        Dicionário com valores de cada imposto e total
    """
    # Garantir que valor_receita é float para evitar TypeError com Decimal
    valor_receita = float(valor_receita)

    if valor_receita <= 0:
        return {
            "pis": 0.0,
            "cofins": 0.0,
            "iss": 0.0,
            "irpj": 0.0,
            "csll": 0.0,
            "total": 0.0
        }

    # Converter alíquotas de percentual para decimal
    aliquota_pis = aliquotas.get("PIS", 0.65) / 100.0
    aliquota_cofins = aliquotas.get("COFINS", 3.0) / 100.0
    aliquota_iss = aliquotas.get("ISS", 5.0) / 100.0
    aliquota_irpj = aliquotas.get("IRPJ", 7.93) / 100.0
    aliquota_csll = aliquotas.get("CSLL", 2.88) / 100.0

    # Calcular cada imposto
    pis = valor_receita * aliquota_pis
    cofins = valor_receita * aliquota_cofins
    iss = valor_receita * aliquota_iss
    irpj = valor_receita * aliquota_irpj
    csll = valor_receita * aliquota_csll

    total = pis + cofins + iss + irpj + csll

    return {
        "pis": round(pis, 2),
        "cofins": round(cofins, 2),
        "iss": round(iss, 2),
        "irpj": round(irpj, 2),
        "csll": round(csll, 2),
        "total": round(total, 2)
    }


def calcular_impostos_mensais(receitas_mes: List[float], aliquotas: Dict[str, float]) -> Dict[str, Any]:
    """
    Calcula impostos sobre receitas mensais usando alíquotas cadastradas

    Args:
        receitas_mes: Lista com 12 valores de receitas mensais
        aliquotas: Dicionário com alíquotas cadastradas (PIS, COFINS, ISS, IRPJ, CSLL)

    Returns:
        Dicionário com impostos por mês e totais
    """
    resultado = {
        "pis_mes": [0.0] * 12,
        "cofins_mes": [0.0] * 12,
        "iss_mes": [0.0] * 12,
        "irpj_mes": [0.0] * 12,
        "csll_mes": [0.0] * 12,
        "total_impostos_mes": [0.0] * 12,
        "pis_total": 0.0,
        "cofins_total": 0.0,
        "iss_total": 0.0,
        "irpj_total": 0.0,
        "csll_total": 0.0,
        "total_impostos": 0.0
    }

    for i in range(12):
        if i < len(receitas_mes):
            impostos = calcular_impostos_sobre_receita(receitas_mes[i], aliquotas)
            resultado["pis_mes"][i] = impostos["pis"]
            resultado["cofins_mes"][i] = impostos["cofins"]
            resultado["iss_mes"][i] = impostos["iss"]
            resultado["irpj_mes"][i] = impostos["irpj"]
            resultado["csll_mes"][i] = impostos["csll"]
            resultado["total_impostos_mes"][i] = impostos["total"]

            # Acumular totais
            resultado["pis_total"] += impostos["pis"]
            resultado["cofins_total"] += impostos["cofins"]
            resultado["iss_total"] += impostos["iss"]
            resultado["irpj_total"] += impostos["irpj"]
            resultado["csll_total"] += impostos["csll"]
            resultado["total_impostos"] += impostos["total"]

    return resultado

@router.get("/relatorios/pl-contabil")
async def get_pl_contabil(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None),
    fornecedor: Optional[int] = Query(None),
    centro_custo: Optional[int] = Query(None)
):
    """
    Relatório P&L Contábil - Demonstração de Resultado por mês
    Filtra apenas transações com entra_no_gerencial=True (checkbox marcado)
    Com hierarquia: Receitas (Cliente > Projeto > Produto), Despesas (Categoria > Subcategoria > Título)
    """
    try:
        # Query base - FILTRAR APENAS transações que entram no gerencial
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.entra_no_gerencial == True  # Checkbox marcado
        )

        # Filtro por ano
        query = query.filter(TransacaoFinanceira.competencia_ano == ano)

        # Aplicar filtros opcionais
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)
        if cliente:
            query = query.filter(TransacaoFinanceira.cliente_id == cliente)
        if projeto:
            query = query.filter(TransacaoFinanceira.projeto_id == projeto)
        if produto_servico:
            query = query.filter(TransacaoFinanceira.produto_servico_id == produto_servico)
        if fornecedor:
            query = query.filter(TransacaoFinanceira.fornecedor_id == fornecedor)
        if centro_custo:
            query = query.filter(TransacaoFinanceira.centro_custo_id == centro_custo)

        # Buscar todas as transações do ano
        transacoes = query.all()

        # Inicializar estrutura de dados hierárquica
        resultado = {
            "ano": ano,
            "empresa_id": empresa,
            "receitas": {},
            "despesas": {},
            "totais": {
                "receitas_mes": [0] * 12,
                "despesas_mes": [0] * 12,
                "resultado_mes": [0] * 12,
                "receitas_total": 0,
                "despesas_total": 0,
                "resultado_total": 0
            }
        }

        # Processar transações por categoria e mês
        for transacao in transacoes:
            try:
                # Validar dados básicos
                if not transacao.competencia_mes or transacao.valor is None:
                    continue

                mes = int(transacao.competencia_mes) - 1  # 0-based para indexação
                if mes < 0 or mes > 11:  # Validar mês
                    continue

                valor = float(transacao.valor)
                tipo_transacao = str(transacao.tipo) if transacao.tipo else 'unknown'

                # RECEITAS: Cliente > Projeto > Produto/Serviço
                if tipo_transacao == 'receita':
                    cliente_nome = "Sem Cliente"
                    if transacao.cliente_id:
                        try:
                            cliente = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                            if cliente and cliente.nome:
                                cliente_nome = str(cliente.nome)
                        except:
                            pass

                    projeto_nome = "Sem Projeto"
                    if transacao.projeto_id:
                        try:
                            projeto_obj = db.query(Projeto).filter(Projeto.id == transacao.projeto_id).first()
                            if projeto_obj and projeto_obj.nome:
                                projeto_nome = str(projeto_obj.nome)
                        except:
                            pass

                    produto_nome = "Sem Produto/Serviço"
                    if transacao.produto_servico_id:
                        try:
                            from app.models.auxiliares import ProdutoServico
                            produto_obj = db.query(ProdutoServico).filter(ProdutoServico.id == transacao.produto_servico_id).first()
                            if produto_obj and produto_obj.nome:
                                produto_nome = str(produto_obj.nome)
                        except:
                            pass

                    # Criar hierarquia Cliente > Projeto > Produto
                    if cliente_nome not in resultado["receitas"]:
                        resultado["receitas"][cliente_nome] = {
                            "meses": [0.0] * 12,
                            "total": 0.0,
                            "projetos": {}
                        }

                    if projeto_nome not in resultado["receitas"][cliente_nome]["projetos"]:
                        resultado["receitas"][cliente_nome]["projetos"][projeto_nome] = {
                            "meses": [0.0] * 12,
                            "total": 0.0,
                            "produtos": {}
                        }

                    if produto_nome not in resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["produtos"]:
                        resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome] = {
                            "meses": [0.0] * 12,
                            "total": 0.0
                        }

                    # Acumular valores em todos os níveis
                    resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["meses"][mes] += valor
                    resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["total"] += valor
                    resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["meses"][mes] += valor
                    resultado["receitas"][cliente_nome]["projetos"][projeto_nome]["total"] += valor
                    resultado["receitas"][cliente_nome]["meses"][mes] += valor
                    resultado["receitas"][cliente_nome]["total"] += valor
                    resultado["totais"]["receitas_mes"][mes] += valor
                    resultado["totais"]["receitas_total"] += valor

                # DESPESAS: Centro de Custo > Categoria Contábil > Título
                elif tipo_transacao == 'despesa':
                    centro_custo_nome = "Sem Centro de Custo"
                    if transacao.centro_custo_id:
                        try:
                            centro = db.query(CentroCusto).filter(CentroCusto.id == transacao.centro_custo_id).first()
                            if centro and centro.nome:
                                centro_custo_nome = str(centro.nome)
                        except:
                            pass

                    categoria_nome = "Sem Categoria"
                    if transacao.categoria_contabil_id:
                        try:
                            categoria = db.query(CategoriaContabil).filter(CategoriaContabil.id == transacao.categoria_contabil_id).first()
                            if categoria and categoria.nome:
                                categoria_nome = str(categoria.nome)
                        except:
                            pass

                    titulo_breve = transacao.titulo_breve or transacao.nome or "Sem Título"

                    valor_abs = abs(valor)

                    # Criar hierarquia Centro de Custo > Categoria > Título
                    if centro_custo_nome not in resultado["despesas"]:
                        resultado["despesas"][centro_custo_nome] = {
                            "meses": [0.0] * 12,
                            "total": 0.0,
                            "categorias": {}
                        }

                    if categoria_nome not in resultado["despesas"][centro_custo_nome]["categorias"]:
                        resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome] = {
                            "meses": [0.0] * 12,
                            "total": 0.0,
                            "titulos": {}
                        }

                    if titulo_breve not in resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["titulos"]:
                        resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["titulos"][titulo_breve] = {
                            "meses": [0.0] * 12,
                            "total": 0.0
                        }

                    # Acumular valores em todos os níveis
                    resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["titulos"][titulo_breve]["meses"][mes] += valor_abs
                    resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["titulos"][titulo_breve]["total"] += valor_abs
                    resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["meses"][mes] += valor_abs
                    resultado["despesas"][centro_custo_nome]["categorias"][categoria_nome]["total"] += valor_abs
                    resultado["despesas"][centro_custo_nome]["meses"][mes] += valor_abs
                    resultado["despesas"][centro_custo_nome]["total"] += valor_abs
                    resultado["totais"]["despesas_mes"][mes] += valor_abs
                    resultado["totais"]["despesas_total"] += valor_abs

            except Exception as e:
                continue

        # Calcular resultado líquido por mês
        for i in range(12):
            resultado["totais"]["resultado_mes"][i] = (
                resultado["totais"]["receitas_mes"][i] - 
                resultado["totais"]["despesas_mes"][i]
            )

        resultado["totais"]["resultado_total"] = (
            resultado["totais"]["receitas_total"] - 
            resultado["totais"]["despesas_total"]
        )

        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar P&L Contábil: {str(e)}")


@router.get("/relatorios/pl-gerencial")
async def get_pl_gerencial(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    mes_corte: Optional[int] = Query(None, description="Mês de corte (1-12) para dividir faturado vs projetado")
):
    """
    Relatório P&L Gerencial - Demonstração com Planejado vs Realizado
    Receita por Cliente, Despesa por Categorias Gerenciais
    Dividendos não aparecem como despesa
    Integra dados de planejamento orçamentário (linhas_orcamentarias)
    Suporta mês de corte para dividir faturado (até o mês) vs projetado (após o mês)
    """
    try:
        # Query para dados realizados (efetivamente pagos)
        query_realizado = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.competencia_ano == ano,
            TransacaoFinanceira.entra_no_gerencial == True,  # Filhos CONTAM, pais desmembrados NÃO
            TransacaoFinanceira.data_pagamento.isnot(None)  # Efetivamente pago
        )

        # Query para dados planejados - transações não pagas
        query_planejado = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.competencia_ano == ano,
            TransacaoFinanceira.entra_no_gerencial == True,  # Filhos CONTAM, pais desmembrados NÃO
            TransacaoFinanceira.data_pagamento.is_(None),  # Ainda não pago
            TransacaoFinanceira.data_vencimento.isnot(None)  # Tem vencimento programado
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query_realizado = query_realizado.filter(TransacaoFinanceira.empresa_id == empresa)
            query_planejado = query_planejado.filter(TransacaoFinanceira.empresa_id == empresa)

        # Buscar transações
        transacoes_realizadas = query_realizado.all()
        transacoes_planejadas = query_planejado.all()

        # Query para dados de planejamento orçamentário (linhas_orcamentarias)
        query_orcamento = db.query(LinhaOrcamentaria).join(
            PlanejamentoVersao,
            LinhaOrcamentaria.versao_id == PlanejamentoVersao.id
        ).filter(
            LinhaOrcamentaria.ano == ano,
            PlanejamentoVersao.is_ativo == True  # Apenas versões ativas
        )

        if empresa:
            query_orcamento = query_orcamento.filter(LinhaOrcamentaria.empresa_id == empresa)

        linhas_orcamento = query_orcamento.all()

        # Inicializar estrutura de dados
        resultado = {
            "ano": ano,
            "empresa_id": empresa,
            "tipo": "gerencial",
            "receitas": {},
            "despesas": {},
            "totais": {
                "receitas_planejado_mes": [0] * 12,
                "receitas_realizado_mes": [0] * 12,
                "despesas_planejado_mes": [0] * 12,
                "despesas_realizado_mes": [0] * 12,
                "resultado_planejado_mes": [0] * 12,
                "resultado_realizado_mes": [0] * 12,
                "receitas_planejado_total": 0,
                "receitas_realizado_total": 0,
                "despesas_planejado_total": 0,
                "despesas_realizado_total": 0,
                "resultado_planejado_total": 0,
                "resultado_realizado_total": 0
            }
        }

        # Função auxiliar para processar transações
        def processar_transacao(transacao, tipo_dados="realizado"):
            try:
                # Validar dados básicos
                if not transacao.competencia_mes or transacao.valor is None:
                    return

                mes = int(transacao.competencia_mes) - 1
                if mes < 0 or mes > 11:
                    return

                valor = float(transacao.valor)
                tipo_transacao = str(transacao.tipo) if transacao.tipo else 'unknown'

                if tipo_transacao == 'receita':
                    # P&L Gerencial: Receita por Cliente
                    cliente_nome = "Sem Cliente"
                    if transacao.cliente_id:
                        try:
                            cliente = db.query(Cliente).filter(
                                Cliente.id == transacao.cliente_id
                            ).first()
                            if cliente and cliente.nome:
                                cliente_nome = str(cliente.nome)
                        except:
                            pass

                    if cliente_nome not in resultado["receitas"]:
                        resultado["receitas"][cliente_nome] = {
                            "planejado_meses": [0.0] * 12,
                            "realizado_meses": [0.0] * 12,
                            "planejado_total": 0.0,
                            "realizado_total": 0.0
                        }

                    if tipo_dados == "realizado":
                        resultado["receitas"][cliente_nome]["realizado_meses"][mes] += valor
                        resultado["receitas"][cliente_nome]["realizado_total"] += valor
                        resultado["totais"]["receitas_realizado_mes"][mes] += valor
                        resultado["totais"]["receitas_realizado_total"] += valor
                    else:  # planejado
                        resultado["receitas"][cliente_nome]["planejado_meses"][mes] += valor
                        resultado["receitas"][cliente_nome]["planejado_total"] += valor
                        resultado["totais"]["receitas_planejado_mes"][mes] += valor
                        resultado["totais"]["receitas_planejado_total"] += valor

                elif tipo_transacao == 'despesa':
                    # P&L Gerencial: Despesa por Categorias Gerenciais (não contábeis)
                    categoria_nome = "Sem Categoria Gerencial"
                    eh_dividendo = False

                    if transacao.categoria_gerencial_id:
                        try:
                            categoria = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == transacao.categoria_gerencial_id
                            ).first()
                            if categoria and categoria.nome:
                                categoria_nome = str(categoria.nome)

                                # Filtrar dividendos - método robusto
                                eh_dividendo = eh_categoria_dividendo(categoria_nome)
                        except:
                            pass

                    # Pular dividendos no gerencial
                    if eh_dividendo:
                        return

                    valor_abs = abs(valor)
                    if categoria_nome not in resultado["despesas"]:
                        resultado["despesas"][categoria_nome] = {
                            "planejado_meses": [0.0] * 12,
                            "realizado_meses": [0.0] * 12,
                            "planejado_total": 0.0,
                            "realizado_total": 0.0
                        }

                    if tipo_dados == "realizado":
                        resultado["despesas"][categoria_nome]["realizado_meses"][mes] += valor_abs
                        resultado["despesas"][categoria_nome]["realizado_total"] += valor_abs
                        resultado["totais"]["despesas_realizado_mes"][mes] += valor_abs
                        resultado["totais"]["despesas_realizado_total"] += valor_abs
                    else:  # planejado
                        resultado["despesas"][categoria_nome]["planejado_meses"][mes] += valor_abs
                        resultado["despesas"][categoria_nome]["planejado_total"] += valor_abs
                        resultado["totais"]["despesas_planejado_mes"][mes] += valor_abs
                        resultado["totais"]["despesas_planejado_total"] += valor_abs

            except Exception as e:
                pass

        # Função para processar linhas orçamentárias (planejamento)
        def processar_linha_orcamento(linha):
            try:
                # Validar dados básicos
                if not linha.mes or linha.valor_previsto is None:
                    return

                mes = int(linha.mes) - 1
                if mes < 0 or mes > 11:
                    return

                valor = float(linha.valor_previsto)
                categoria_tipo = str(linha.categoria) if linha.categoria else 'outros'

                # Processar RECEITAS do orçamento
                if categoria_tipo == 'receita':
                    cliente_nome = "Sem Cliente"
                    if linha.cliente_id:
                        try:
                            cliente = db.query(Cliente).filter(
                                Cliente.id == linha.cliente_id
                            ).first()
                            if cliente and cliente.nome:
                                cliente_nome = str(cliente.nome)
                        except:
                            pass

                    if cliente_nome not in resultado["receitas"]:
                        resultado["receitas"][cliente_nome] = {
                            "planejado_meses": [0.0] * 12,
                            "realizado_meses": [0.0] * 12,
                            "planejado_total": 0.0,
                            "realizado_total": 0.0
                        }

                    resultado["receitas"][cliente_nome]["planejado_meses"][mes] += valor
                    resultado["receitas"][cliente_nome]["planejado_total"] += valor
                    resultado["totais"]["receitas_planejado_mes"][mes] += valor
                    resultado["totais"]["receitas_planejado_total"] += valor

                # Processar DESPESAS do orçamento
                elif categoria_tipo == 'despesa':
                    categoria_nome = "Sem Categoria Gerencial"
                    eh_dividendo = False

                    if linha.categoria_gerencial_id:
                        try:
                            categoria = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == linha.categoria_gerencial_id
                            ).first()
                            if categoria and categoria.nome:
                                categoria_nome = str(categoria.nome)
                                eh_dividendo = eh_categoria_dividendo(categoria_nome)
                        except:
                            pass

                    # Pular dividendos
                    if eh_dividendo:
                        return

                    valor_abs = abs(valor)
                    if categoria_nome not in resultado["despesas"]:
                        resultado["despesas"][categoria_nome] = {
                            "planejado_meses": [0.0] * 12,
                            "realizado_meses": [0.0] * 12,
                            "planejado_total": 0.0,
                            "realizado_total": 0.0
                        }

                    resultado["despesas"][categoria_nome]["planejado_meses"][mes] += valor_abs
                    resultado["despesas"][categoria_nome]["planejado_total"] += valor_abs
                    resultado["totais"]["despesas_planejado_mes"][mes] += valor_abs
                    resultado["totais"]["despesas_planejado_total"] += valor_abs

            except Exception as e:
                pass

        # Processar transações realizadas
        for transacao in transacoes_realizadas:
            processar_transacao(transacao, "realizado")

        # Processar transações planejadas (contas a pagar não lançadas)
        for transacao in transacoes_planejadas:
            processar_transacao(transacao, "planejado")

        # Processar linhas orçamentárias (planejamento)
        for linha in linhas_orcamento:
            processar_linha_orcamento(linha)

        # Calcular resultados líquidos
        for i in range(12):
            resultado["totais"]["resultado_planejado_mes"][i] = (
                resultado["totais"]["receitas_planejado_mes"][i] - 
                resultado["totais"]["despesas_planejado_mes"][i]
            )
            resultado["totais"]["resultado_realizado_mes"][i] = (
                resultado["totais"]["receitas_realizado_mes"][i] - 
                resultado["totais"]["despesas_realizado_mes"][i]
            )

        resultado["totais"]["resultado_planejado_total"] = (
            resultado["totais"]["receitas_planejado_total"] - 
            resultado["totais"]["despesas_planejado_total"]
        )
        resultado["totais"]["resultado_realizado_total"] = (
            resultado["totais"]["receitas_realizado_total"] - 
            resultado["totais"]["despesas_realizado_total"]
        )

        # Adicionar mês de corte ao resultado se fornecido
        resultado["mes_corte"] = mes_corte

        # Se houver mês de corte, calcular totais de faturado (até o mês) vs projetado (após o mês)
        if mes_corte and 1 <= mes_corte <= 12:
            resultado["totais"]["faturado"] = {
                "receitas": sum(resultado["totais"]["receitas_realizado_mes"][:mes_corte]),
                "despesas": sum(resultado["totais"]["despesas_realizado_mes"][:mes_corte]),
                "resultado": 0
            }
            resultado["totais"]["faturado"]["resultado"] = (
                resultado["totais"]["faturado"]["receitas"] - 
                resultado["totais"]["faturado"]["despesas"]
            )

            resultado["totais"]["projetado"] = {
                "receitas": sum(resultado["totais"]["receitas_planejado_mes"][mes_corte:]),
                "despesas": sum(resultado["totais"]["despesas_planejado_mes"][mes_corte:]),
                "resultado": 0
            }
            resultado["totais"]["projetado"]["resultado"] = (
                resultado["totais"]["projetado"]["receitas"] - 
                resultado["totais"]["projetado"]["despesas"]
            )

        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar P&L Gerencial: {str(e)}")


@router.get("/relatorios/receita-por-cliente")
async def get_receita_por_cliente(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    empresa: Optional[int] = Query(None)
):
    """
    Relatório de Receita por Cliente
    """
    try:
        # Query base para receitas
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'receita'
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Filtros
        if ano:
            query = query.filter(TransacaoFinanceira.competencia_ano == ano)
        if mes:
            query = query.filter(TransacaoFinanceira.competencia_mes == mes)

        # Join com clientes
        query = query.outerjoin(Cliente, TransacaoFinanceira.cliente_id == Cliente.id)

        # Agrupar por cliente
        receitas_cliente = query.with_entities(
            Cliente.nome.label('cliente_nome'),
            func.sum(TransacaoFinanceira.valor).label('total_receita'),
            func.count(TransacaoFinanceira.id).label('qtd_transacoes')
        ).group_by(Cliente.id, Cliente.nome).all()

        # Processar resultados
        resultado = []
        total_geral = 0

        for item in receitas_cliente:
            cliente_nome = item.cliente_nome or "Cliente não informado"
            valor = float(item.total_receita or 0)
            total_geral += valor

            resultado.append({
                "cliente": cliente_nome,
                "valor": valor,
                "transacoes": item.qtd_transacoes,
                "percentual": 0  # Será calculado após ter o total
            })

        # Calcular percentuais e ordenar
        if total_geral > 0:
            for item in resultado:
                item["percentual"] = (item["valor"] / total_geral) * 100

        # Ordenar por valor decrescente
        resultado.sort(key=lambda x: x["valor"], reverse=True)

        return {
            "dados": resultado,
            "total_geral": total_geral,
            "filtros": {
                "ano": ano,
                "mes": mes,
                "empresa": empresa
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório de receita por cliente: {str(e)}")


def calcular_rateio_despesas(
    db: Session,
    transacoes_despesas: List[TransacaoFinanceira],
    receitas_por_cliente_mes: Dict[str, Dict[str, Any]]
) -> Dict[str, Dict[str, Any]]:
    """
    Calcula o rateio proporcional de despesas sem cliente baseado no faturamento mensal

    Args:
        db: Sessão do banco de dados
        transacoes_despesas: Lista de transações de despesa
        receitas_por_cliente_mes: Dicionário com receitas por cliente { cliente_nome: { "realizado_meses": [...], ... } }

    Returns:
        Dicionário com despesas rateadas por cliente e mês
        {
            cliente_nome: {
                "rateio_meses": [0.0] * 12,  # Despesas rateadas por mês
                "rateio_total": 0.0,          # Total de despesas rateadas
                "detalhes": {                 # Detalhamento por categoria
                    categoria_nome: {
                        "meses": [0.0] * 12,
                        "total": 0.0
                    }
                }
            }
        }
    """
    print(f"\n🔄 Iniciando cálculo de rateio automático de despesas...")

    # Estrutura para armazenar despesas rateadas
    rateio_por_cliente = {}

    # Para cada mês, calcular:
    # 1. Faturamento total de todos os clientes
    # 2. Proporção de cada cliente
    # 3. Distribuir despesas sem cliente proporcionalmente

    for mes in range(12):
        # Calcular faturamento total do mês (soma de todos os clientes)
        faturamento_total_mes = 0.0
        faturamento_por_cliente = {}

        for cliente_nome, dados_cliente in receitas_por_cliente_mes.items():
            faturamento_cliente = dados_cliente.get("realizado_meses", [0.0] * 12)[mes]
            if faturamento_cliente > 0:
                faturamento_por_cliente[cliente_nome] = faturamento_cliente
                faturamento_total_mes += faturamento_cliente

        if faturamento_total_mes == 0:
            # Não há faturamento no mês, não há o que ratear
            continue

        # Calcular proporção de cada cliente no mês
        proporcoes = {}
        for cliente_nome, faturamento in faturamento_por_cliente.items():
            proporcoes[cliente_nome] = faturamento / faturamento_total_mes

        # Filtrar despesas SEM cliente_id no mês específico (usa competencia contábil com fallback)
        despesas_sem_cliente = []
        for transacao in transacoes_despesas:
            mes_contabil_desp = transacao.competencia_mes_contabil or transacao.competencia_mes
            if transacao.cliente_id is None and mes_contabil_desp == (mes + 1):
                despesas_sem_cliente.append(transacao)

        if not despesas_sem_cliente:
            continue

        print(f"  📊 Mês {mes + 1}: {len(despesas_sem_cliente)} despesas sem cliente, faturamento total R$ {faturamento_total_mes:,.2f}")

        # Distribuir cada despesa proporcionalmente entre os clientes
        for transacao in despesas_sem_cliente:
            valor_despesa = abs(float(transacao.valor))

            # Obter nome da categoria para detalhamento
            categoria_nome = "Sem Categoria"
            if transacao.categoria_gerencial_id:
                try:
                    categoria_obj = db.query(CategoriaGerencial).filter(
                        CategoriaGerencial.id == transacao.categoria_gerencial_id
                    ).first()
                    if categoria_obj and categoria_obj.nome:
                        categoria_nome = str(categoria_obj.nome)
                except:
                    pass

            # Distribuir proporcionalmente entre os clientes
            for cliente_nome, proporcao in proporcoes.items():
                valor_rateado = valor_despesa * proporcao

                # Inicializar estrutura do cliente se não existir
                if cliente_nome not in rateio_por_cliente:
                    rateio_por_cliente[cliente_nome] = {
                        "rateio_meses": [0.0] * 12,
                        "rateio_total": 0.0,
                        "detalhes": {}
                    }

                # Acumular valor rateado
                rateio_por_cliente[cliente_nome]["rateio_meses"][mes] += valor_rateado
                rateio_por_cliente[cliente_nome]["rateio_total"] += valor_rateado

                # Detalhamento por categoria
                if categoria_nome not in rateio_por_cliente[cliente_nome]["detalhes"]:
                    rateio_por_cliente[cliente_nome]["detalhes"][categoria_nome] = {
                        "meses": [0.0] * 12,
                        "total": 0.0
                    }

                rateio_por_cliente[cliente_nome]["detalhes"][categoria_nome]["meses"][mes] += valor_rateado
                rateio_por_cliente[cliente_nome]["detalhes"][categoria_nome]["total"] += valor_rateado

                print(f"    ✓ {cliente_nome}: R$ {valor_rateado:,.2f} ({proporcao*100:.1f}% de R$ {valor_despesa:,.2f} - {categoria_nome})")

    # Resumo do rateio
    total_rateado_geral = sum(dados["rateio_total"] for dados in rateio_por_cliente.values())
    print(f"\n💰 Rateio concluído: R$ {total_rateado_geral:,.2f} distribuídos entre {len(rateio_por_cliente)} clientes")

    return rateio_por_cliente


@router.get("/relatorios/pl-consolidado")
async def get_pl_consolidado(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None),
    mes_corte: Optional[int] = Query(None, description="Mês de corte (1-12) para dividir faturado vs projetado")
):
    """
    P&L Consolidado com cálculo automático de impostos sobre receitas
    Impostos calculados pela competência, independente da data de pagamento
    Integra dados de planejamento orçamentário (linhas_orcamentarias)
    Suporta mês de corte para dividir faturado (até o mês) vs projetado (após o mês)
    """
    try:
        print(f"🔍 P&L Consolidado - Iniciando para ano={ano}, empresa={empresa}")

        # Determinar empresa para buscar alíquotas
        # Se filtrado por empresa específica, usar alíquotas daquela empresa
        # Senão, usar alíquotas da empresa do usuário
        empresa_id_usuario = getattr(current_user, 'empresa_id', None)
        empresa_id_aliquotas = empresa if empresa else empresa_id_usuario

        # Buscar alíquotas cadastradas de impostos
        aliquotas = obter_aliquotas_impostos(db, empresa_id_aliquotas) if empresa_id_aliquotas else {
            "PIS": 0.65, "COFINS": 3.0, "ISS": 5.0, "IRPJ": 7.93, "CSLL": 2.88
        }
        print(f"💰 Alíquotas de impostos carregadas: PIS={aliquotas['PIS']}%, COFINS={aliquotas['COFINS']}%, ISS={aliquotas['ISS']}%, IRPJ={aliquotas['IRPJ']}%, CSLL={aliquotas['CSLL']}%")

        # Query base para transações do ano — usa competencia_contabil (com fallback para campo antigo)
        from sqlalchemy import or_ as sql_or, and_ as sql_and
        query = db.query(TransacaoFinanceira).filter(
            sql_or(
                TransacaoFinanceira.competencia_ano_contabil == ano,
                sql_and(
                    TransacaoFinanceira.competencia_ano_contabil.is_(None),
                    TransacaoFinanceira.competencia_ano == ano
                )
            )
        )

        # Aplicar filtros opcionais
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)
        if cliente:
            query = query.filter(TransacaoFinanceira.cliente_id == cliente)
        if projeto:
            query = query.filter(TransacaoFinanceira.projeto_id == projeto)
        if produto_servico:
            query = query.filter(TransacaoFinanceira.produto_servico_id == produto_servico)

        # INCLUIR apenas transações que entram no gerencial (filhos CONTAM, pais desmembrados NÃO)
        query = query.filter(TransacaoFinanceira.entra_no_gerencial == True)

        transacoes = query.all()
        print(f"📊 Total de transações encontradas: {len(transacoes)}")

        # Query para dados de planejamento orçamentário (linhas_orcamentarias)
        query_orcamento = db.query(LinhaOrcamentaria).join(
            PlanejamentoVersao,
            LinhaOrcamentaria.versao_id == PlanejamentoVersao.id
        ).filter(
            LinhaOrcamentaria.ano == ano,
            PlanejamentoVersao.is_ativo == True
        )

        # Aplicar mesmo filtro de empresa se fornecido
        if empresa:
            query_orcamento = query_orcamento.filter(LinhaOrcamentaria.empresa_id == empresa)
        if cliente:
            query_orcamento = query_orcamento.filter(LinhaOrcamentaria.cliente_id == cliente)
        if projeto:
            query_orcamento = query_orcamento.filter(LinhaOrcamentaria.projeto_id == projeto)
        if produto_servico:
            query_orcamento = query_orcamento.filter(LinhaOrcamentaria.produto_servico_id == produto_servico)

        linhas_orcamento = query_orcamento.all()
        print(f"📋 Total de linhas orçamentárias encontradas: {len(linhas_orcamento)}")

        # Estruturas para armazenar dados
        receitas_mes = [0.0] * 12
        despesas_mes = [0.0] * 12
        receitas_dict = {}  # Por cliente
        despesas_dict = {}  # Por categoria contábil
        impostos_mes = {"pis": [0.0]*12, "cofins": [0.0]*12, "iss": [0.0]*12, "irpj": [0.0]*12, "csll": [0.0]*12}

        # Dicionário para armazenar títulos individuais de cada tipo de imposto
        impostos_titulos = {
            "pis": {},
            "cofins": {},
            "iss": {},
            "irpj": {},
            "csll": {}
        }

        # Cache de alíquotas por empresa para visão consolidada
        cache_aliquotas = {}

        # Processar transações
        for transacao in transacoes:
            try:
                mes_ref = transacao.competencia_mes_contabil or transacao.competencia_mes
                if not mes_ref or transacao.valor is None:
                    continue

                mes = int(mes_ref) - 1
                if mes < 0 or mes > 11:
                    continue

                # Se houver mês de corte, processar apenas transações ATÉ o mês de corte
                # Transações APÓS o corte serão substituídas por dados de planejamento
                # mes é 0-indexed (0-11), mes_corte é 1-indexed (1-12)
                if mes_corte and (mes + 1) > mes_corte:
                    continue

                valor = float(transacao.valor)

                if transacao.tipo == 'receita' and valor > 0:
                    # Receitas: Hierarquia Cliente > Projeto > Produto/Serviço
                    cliente_nome = "Sem Cliente"
                    if transacao.cliente_id:
                        try:
                            cliente_obj = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                            if cliente_obj and cliente_obj.nome:
                                cliente_nome = str(cliente_obj.nome)
                        except:
                            pass

                    projeto_nome = "Sem Projeto"
                    if transacao.projeto_id:
                        try:
                            projeto_obj = db.query(Projeto).filter(Projeto.id == transacao.projeto_id).first()
                            if projeto_obj and projeto_obj.nome:
                                projeto_nome = str(projeto_obj.nome)
                        except:
                            pass

                    produto_nome = "Sem Produto/Serviço"
                    if transacao.produto_servico_id:
                        try:
                            from app.models.auxiliares import ProdutoServico
                            produto_obj = db.query(ProdutoServico).filter(ProdutoServico.id == transacao.produto_servico_id).first()
                            if produto_obj and produto_obj.nome:
                                produto_nome = str(produto_obj.nome)
                        except:
                            pass

                    # Criar hierarquia Cliente > Projeto > Produto
                    if cliente_nome not in receitas_dict:
                        receitas_dict[cliente_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "projetos": {}
                        }

                    if projeto_nome not in receitas_dict[cliente_nome]["projetos"]:
                        receitas_dict[cliente_nome]["projetos"][projeto_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "produtos": {}
                        }

                    if produto_nome not in receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"]:
                        receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0
                        }

                    # Acumular valores em todos os níveis
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["realizado_meses"][mes] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["realizado_total"] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["realizado_meses"][mes] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["realizado_total"] += valor
                    receitas_dict[cliente_nome]["realizado_meses"][mes] += valor
                    receitas_dict[cliente_nome]["realizado_total"] += valor
                    receitas_mes[mes] += valor

                    # Calcular impostos por transação quando não há filtro de empresa
                    if not empresa and transacao.empresa_id:
                        # Buscar alíquotas da empresa da transação (com cache)
                        if transacao.empresa_id not in cache_aliquotas:
                            cache_aliquotas[transacao.empresa_id] = obter_aliquotas_impostos(db, transacao.empresa_id)

                        aliq_transacao = cache_aliquotas[transacao.empresa_id]
                        impostos_transacao = calcular_impostos_sobre_receita(valor, aliq_transacao)

                        impostos_mes["pis"][mes] += impostos_transacao["pis"]
                        impostos_mes["cofins"][mes] += impostos_transacao["cofins"]
                        impostos_mes["iss"][mes] += impostos_transacao["iss"]
                        impostos_mes["irpj"][mes] += impostos_transacao["irpj"]
                        impostos_mes["csll"][mes] += impostos_transacao["csll"]

                        # Salvar títulos individuais para drill-down de impostos
                        titulo_imposto = transacao.titulo_breve or transacao.nome or "Sem Título"

                        for tipo_imposto in ["pis", "cofins", "iss", "irpj", "csll"]:
                            if titulo_imposto not in impostos_titulos[tipo_imposto]:
                                impostos_titulos[tipo_imposto][titulo_imposto] = {
                                    "realizado_meses": [0.0] * 12,
                                    "previsto_meses": [0.0] * 12,
                                    "realizado_total": 0.0,
                                    "previsto_total": 0.0
                                }

                            impostos_titulos[tipo_imposto][titulo_imposto]["realizado_meses"][mes] += impostos_transacao[tipo_imposto]
                            impostos_titulos[tipo_imposto][titulo_imposto]["realizado_total"] += impostos_transacao[tipo_imposto]

                elif transacao.tipo == 'despesa':
                    # Despesas: Hierarquia Categoria Gerencial > Subcategoria Gerencial > Título
                    categoria_nome = "Sem Categoria Gerencial"
                    if transacao.categoria_gerencial_id:
                        try:
                            categoria_obj = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == transacao.categoria_gerencial_id
                            ).first()
                            if categoria_obj and categoria_obj.nome:
                                categoria_nome = str(categoria_obj.nome)
                        except:
                            pass

                    subcategoria_nome = "Sem Subcategoria"
                    if transacao.subcategoria_gerencial_id:
                        try:
                            subcategoria_obj = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == transacao.subcategoria_gerencial_id,
                                CategoriaGerencial.pai_id.isnot(None)
                            ).first()
                            if subcategoria_obj and subcategoria_obj.nome:
                                subcategoria_nome = str(subcategoria_obj.nome)
                        except Exception as e:
                            print(f"Erro ao buscar subcategoria gerencial ID {transacao.subcategoria_gerencial_id}: {e}")
                            pass

                    titulo_breve = transacao.titulo_breve or transacao.nome or "Sem Título"

                    # Criar hierarquia Categoria > Subcategoria > Título
                    if categoria_nome not in despesas_dict:
                        despesas_dict[categoria_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "subcategorias": {}
                        }

                    if subcategoria_nome not in despesas_dict[categoria_nome]["subcategorias"]:
                        despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "titulos": {}
                        }

                    if titulo_breve not in despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"]:
                        despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo_breve] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0
                        }

                    valor_abs = abs(valor)

                    # Acumular valores em todos os níveis
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo_breve]["realizado_meses"][mes] += valor_abs
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo_breve]["realizado_total"] += valor_abs
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["realizado_meses"][mes] += valor_abs
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["realizado_total"] += valor_abs
                    despesas_dict[categoria_nome]["realizado_meses"][mes] += valor_abs
                    despesas_dict[categoria_nome]["realizado_total"] += valor_abs
                    despesas_mes[mes] += valor_abs

            except Exception as e:
                print(f"⚠️ Erro ao processar transação ID {transacao.id}: {e}")
                continue

        # ========================================
        # PROCESSAR DADOS DE PLANEJAMENTO ORÇAMENTÁRIO
        # ========================================
        print(f"📋 Processando {len(linhas_orcamento)} linhas orçamentárias...")
        receitas_previsto_mes = [0.0] * 12
        despesas_previsto_mes = [0.0] * 12

        for linha in linhas_orcamento:
            try:
                if linha.mes is None or linha.mes < 1 or linha.mes > 12:
                    continue

                mes_idx = linha.mes - 1
                valor = float(linha.valor_previsto) if linha.valor_previsto else 0.0

                if valor == 0:
                    continue

                # Se houver mês de corte, só usar planejamento APÓS o mês de corte
                if mes_corte and linha.mes <= mes_corte:
                    continue

                # RECEITAS de planejamento
                if linha.categoria == 'receita':
                    # Hierarquia Cliente > Projeto > Produto/Serviço
                    cliente_nome = "Sem Cliente"
                    if linha.cliente_id:
                        try:
                            cliente_obj = db.query(Cliente).filter(Cliente.id == linha.cliente_id).first()
                            if cliente_obj and cliente_obj.nome:
                                cliente_nome = str(cliente_obj.nome)
                        except:
                            pass

                    projeto_nome = "Sem Projeto"
                    if linha.projeto_id:
                        try:
                            projeto_obj = db.query(Projeto).filter(Projeto.id == linha.projeto_id).first()
                            if projeto_obj and projeto_obj.nome:
                                projeto_nome = str(projeto_obj.nome)
                        except:
                            pass

                    produto_nome = linha.descricao or "Sem Produto/Serviço"
                    if linha.produto_servico_id:
                        try:
                            from app.models.auxiliares import ProdutoServico
                            produto_obj = db.query(ProdutoServico).filter(ProdutoServico.id == linha.produto_servico_id).first()
                            if produto_obj and produto_obj.nome:
                                produto_nome = str(produto_obj.nome)
                        except:
                            pass

                    # Criar estrutura se não existir
                    if cliente_nome not in receitas_dict:
                        receitas_dict[cliente_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "projetos": {}
                        }

                    if projeto_nome not in receitas_dict[cliente_nome]["projetos"]:
                        receitas_dict[cliente_nome]["projetos"][projeto_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "produtos": {}
                        }

                    if produto_nome not in receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"]:
                        receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0
                        }

                    # Acumular valores previstos
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["previsto_meses"][mes_idx] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["produtos"][produto_nome]["previsto_total"] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["previsto_meses"][mes_idx] += valor
                    receitas_dict[cliente_nome]["projetos"][projeto_nome]["previsto_total"] += valor
                    receitas_dict[cliente_nome]["previsto_meses"][mes_idx] += valor
                    receitas_dict[cliente_nome]["previsto_total"] += valor
                    receitas_previsto_mes[mes_idx] += valor

                # DESPESAS de planejamento
                elif linha.categoria == 'despesa':
                    # Usar categoria gerencial se disponível, senão usar descrição
                    categoria_nome = "Sem Categoria Gerencial"

                    if linha.categoria_gerencial_id:
                        try:
                            cat_obj = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == linha.categoria_gerencial_id
                            ).first()
                            if cat_obj and cat_obj.nome:
                                categoria_nome = str(cat_obj.nome)
                        except:
                            pass

                    # Buscar subcategoria gerencial
                    subcategoria_nome = "Sem Subcategoria"
                    if linha.subcategoria_gerencial_id:
                        try:
                            subcat_obj = db.query(CategoriaGerencial).filter(
                                CategoriaGerencial.id == linha.subcategoria_gerencial_id,
                                CategoriaGerencial.pai_id.isnot(None)
                            ).first()
                            if subcat_obj and subcat_obj.nome:
                                subcategoria_nome = str(subcat_obj.nome)
                        except:
                            pass

                    # Criar estrutura se não existir
                    if categoria_nome not in despesas_dict:
                        despesas_dict[categoria_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "subcategorias": {}
                        }

                    # Usar subcategoria gerencial da linha orçamentária
                    if subcategoria_nome not in despesas_dict[categoria_nome]["subcategorias"]:
                        despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0,
                            "titulos": {}
                        }

                    # Usar descrição como título
                    titulo = linha.descricao or "Sem Título"
                    if titulo not in despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"]:
                        despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo] = {
                            "realizado_meses": [0.0] * 12,
                            "previsto_meses": [0.0] * 12,
                            "realizado_total": 0.0,
                            "previsto_total": 0.0
                        }

                    # Acumular valores previstos
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo]["previsto_meses"][mes_idx] += valor
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["titulos"][titulo]["previsto_total"] += valor
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["previsto_meses"][mes_idx] += valor
                    despesas_dict[categoria_nome]["subcategorias"][subcategoria_nome]["previsto_total"] += valor
                    despesas_dict[categoria_nome]["previsto_meses"][mes_idx] += valor
                    despesas_dict[categoria_nome]["previsto_total"] += valor
                    despesas_previsto_mes[mes_idx] += valor

            except Exception as e:
                print(f"⚠️ Erro ao processar linha orçamentária ID {linha.id}: {e}")
                continue

        print(f"✅ Planejamento processado: Receitas previstas=R$ {sum(receitas_previsto_mes):,.2f}, Despesas previstas=R$ {sum(despesas_previsto_mes):,.2f}")

        # ========================================
        # CALCULAR RATEIO AUTOMÁTICO DE DESPESAS SEM CLIENTE
        # ========================================
        # Filtrar apenas despesas SEM cliente_id para rateio proporcional
        transacoes_despesas_sem_cliente = [
            t for t in transacoes 
            if t.tipo == 'despesa' and t.cliente_id is None
        ]

        # Chamar função de rateio se houver despesas sem cliente E houver receitas
        rateio_por_cliente = {}
        if transacoes_despesas_sem_cliente and receitas_dict:
            print(f"\n🔄 Rateio: {len(transacoes_despesas_sem_cliente)} despesas sem cliente serão rateadas")
            rateio_por_cliente = calcular_rateio_despesas(
                db=db,
                transacoes_despesas=transacoes_despesas_sem_cliente,
                receitas_por_cliente_mes=receitas_dict
            )
        else:
            if not transacoes_despesas_sem_cliente:
                print(f"ℹ️ Rateio: Não há despesas sem cliente para ratear")
            if not receitas_dict:
                print(f"ℹ️ Rateio: Não há receitas para calcular proporções")

        # Calcular ou usar impostos já calculados
        if empresa:
            # Se filtrado por empresa, recalcular com alíquota da empresa
            impostos_calculados = calcular_impostos_mensais(receitas_mes, aliquotas)
            print(f"📈 Total de impostos calculados: R$ {impostos_calculados['total_impostos']:,.2f}")

            # Popular títulos individuais de impostos para drill-down (quando filtrado por empresa)
            for transacao in transacoes:
                try:
                    mes_ref_imp = transacao.competencia_mes_contabil or transacao.competencia_mes
                    if transacao.tipo != 'receita' or not mes_ref_imp or transacao.valor is None:
                        continue

                    valor = float(transacao.valor)
                    if valor <= 0:
                        continue

                    mes = int(mes_ref_imp) - 1
                    if mes < 0 or mes > 11:
                        continue

                    # Se houver mês de corte, processar apenas transações ATÉ o mês de corte
                    # mes é 0-indexed (0-11), mes_corte é 1-indexed (1-12)
                    if mes_corte and (mes + 1) > mes_corte:
                        continue

                    # Calcular impostos desta transação com alíquotas da empresa filtrada
                    impostos_transacao = calcular_impostos_sobre_receita(valor, aliquotas)
                    titulo_imposto = transacao.titulo_breve or transacao.nome or "Sem Título"

                    for tipo_imposto in ["pis", "cofins", "iss", "irpj", "csll"]:
                        if titulo_imposto not in impostos_titulos[tipo_imposto]:
                            impostos_titulos[tipo_imposto][titulo_imposto] = {
                                "realizado_meses": [0.0] * 12,
                                "previsto_meses": [0.0] * 12,
                                "realizado_total": 0.0,
                                "previsto_total": 0.0
                            }

                        impostos_titulos[tipo_imposto][titulo_imposto]["realizado_meses"][mes] += impostos_transacao[tipo_imposto]
                        impostos_titulos[tipo_imposto][titulo_imposto]["realizado_total"] += impostos_transacao[tipo_imposto]
                except Exception as e:
                    print(f"⚠️ Erro ao processar título de imposto para transação ID {transacao.id}: {e}")
                    continue
        else:
            # Usar impostos já calculados por transação (consolidado multi-empresa)
            impostos_calculados = {
                "pis_mes": impostos_mes["pis"],
                "cofins_mes": impostos_mes["cofins"],
                "iss_mes": impostos_mes["iss"],
                "irpj_mes": impostos_mes["irpj"],
                "csll_mes": impostos_mes["csll"],
                "pis_total": sum(impostos_mes["pis"]),
                "cofins_total": sum(impostos_mes["cofins"]),
                "iss_total": sum(impostos_mes["iss"]),
                "irpj_total": sum(impostos_mes["irpj"]),
                "csll_total": sum(impostos_mes["csll"]),
                "total_impostos_mes": [impostos_mes["pis"][i] + impostos_mes["cofins"][i] + impostos_mes["iss"][i] + impostos_mes["irpj"][i] + impostos_mes["csll"][i] for i in range(12)],
                "total_impostos": sum(impostos_mes["pis"]) + sum(impostos_mes["cofins"]) + sum(impostos_mes["iss"]) + sum(impostos_mes["irpj"]) + sum(impostos_mes["csll"])
            }
            print(f"📈 Total de impostos consolidados (multi-empresa): R$ {impostos_calculados['total_impostos']:,.2f}")

        # ========================================
        # CALCULAR IMPOSTOS SOBRE RECEITAS PREVISTAS (PLANEJAMENTO)
        # ========================================
        impostos_previstos_mes = {"pis": [0.0]*12, "cofins": [0.0]*12, "iss": [0.0]*12, "irpj": [0.0]*12, "csll": [0.0]*12}

        # Calcular impostos sobre receitas previstas
        impostos_calculados_previstos = calcular_impostos_mensais(receitas_previsto_mes, aliquotas)

        # Adicionar aos totais
        impostos_previstos_mes["pis"] = impostos_calculados_previstos["pis_mes"]
        impostos_previstos_mes["cofins"] = impostos_calculados_previstos["cofins_mes"]
        impostos_previstos_mes["iss"] = impostos_calculados_previstos["iss_mes"]
        impostos_previstos_mes["irpj"] = impostos_calculados_previstos["irpj_mes"]
        impostos_previstos_mes["csll"] = impostos_calculados_previstos["csll_mes"]

        print(f"💰 Impostos sobre receitas previstas: R$ {impostos_calculados_previstos['total_impostos']:,.2f}")

        # Estruturar impostos por tipo (para exibição detalhada)
        # Quando consolidado, não mostrar alíquota específica pois podem ser múltiplas
        if empresa and aliquotas:
            label_pis = f"PIS ({aliquotas['PIS']}%)"
            label_cofins = f"COFINS ({aliquotas['COFINS']}%)"
            label_iss = f"ISS ({aliquotas['ISS']}%)"
            label_irpj = f"IRPJ ({aliquotas['IRPJ']}%)"
            label_csll = f"CSLL ({aliquotas['CSLL']}%)"
        else:
            label_pis = "PIS (variável)"
            label_cofins = "COFINS (variável)"
            label_iss = "ISS (variável)"
            label_irpj = "IRPJ (variável)"
            label_csll = "CSLL (variável)"

        impostos_dict = {
            label_pis: {
                "realizado_meses": impostos_calculados["pis_mes"],
                "realizado_total": impostos_calculados["pis_total"],
                "previsto_meses": impostos_previstos_mes["pis"],
                "previsto_total": sum(impostos_previstos_mes["pis"]),
                "titulos": impostos_titulos["pis"]
            },
            label_cofins: {
                "realizado_meses": impostos_calculados["cofins_mes"],
                "realizado_total": impostos_calculados["cofins_total"],
                "previsto_meses": impostos_previstos_mes["cofins"],
                "previsto_total": sum(impostos_previstos_mes["cofins"]),
                "titulos": impostos_titulos["cofins"]
            },
            label_iss: {
                "realizado_meses": impostos_calculados["iss_mes"],
                "realizado_total": impostos_calculados["iss_total"],
                "previsto_meses": impostos_previstos_mes["iss"],
                "previsto_total": sum(impostos_previstos_mes["iss"]),
                "titulos": impostos_titulos["iss"]
            },
            label_irpj: {
                "realizado_meses": impostos_calculados["irpj_mes"],
                "realizado_total": impostos_calculados["irpj_total"],
                "previsto_meses": impostos_previstos_mes["irpj"],
                "previsto_total": sum(impostos_previstos_mes["irpj"]),
                "titulos": impostos_titulos["irpj"]
            },
            label_csll: {
                "realizado_meses": impostos_calculados["csll_mes"],
                "realizado_total": impostos_calculados["csll_total"],
                "previsto_meses": impostos_previstos_mes["csll"],
                "previsto_total": sum(impostos_previstos_mes["csll"]),
                "titulos": impostos_titulos["csll"]
            }
        }

        # Calcular totais REALIZADOS
        receitas_total = sum(receitas_mes)
        despesas_total_sem_impostos = sum(despesas_mes)
        impostos_total = impostos_calculados["total_impostos"]

        # Despesas totais = Despesas operacionais + Impostos
        despesas_total_com_impostos = despesas_total_sem_impostos + impostos_total

        # Receita líquida = Receita bruta - Impostos
        receita_liquida_mes = [receitas_mes[i] - impostos_calculados["total_impostos_mes"][i] for i in range(12)]
        receita_liquida_total = receitas_total - impostos_total

        # Resultado = Receita líquida - Despesas operacionais
        # OU equivalentemente: Resultado = Receitas - (Despesas + Impostos)
        resultado_mes = [receitas_mes[i] - despesas_mes[i] - impostos_calculados["total_impostos_mes"][i] for i in range(12)]
        resultado_total = receitas_total - despesas_total_com_impostos

        # Calcular totais PREVISTOS (de planejamento)
        receitas_previsto_total = sum(receitas_previsto_mes)
        despesas_previsto_total = sum(despesas_previsto_mes)
        impostos_previsto_total = impostos_calculados_previstos["total_impostos"]

        # Despesas totais previstas = Despesas operacionais + Impostos
        despesas_previsto_total_com_impostos = despesas_previsto_total + impostos_previsto_total

        # Receita líquida prevista = Receita bruta - Impostos
        receita_liquida_previsto_mes = [receitas_previsto_mes[i] - impostos_calculados_previstos["total_impostos_mes"][i] for i in range(12)]
        receita_liquida_previsto_total = receitas_previsto_total - impostos_previsto_total

        # Resultado previsto = Receita líquida - Despesas operacionais
        resultado_previsto_mes = [receitas_previsto_mes[i] - despesas_previsto_mes[i] - impostos_calculados_previstos["total_impostos_mes"][i] for i in range(12)]
        resultado_previsto_total = receitas_previsto_total - despesas_previsto_total_com_impostos

        # Montar resposta
        resultado = {
            "ano": ano,
            "empresa_filtro": empresa,
            "aliquotas": aliquotas,  # Retornar todas as alíquotas cadastradas (PIS, COFINS, ISS, IRPJ, CSLL)
            "filtros": {
                "cliente": cliente,
                "projeto": projeto,
                "produto_servico": produto_servico
            },
            "receitas": receitas_dict,
            "impostos": impostos_dict,
            "despesas": despesas_dict,
            "rateio": rateio_por_cliente,  # Despesas rateadas automaticamente por cliente baseado em faturamento
            "totais": {
                "receitas_realizado_mes": receitas_mes,
                "receitas_previsto_mes": receitas_previsto_mes,
                "receitas_realizado_total": receitas_total,
                "receitas_previsto_total": receitas_previsto_total,

                "impostos_realizado_mes": impostos_calculados["total_impostos_mes"],
                "impostos_previsto_mes": impostos_calculados_previstos["total_impostos_mes"],
                "impostos_realizado_total": impostos_total,
                "impostos_previsto_total": impostos_previsto_total,

                "receita_liquida_realizado_mes": receita_liquida_mes,
                "receita_liquida_previsto_mes": receita_liquida_previsto_mes,
                "receita_liquida_realizado_total": receita_liquida_total,
                "receita_liquida_previsto_total": receita_liquida_previsto_total,

                "despesas_realizado_mes": despesas_mes,
                "despesas_previsto_mes": despesas_previsto_mes,
                "despesas_realizado_total": despesas_total_sem_impostos,
                "despesas_previsto_total": despesas_previsto_total,

                "resultado_realizado_mes": resultado_mes,
                "resultado_previsto_mes": resultado_previsto_mes,
                "resultado_realizado_total": resultado_total,
                "resultado_previsto_total": resultado_previsto_total
            }
        }

        print(f"✅ P&L Consolidado gerado: Receitas=R$ {receitas_total:,.2f}, Impostos=R$ {impostos_total:,.2f}, Despesas=R$ {despesas_total_sem_impostos:,.2f}, Resultado=R$ {resultado_total:,.2f}")
        return resultado

    except Exception as e:
        print(f"❌ P&L Consolidado - Erro: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao gerar P&L Consolidado: {str(e)}")


@router.get("/relatorios/rateio-despesas")
async def get_rateio_despesas(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    mes_inicio: Optional[int] = Query(None, ge=1, le=12, description="Mês inicial do período (1-12)"),
    mes_fim: Optional[int] = Query(None, ge=1, le=12, description="Mês final do período (1-12)"),
    empresa: Optional[int] = Query(None, description="Filtrar por empresa"),
    cliente: Optional[int] = Query(None, description="Filtrar receitas por cliente específico"),
    tipo: Optional[str] = Query("casado", description="Tipo de dados: realizado, planejado ou casado")
):
    """
    Consulta o rateio automático de despesas por período

    Retorna despesas sem cliente distribuídas proporcionalmente baseado no faturamento mensal
    Suporta dados realizados (transações), planejados (orçamento) ou casados (ambos)
    """
    try:
        print(f"🔍 Rateio de Despesas - Ano={ano}, Meses={mes_inicio or 1} a {mes_fim or 12}, Empresa={empresa}, Tipo={tipo}")

        # Definir período
        mes_inicial = mes_inicio or 1
        mes_final = mes_fim or 12

        receitas_por_cliente = {}
        despesas_sem_cliente = []

        # === DADOS REALIZADOS (transações) ===
        from sqlalchemy import or_ as sql_or2, and_ as sql_and2
        if tipo in ["realizado", "casado"]:
            query_receitas = db.query(TransacaoFinanceira).filter(
                sql_or2(
                    TransacaoFinanceira.competencia_ano_contabil == ano,
                    sql_and2(TransacaoFinanceira.competencia_ano_contabil.is_(None), TransacaoFinanceira.competencia_ano == ano)
                ),
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.entra_no_gerencial == True
            )

            query_despesas = db.query(TransacaoFinanceira).filter(
                sql_or2(
                    TransacaoFinanceira.competencia_ano_contabil == ano,
                    sql_and2(TransacaoFinanceira.competencia_ano_contabil.is_(None), TransacaoFinanceira.competencia_ano == ano)
                ),
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.entra_no_gerencial == True
            )

            if empresa:
                query_receitas = query_receitas.filter(TransacaoFinanceira.empresa_id == empresa)
                query_despesas = query_despesas.filter(TransacaoFinanceira.empresa_id == empresa)

            if cliente:
                query_receitas = query_receitas.filter(TransacaoFinanceira.cliente_id == cliente)

            transacoes_receitas = query_receitas.all()
            transacoes_despesas = query_despesas.all()

            print(f"📊 Realizado - Receitas: {len(transacoes_receitas)} | Despesas: {len(transacoes_despesas)}")

            for transacao in transacoes_receitas:
                cliente_nome = "Sem Cliente"
                if transacao.cliente_id:
                    cliente_obj = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                    if cliente_obj:
                        cliente_nome = cliente_obj.nome

                if cliente_nome not in receitas_por_cliente:
                    receitas_por_cliente[cliente_nome] = {
                        "realizado_meses": [0.0] * 12,
                        "realizado_total": 0.0
                    }

                mes_idx = (transacao.competencia_mes_contabil or transacao.competencia_mes) - 1
                valor = abs(float(transacao.valor))

                receitas_por_cliente[cliente_nome]["realizado_meses"][mes_idx] += valor
                receitas_por_cliente[cliente_nome]["realizado_total"] += valor

            despesas_sem_cliente.extend(transacoes_despesas)

        # === DADOS PLANEJADOS (orçamento) ===
        if tipo in ["planejado", "casado"]:
            from app.models.planejamento import LinhaOrcamentaria, PlanejamentoVersao, StatusPlanejamentoEnum, CategoriaLinhaEnum

            query_linhas_receita = db.query(LinhaOrcamentaria).join(
                PlanejamentoVersao, LinhaOrcamentaria.versao_id == PlanejamentoVersao.id
            ).filter(
                LinhaOrcamentaria.ano == ano,
                LinhaOrcamentaria.categoria == CategoriaLinhaEnum.receita,
                PlanejamentoVersao.status == StatusPlanejamentoEnum.publicado
            )

            query_linhas_despesa = db.query(LinhaOrcamentaria).join(
                PlanejamentoVersao, LinhaOrcamentaria.versao_id == PlanejamentoVersao.id
            ).filter(
                LinhaOrcamentaria.ano == ano,
                LinhaOrcamentaria.categoria == CategoriaLinhaEnum.despesa,
                PlanejamentoVersao.status == StatusPlanejamentoEnum.publicado
            )

            if empresa:
                query_linhas_receita = query_linhas_receita.filter(LinhaOrcamentaria.empresa_id == empresa)
                query_linhas_despesa = query_linhas_despesa.filter(LinhaOrcamentaria.empresa_id == empresa)

            if cliente:
                query_linhas_receita = query_linhas_receita.filter(LinhaOrcamentaria.cliente_id == cliente)

            linhas_receita = query_linhas_receita.all()
            linhas_despesa = query_linhas_despesa.all()

            print(f"📊 Planejado - Receitas: {len(linhas_receita)} | Despesas: {len(linhas_despesa)}")

            for linha in linhas_receita:
                cliente_nome = "Sem Cliente"
                if linha.cliente_id:
                    cliente_obj = db.query(Cliente).filter(Cliente.id == linha.cliente_id).first()
                    if cliente_obj:
                        cliente_nome = cliente_obj.nome

                if cliente_nome not in receitas_por_cliente:
                    receitas_por_cliente[cliente_nome] = {
                        "realizado_meses": [0.0] * 12,
                        "realizado_total": 0.0
                    }

                mes_idx = linha.mes - 1
                valor = abs(float(linha.valor_previsto))

                receitas_por_cliente[cliente_nome]["realizado_meses"][mes_idx] += valor
                receitas_por_cliente[cliente_nome]["realizado_total"] += valor

            class LinhaComoDespesa:
                def __init__(self, linha):
                    self.competencia_mes = linha.mes
                    self.valor = float(linha.valor_previsto)
                    self.cliente_id = linha.cliente_id
                    self.categoria_gerencial_id = linha.categoria_gerencial_id

            despesas_planejadas_sem_cliente = [l for l in linhas_despesa if l.cliente_id is None]
            despesas_sem_cliente.extend([LinhaComoDespesa(l) for l in despesas_planejadas_sem_cliente])

        print(f"📊 Total - Clientes com receita: {len(receitas_por_cliente)} | Despesas para rateio: {len(despesas_sem_cliente)}")

        # Calcular rateio de despesas
        rateio_result = calcular_rateio_despesas(db, despesas_sem_cliente, receitas_por_cliente)

        # Filtrar apenas os meses do período solicitado e calcular totais
        rateio_filtrado = {}
        for cliente_nome, dados in rateio_result.items():
            # Somar apenas os meses do período
            total_periodo = sum(dados["rateio_meses"][mes_inicial - 1:mes_final])

            # Filtrar detalhes por categoria também
            detalhes_filtrado = {}
            for categoria_nome, categoria_dados in dados["detalhes"].items():
                total_categoria_periodo = sum(categoria_dados["meses"][mes_inicial - 1:mes_final])

                if total_categoria_periodo > 0:
                    detalhes_filtrado[categoria_nome] = {
                        "meses": categoria_dados["meses"][mes_inicial - 1:mes_final],
                        "total": total_categoria_periodo
                    }

            if total_periodo > 0:
                rateio_filtrado[cliente_nome] = {
                    "rateio_meses": dados["rateio_meses"][mes_inicial - 1:mes_final],
                    "rateio_total": total_periodo,
                    "detalhes": detalhes_filtrado,
                    "faturamento_meses": receitas_por_cliente.get(cliente_nome, {}).get("realizado_meses", [0.0] * 12)[mes_inicial - 1:mes_final],
                    "faturamento_total": sum(receitas_por_cliente.get(cliente_nome, {}).get("realizado_meses", [0.0] * 12)[mes_inicial - 1:mes_final])
                }

        # Calcular total geral rateado no período
        total_rateado_periodo = sum(dados["rateio_total"] for dados in rateio_filtrado.values())

        # Calcular percentuais de cada cliente
        for cliente_nome in rateio_filtrado:
            percentual = (rateio_filtrado[cliente_nome]["rateio_total"] / total_rateado_periodo * 100) if total_rateado_periodo > 0 else 0
            rateio_filtrado[cliente_nome]["percentual"] = round(percentual, 2)

        # Ordenar por valor decrescente
        rateio_ordenado = dict(sorted(rateio_filtrado.items(), key=lambda x: x[1]["rateio_total"], reverse=True))

        print(f"✅ Rateio calculado: R$ {total_rateado_periodo:,.2f} distribuídos entre {len(rateio_ordenado)} clientes")

        return {
            "ano": ano,
            "mes_inicio": mes_inicial,
            "mes_fim": mes_final,
            "total_rateado": total_rateado_periodo,
            "quantidade_clientes": len(rateio_ordenado),
            "rateio_por_cliente": rateio_ordenado,
            "filtros": {
                "empresa": empresa,
                "cliente": cliente
            }
        }

    except Exception as e:
        print(f"❌ Erro ao calcular rateio: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao calcular rateio de despesas: {str(e)}")


@router.get("/relatorios/pl-consolidado/filtros")
async def get_pl_consolidado_filtros(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Buscar dados para filtros do P&L Consolidado
    """
    try:
        resultado = {
            "empresas": [],
            "clientes": [],
            "projetos": [],
            "produtos_servicos": []
        }

        # Buscar TODAS as empresas (sem restrições)
        try:
            from app.models.empresas import Empresa
            empresas = db.query(Empresa).filter(Empresa.ativo == True).all()
            resultado["empresas"] = [{"id": e.id, "nome": e.nome_fantasia or e.razao_social} for e in empresas]
        except Exception as e:
            print(f"❌ Erro ao carregar empresas: {e}")
            resultado["empresas"] = []

        # Buscar apenas clientes que possuem transações de receita no ano atual
        try:
            ano_atual = datetime.now().year
            clientes_com_valores = db.query(Cliente).join(TransacaoFinanceira).filter(
                Cliente.status == 'ativo',
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.competencia_ano == ano_atual,
                TransacaoFinanceira.valor > 0
            ).distinct().all()
            resultado["clientes"] = [{"id": c.id, "nome": c.nome} for c in clientes_com_valores]
        except Exception as e:
            print(f"❌ Erro ao carregar clientes com valores: {e}")
            resultado["clientes"] = []

        # Buscar TODOS os projetos (sem restrições por empresa)
        try:
            from app.models.auxiliares import Projeto
            projetos = db.query(Projeto).filter(Projeto.ativo == True).all()
            resultado["projetos"] = [{"id": p.id, "nome": p.nome} for p in projetos]
        except Exception as e:
            print(f"❌ Projetos não disponíveis: {e}")
            resultado["projetos"] = []

        # Buscar TODOS os produtos/serviços (sem restrições por empresa)
        try:
            from app.models.auxiliares import ProdutoServico
            produtos_servicos = db.query(ProdutoServico).filter(ProdutoServico.ativo == True).all()
            resultado["produtos_servicos"] = [{"id": ps.id, "nome": ps.nome, "tipo": ps.tipo} for ps in produtos_servicos]
        except Exception as e:
            print(f"❌ Erro ao carregar produtos/serviços: {e}")
            resultado["produtos_servicos"] = []

        print(f"✅ Filtros P&L carregados: empresas={len(resultado['empresas'])}, clientes={len(resultado['clientes'])}")
        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar filtros: {str(e)}")


@router.get("/relatorios/pl-consolidado/detalhes-categoria")
async def get_detalhes_categoria_pl_consolidado(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    categoria: str = Query(...),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None)
):
    """
    Buscar detalhes hierárquicos de uma categoria específica para drill down
    Estrutura: Categoria > Subcategoria > Breve Título
    """
    try:
        print(f"🔍 Detalhes categoria: '{categoria}' para ano={ano}, empresa={empresa}")

        # Query base para despesas da categoria específica
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'despesa',
            TransacaoFinanceira.competencia_ano == ano
        )

        # Aplicar filtros opcionais
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)
        if cliente:
            query = query.filter(TransacaoFinanceira.cliente_id == cliente)
        if projeto:
            query = query.filter(TransacaoFinanceira.projeto_id == projeto)
        if produto_servico:
            query = query.filter(TransacaoFinanceira.produto_servico_id == produto_servico)

        # Join com categoria contábil para filtrar pela categoria específica
        query = query.join(
            CategoriaContabil, 
            TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        ).filter(CategoriaContabil.nome == categoria)

        transacoes = query.all()
        print(f"📊 Encontradas {len(transacoes)} transações para categoria '{categoria}'")

        # Estruturar dados hierárquicos
        resultado = {
            "categoria": categoria,
            "subcategorias": {},
            "itens": {},
            "total_realizado": 0,
            "total_previsto": 0
        }

        # Processar transações
        for transacao in transacoes:
            try:
                if not transacao.competencia_mes or transacao.valor is None:
                    continue

                mes = int(transacao.competencia_mes) - 1
                if mes < 0 or mes > 11:
                    continue

                valor_abs = abs(float(transacao.valor))
                breve_titulo = transacao.titulo_breve or transacao.nome or "Sem título"

                # Buscar subcategoria se existir
                subcategoria_nome = "Sem Subcategoria"
                if transacao.subcategoria_contabil_id:
                    try:
                        subcategoria = db.query(CategoriaContabil).filter(
                            CategoriaContabil.id == transacao.subcategoria_contabil_id
                        ).first()
                        if subcategoria:
                            subcategoria_nome = subcategoria.nome
                    except Exception as e:
                        print(f"⚠️ Erro ao buscar subcategoria: {e}")

                # Inicializar subcategoria no dicionário se não existir
                if subcategoria_nome not in resultado["subcategorias"]:
                    resultado["subcategorias"][subcategoria_nome] = {
                        "nome": subcategoria_nome,
                        "itens": {},
                        "total_realizado": 0,
                        "total_previsto": 0
                    }

                # Inicializar item no dicionário se não existir
                if breve_titulo not in resultado["subcategorias"][subcategoria_nome]["itens"]:
                    resultado["subcategorias"][subcategoria_nome]["itens"][breve_titulo] = {
                        "titulo": breve_titulo,
                        "realizado_meses": [0.0] * 12,
                        "previsto_meses": [0.0] * 12,
                        "realizado_total": 0.0,
                        "previsto_total": 0.0
                    }

                # Acumular valores realizados
                resultado["subcategorias"][subcategoria_nome]["itens"][breve_titulo]["realizado_meses"][mes] += valor_abs
                resultado["subcategorias"][subcategoria_nome]["itens"][breve_titulo]["realizado_total"] += valor_abs
                resultado["subcategorias"][subcategoria_nome]["total_realizado"] += valor_abs
                resultado["total_realizado"] += valor_abs

            except Exception as e:
                print(f"⚠️ Erro ao processar transação ID {transacao.id}: {e}")
                continue

        print(f"✅ Detalhes da categoria '{categoria}' processados: {len(resultado['subcategorias'])} subcategorias")
        return resultado

    except Exception as e:
        print(f"❌ Erro ao buscar detalhes da categoria: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao buscar detalhes: {str(e)}")


@router.get("/relatorios/pl-consolidado/export-excel")
async def export_pl_consolidado_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None),
    mes_corte: Optional[int] = Query(None)
):
    """
    Exportar P&L Consolidado em formato Excel (.xlsx)
    Inclui todos os detalhes hierárquicos: Receitas (Cliente > Projeto > Produto),
    Impostos detalhados, Despesas (Categoria > Subcategoria > Título)
    """
    try:
        dados_pl = await get_pl_consolidado(
            current_user=current_user,
            db=db,
            ano=ano,
            empresa=empresa,
            cliente=cliente,
            projeto=projeto,
            produto_servico=produto_servico,
            mes_corte=mes_corte
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f"P&L Consolidado {ano}"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=10)
        group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        group_font = Font(bold=True, size=10)
        subgroup_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        subgroup_font = Font(bold=True, size=9)
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=10)
        imposto_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        imposto_font = Font(bold=True, size=9)

        ws.merge_cells('A1:N1')
        ws['A1'] = f"P&L CONSOLIDADO - {ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 2
        if empresa:
            ws[f'A{row}'] = f"Empresa: {empresa}"
            row += 1
        if mes_corte:
            ws[f'A{row}'] = f"Mês de Corte: {mes_corte}"
            row += 1

        row += 1
        meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

        ws[f'A{row}'] = 'CONCEITO'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        col = 2
        for mes in meses:
            ws.cell(row=row, column=col, value=mes)
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
            col += 1
        ws.cell(row=row, column=col, value='TOTAL')
        ws.cell(row=row, column=col).font = header_font
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1

        def write_row(ws, row_num, label, meses_vals, total_val, fill=None, font=None, indent=0):
            label_with_indent = "  " * indent + label
            ws[f'A{row_num}'] = label_with_indent
            if fill:
                ws[f'A{row_num}'].fill = fill
            if font:
                ws[f'A{row_num}'].font = font
            for i, val in enumerate(meses_vals):
                cell = ws.cell(row=row_num, column=i+2, value=val if val != 0 else None)
                cell.number_format = '#,##0'
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
            total_cell = ws.cell(row=row_num, column=14, value=total_val if total_val != 0 else None)
            total_cell.number_format = '#,##0'
            if fill:
                total_cell.fill = fill
            if font:
                total_cell.font = font
            return row_num + 1

        row = write_row(ws, row, 'RECEITAS BRUTAS', dados_pl['totais']['receitas_realizado_mes'], dados_pl['totais']['receitas_realizado_total'], section_fill, section_font)

        receitas = dados_pl.get('receitas', {})
        ordem_clientes = ['TIM', 'VIVO', 'CLARO']
        clientes_ordenados = []
        outros_clientes = []
        for cliente_pref in ordem_clientes:
            for cliente_nome in receitas.keys():
                if cliente_pref.upper() in cliente_nome.upper() and cliente_nome not in clientes_ordenados:
                    clientes_ordenados.append(cliente_nome)
        for cliente_nome in sorted(receitas.keys()):
            if cliente_nome not in clientes_ordenados:
                if 'OUTRO' in cliente_nome.upper() or 'SEM CLIENTE' in cliente_nome.upper():
                    outros_clientes.append(cliente_nome)
                else:
                    clientes_ordenados.append(cliente_nome)
        clientes_ordenados.extend(outros_clientes)

        for cliente_nome in clientes_ordenados:
            cliente_data = receitas[cliente_nome]
            row = write_row(ws, row, cliente_nome, cliente_data.get('realizado_meses', [0]*12), cliente_data.get('realizado_total', 0), group_fill, group_font, indent=1)
            
            projetos = cliente_data.get('projetos', {})
            for projeto_nome, projeto_data in projetos.items():
                row = write_row(ws, row, projeto_nome, projeto_data.get('realizado_meses', [0]*12), projeto_data.get('realizado_total', 0), subgroup_fill, subgroup_font, indent=2)
                
                produtos = projeto_data.get('produtos', {})
                for produto_nome, produto_data in produtos.items():
                    row = write_row(ws, row, produto_nome, produto_data.get('realizado_meses', [0]*12), produto_data.get('realizado_total', 0), indent=3)

        row += 1
        row = write_row(ws, row, 'IMPOSTOS SOBRE RECEITAS', dados_pl['totais']['impostos_realizado_mes'], dados_pl['totais']['impostos_realizado_total'], section_fill, section_font)

        impostos = dados_pl.get('impostos', {})
        impostos_ordem = ['pis', 'cofins', 'iss', 'irpj', 'csll']
        for tipo_imposto in impostos_ordem:
            if tipo_imposto in impostos:
                imposto_data = impostos[tipo_imposto]
                imposto_meses = imposto_data.get('realizado_meses', [0]*12)
                imposto_total = imposto_data.get('realizado_total', 0)
                row = write_row(ws, row, tipo_imposto.upper(), imposto_meses, imposto_total, imposto_fill, imposto_font, indent=1)
                
                titulos = imposto_data.get('titulos', {})
                for titulo_nome, titulo_data in titulos.items():
                    row = write_row(ws, row, titulo_nome, titulo_data.get('realizado_meses', [0]*12), titulo_data.get('realizado_total', 0), indent=2)

        row += 1
        row = write_row(ws, row, 'RECEITA LÍQUIDA', dados_pl['totais']['receita_liquida_realizado_mes'], dados_pl['totais']['receita_liquida_realizado_total'], total_fill, total_font)

        row += 1
        row = write_row(ws, row, 'DESPESAS OPERACIONAIS', dados_pl['totais']['despesas_realizado_mes'], dados_pl['totais']['despesas_realizado_total'], section_fill, section_font)

        despesas = dados_pl.get('despesas', {})
        for categoria_nome in sorted(despesas.keys()):
            categoria_data = despesas[categoria_nome]
            row = write_row(ws, row, categoria_nome, categoria_data.get('realizado_meses', [0]*12), categoria_data.get('realizado_total', 0), group_fill, group_font, indent=1)
            
            subcategorias = categoria_data.get('subcategorias', {})
            for subcategoria_nome, subcategoria_data in subcategorias.items():
                row = write_row(ws, row, subcategoria_nome, subcategoria_data.get('realizado_meses', [0]*12), subcategoria_data.get('realizado_total', 0), subgroup_fill, subgroup_font, indent=2)
                
                titulos = subcategoria_data.get('titulos', {})
                for titulo_nome, titulo_data in titulos.items():
                    row = write_row(ws, row, titulo_nome, titulo_data.get('realizado_meses', [0]*12), titulo_data.get('realizado_total', 0), indent=3)

        row += 1
        row = write_row(ws, row, 'RESULTADO LÍQUIDO', dados_pl['totais']['resultado_realizado_mes'], dados_pl['totais']['resultado_realizado_total'], total_fill, total_font)

        rateio = dados_pl.get('rateio', {})
        if rateio:
            row += 2
            row = write_row(ws, row, 'RATEIO DE DESPESAS (por cliente)', [0]*12, 0, section_fill, section_font)
            for cliente_nome, cliente_rateio in rateio.items():
                rateio_meses = cliente_rateio.get('rateio_meses', [0]*12)
                rateio_total = cliente_rateio.get('rateio_total', 0)
                row = write_row(ws, row, cliente_nome, rateio_meses, rateio_total, group_fill, group_font, indent=1)

        ws.column_dimensions['A'].width = 50
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"pl_consolidado_{ano}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar P&L Consolidado: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar P&L Consolidado: {str(e)}")


@router.get("/relatorios/pl-consolidado/detalhes-imposto")
async def get_detalhes_imposto_pl_consolidado(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    tipo_imposto: str = Query(...), # PIS, COFINS, ISS, IRPJ, CSLL
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None)
):
    """
    Buscar detalhes de impostos por tipo para drill down
    Estrutura: Imposto > Título Breve > Receita Associada
    """
    try:
        print(f"🔍 Detalhes imposto: '{tipo_imposto}' para ano={ano}, empresa={empresa}")

        # Query base para transações de receita
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'receita',
            TransacaoFinanceira.competencia_ano == ano,
            TransacaoFinanceira.valor > 0
        )

        # Aplicar filtros opcionais
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)
        if cliente:
            query = query.filter(TransacaoFinanceira.cliente_id == cliente)
        if projeto:
            query = query.filter(TransacaoFinanceira.projeto_id == projeto)
        if produto_servico:
            query = query.filter(TransacaoFinanceira.produto_servico_id == produto_servico)

        transacoes = query.all()
        print(f"📊 Encontradas {len(transacoes)} transações de receita para o período.")

        # Dicionário para armazenar os detalhes
        detalhes_imposto = {
            "tipo_imposto": tipo_imposto,
            "titulos": {},
            "total_imposto": 0.0,
            "total_receita_base": 0.0
        }

        # Cache de alíquotas por empresa
        cache_aliquotas = {}

        # Processar transações
        for transacao in transacoes:
            try:
                if not transacao.competencia_mes or transacao.valor is None:
                    continue

                mes = int(transacao.competencia_mes) - 1
                if mes < 0 or mes > 11:
                    continue

                valor_receita = float(transacao.valor)

                # Obter alíquotas da empresa da transação
                empresa_id = transacao.empresa_id
                if empresa_id not in cache_aliquotas:
                    cache_aliquotas[empresa_id] = obter_aliquotas_impostos(db, empresa_id)

                aliquotas = cache_aliquotas[empresa_id]

                # Calcular imposto específico
                valor_imposto_tipo = 0.0
                if tipo_imposto.upper() == "PIS":
                    valor_imposto_tipo = valor_receita * (aliquotas.get("PIS", 0) / 100.0)
                elif tipo_imposto.upper() == "COFINS":
                    valor_imposto_tipo = valor_receita * (aliquotas.get("COFINS", 0) / 100.0)
                elif tipo_imposto.upper() == "ISS":
                    valor_imposto_tipo = valor_receita * (aliquotas.get("ISS", 0) / 100.0)
                elif tipo_imposto.upper() == "IRPJ":
                    valor_imposto_tipo = valor_receita * (aliquotas.get("IRPJ", 0) / 100.0)
                elif tipo_imposto.upper() == "CSLL":
                    valor_imposto_tipo = valor_receita * (aliquotas.get("CSLL", 0) / 100.0)

                if valor_imposto_tipo > 0:
                    titulo_breve = transacao.titulo_breve or transacao.nome or "Sem Título"

                    # Inicializar título se não existir
                    if titulo_breve not in detalhes_imposto["titulos"]:
                        detalhes_imposto["titulos"][titulo_breve] = {
                            "meses": [0.0] * 12,
                            "total": 0.0,
                            "receita_base_total": 0.0
                        }

                    # Acumular valores
                    detalhes_imposto["titulos"][titulo_breve]["meses"][mes] += valor_imposto_tipo
                    detalhes_imposto["titulos"][breve_titulo]["total"] += valor_imposto_tipo
                    detalhes_imposto["titulos"][breve_titulo]["receita_base_total"] += valor_receita

                    detalhes_imposto["total_imposto"] += valor_imposto_tipo
                    detalhes_imposto["total_receita_base"] += valor_receita

            except Exception as e:
                print(f"⚠️ Erro ao processar transação ID {transacao.id} para imposto {tipo_imposto}: {e}")
                continue

        print(f"✅ Detalhes do imposto '{tipo_imposto}' processados. Total imposto: R$ {detalhes_imposto['total_imposto']:,.2f}")
        return detalhes_imposto

    except Exception as e:
        print(f"❌ Erro ao buscar detalhes do imposto: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao buscar detalhes do imposto: {str(e)}")


@router.get("/relatorios/cashflow-gerencial")
async def get_cashflow_gerencial(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None),
    base: Optional[str] = Query("competencia", description="Base de alocação: 'competencia', 'competencia_gerencial', 'lancamento' ou 'pagamento'")
):
    """
    Cashflow Gerencial - Hierarquia: Projeto > Cliente > Produto/Serviço
    Base configurável: competência contábil, competência gerencial, lançamento ou pagamento
    """
    try:
        print(f"🔍 Cash Flow: Carregando transações para empresa {empresa if empresa else 'TODAS'} (base: {base})")

        from sqlalchemy import or_, extract
        if base == "lancamento":
            query_base = db.query(TransacaoFinanceira).filter(
                extract('year', TransacaoFinanceira.data_lancamento) == ano
            )
        elif base == "pagamento":
            query_base = db.query(TransacaoFinanceira).filter(
                or_(
                    extract('year', TransacaoFinanceira.data_pagamento) == ano,
                    extract('year', TransacaoFinanceira.data_vencimento) == ano
                )
            )
        elif base == "competencia_gerencial":
            query_base = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.competencia_ano_gerencial == ano
            )
        else:
            query_base = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.competencia_ano_contabil == ano
            )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query_base = query_base.filter(TransacaoFinanceira.empresa_id == empresa)

        # Opcional: Filtro por cliente específico
        if cliente:
            print(f"🔍 Cash Flow: Aplicando filtro de cliente {cliente}")
            query_base = query_base.filter(TransacaoFinanceira.cliente_id == cliente)

        # Opcional: Filtro por projeto específico
        if projeto:
            print(f"🔍 Cash Flow: Aplicando filtro de projeto {projeto}")
            query_base = query_base.filter(TransacaoFinanceira.projeto_id == projeto)

        # Opcional: Filtro por produto/serviço específico
        if produto_servico:
            print(f"🔍 Cash Flow: Aplicando filtro de produto/serviço {produto_servico}")
            query_base = query_base.filter(TransacaoFinanceira.produto_servico_id == produto_servico)

        # Separar realizados (pagos) e projeções (pendentes)
        transacoes_realizadas = query_base.filter(
            TransacaoFinanceira.data_pagamento.isnot(None)
        ).all()
        transacoes_projecao = query_base.filter(
            TransacaoFinanceira.data_pagamento.is_(None),
            TransacaoFinanceira.data_vencimento.isnot(None)
        ).all()

        # Obter saldo inicial
        saldo_inicial = 0

        # Estrutura hierárquica
        resultado = {
            "ano": ano,
            "empresa_id": empresa,
            "tipo": "cashflow_gerencial",
            "saldo_inicial": saldo_inicial,
            "projetos": {},
            "categorias_despesas": {},
            "impostos": {},
            "totais": {
                "entrada_mes": [0] * 12,
                "saida_mes": [0] * 12,
                "fluxo_liquido_mes": [0] * 12,
                "saldo_acumulado_mes": [0] * 12,
                "entrada_total": 0,
                "saida_total": 0,
                "fluxo_liquido_total": 0
            }
        }

        # Cache de alíquotas e impostos calculados por transação (para visão consolidada)
        cache_aliquotas_cf = {}
        impostos_mes_cf = {"pis": [0.0]*12, "cofins": [0.0]*12, "iss": [0.0]*12, "irpj": [0.0]*12, "csll": [0.0]*12}

        def obter_mes_alocacao(transacao, tipo_dados):
            if base == "lancamento":
                if transacao.data_lancamento and transacao.data_lancamento.year == ano:
                    return transacao.data_lancamento.month
                return None
            elif base == "pagamento":
                data_ref = transacao.data_pagamento if tipo_dados == "realizado" else transacao.data_vencimento
                if data_ref and data_ref.year == ano:
                    return data_ref.month
                return None
            elif base == "competencia_gerencial":
                if transacao.competencia_ano_gerencial == ano and transacao.competencia_mes_gerencial:
                    return transacao.competencia_mes_gerencial
                data_ref = transacao.data_pagamento if tipo_dados == "realizado" else transacao.data_vencimento
                if data_ref and data_ref.year == ano:
                    return data_ref.month
                return None
            else:
                if transacao.competencia_ano_contabil == ano and transacao.competencia_mes_contabil:
                    return transacao.competencia_mes_contabil
                data_ref = transacao.data_pagamento if tipo_dados == "realizado" else transacao.data_vencimento
                if data_ref and data_ref.year == ano:
                    return data_ref.month
                return None

        def processar_transacoes(transacoes, tipo_dados="realizado"):
            print(f"🔍 Cash Flow: Processando {len(transacoes)} transações ({tipo_dados}, base={base})")
            for transacao in transacoes:
                try:
                    if transacao.valor is None:
                        continue

                    mes_alocacao = obter_mes_alocacao(transacao, tipo_dados)
                    if not mes_alocacao:
                        continue

                    mes = mes_alocacao - 1
                    if mes < 0 or mes > 11:
                        continue

                    valor = float(transacao.valor)

                    if transacao.tipo == 'receita':
                        print(f"  📊 Receita ID {transacao.id}: Cliente={transacao.cliente_id}, Projeto={transacao.projeto_id}, Valor={valor}")

                    # Hierarquia: Projeto - buscar nome real do projeto
                    projeto_nome = "Sem Projeto"
                    if transacao.projeto_id:
                        try:
                            # Assumindo que existe uma tabela de projetos
                            # Por enquanto, usar descrição da transação ou nome genérico
                            projeto_nome = transacao.nome or f"Projeto {transacao.projeto_id}"
                        except:
                            projeto_nome = f"Projeto {transacao.projeto_id}"

                    # Hierarquia: Cliente
                    cliente_nome = "Sem Cliente"
                    if transacao.cliente_id:
                        try:
                            cliente_obj = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                            if cliente_obj and cliente_obj.nome:
                                cliente_nome = str(cliente_obj.nome)
                        except:
                            pass

                    # Hierarquia: Produto/Serviço - buscar nome real
                    produto_nome = "Sem Produto/Serviço"
                    if transacao.produto_servico_id:
                        try:
                            from app.models.auxiliares import ProdutoServico
                            produto_obj = db.query(ProdutoServico).filter(
                                ProdutoServico.id == transacao.produto_servico_id
                            ).first()
                            if produto_obj and produto_obj.nome:
                                produto_nome = str(produto_obj.nome)
                            else:
                                produto_nome = f"Produto/Serviço {transacao.produto_servico_id}"
                        except:
                            produto_nome = f"Produto/Serviço {transacao.produto_servico_id}"

                    # Inicializar estruturas se necessário
                    if transacao.tipo == 'receita':
                        # RECEITAS - Hierarquia: Projeto > Cliente > Produto/Serviço
                        # Buscar nomes dos relacionamentos usando os IDs
                        projeto_nome = "Outros Projetos"
                        if transacao.projeto_id:
                            try:
                                projeto_obj = db.query(Projeto).filter(Projeto.id == transacao.projeto_id).first()
                                if projeto_obj and projeto_obj.nome:
                                    projeto_nome = str(projeto_obj.nome)
                                    print(f"  📊 Projeto encontrado: ID={transacao.projeto_id}, Nome={projeto_nome}")
                                else:
                                    print(f"  ⚠️ Projeto ID {transacao.projeto_id} não encontrado na tabela")
                            except Exception as e:
                                print(f"  ⚠️ Erro ao buscar projeto ID {transacao.projeto_id}: {e}")
                                projeto_nome = f"Projeto {transacao.projeto_id}"

                        # Cliente já foi processado acima
                        # produto_nome já foi processado acima

                        print(f"  📊 Estruturando receita: Projeto={projeto_nome}, Cliente={cliente_nome}, Produto={produto_nome}, Valor={valor}")

                        if projeto_nome not in resultado["projetos"]:
                            resultado["projetos"][projeto_nome] = {
                                "entrada_mes": [0] * 12,
                                "entrada_total": 0,
                                "clientes": {}
                            }

                        if cliente_nome not in resultado["projetos"][projeto_nome]["clientes"]:
                            resultado["projetos"][projeto_nome]["clientes"][cliente_nome] = {
                                "entrada_mes": [0] * 12,
                                "entrada_total": 0,
                                "produtos_servicos": {}
                            }

                        if produto_nome not in resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["produtos_servicos"]:
                            resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["produtos_servicos"][produto_nome] = {
                                "entrada_mes": [0] * 12,
                                "entrada_total": 0
                            }

                        # Somar receita em todos os níveis da hierarquia
                        resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["produtos_servicos"][produto_nome]["entrada_mes"][mes] += valor
                        resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["produtos_servicos"][produto_nome]["entrada_total"] += valor

                        resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["entrada_mes"][mes] += valor
                        resultado["projetos"][projeto_nome]["clientes"][cliente_nome]["entrada_total"] += valor

                        resultado["projetos"][projeto_nome]["entrada_mes"][mes] += valor
                        resultado["projetos"][projeto_nome]["entrada_total"] += valor

                        resultado["totais"]["entrada_mes"][mes] += valor
                        resultado["totais"]["entrada_total"] += valor

                        # Calcular impostos por transação quando não há filtro de empresa
                        if not empresa and transacao.empresa_id:
                            # Buscar alíquotas da empresa da transação (com cache)
                            if transacao.empresa_id not in cache_aliquotas_cf:
                                cache_aliquotas_cf[transacao.empresa_id] = obter_aliquotas_impostos(db, transacao.empresa_id)

                            aliq_transacao = cache_aliquotas_cf[transacao.empresa_id]
                            impostos_transacao = calcular_impostos_sobre_receita(valor, aliq_transacao)

                            impostos_mes_cf["pis"][mes] += impostos_transacao["pis"]
                            impostos_mes_cf["cofins"][mes] += impostos_transacao["cofins"]
                            impostos_mes_cf["iss"][mes] += impostos_transacao["iss"]
                            impostos_mes_cf["irpj"][mes] += impostos_transacao["irpj"]
                            impostos_mes_cf["csll"][mes] += impostos_transacao["csll"]

                    elif transacao.tipo == 'despesa':
                        # Nova estrutura hierárquica: Centro de Custo > Categoria Contábil > Breve Título
                        centro_custo_nome = "Sem Centro de Custo"
                        categoria_nome = "Sem Categoria Contábil"
                        breve_titulo = transacao.titulo_breve or transacao.nome or "Sem Título"

                        # Buscar centro de custo
                        if transacao.centro_custo_id:
                            try:
                                centro = db.query(CentroCusto).filter(
                                    CentroCusto.id == transacao.centro_custo_id
                                ).first()
                                if centro and centro.nome:
                                    centro_custo_nome = str(centro.nome)
                            except:
                                pass

                        # Buscar categoria contábil
                        if transacao.categoria_contabil_id:
                            try:
                                categoria = db.query(CategoriaContabil).filter(
                                    CategoriaContabil.id == transacao.categoria_contabil_id
                                ).first()
                                if categoria and categoria.nome:
                                    categoria_nome = str(categoria.nome)
                            except:
                                pass

                        # Criar estrutura hierárquica: Centro de Custo > Categoria > Título
                        if centro_custo_nome not in resultado["categorias_despesas"]:
                            resultado["categorias_despesas"][centro_custo_nome] = {
                                "saida_mes": [0] * 12,
                                "saida_total": 0,
                                "subcategorias": {}
                            }

                        if categoria_nome not in resultado["categorias_despesas"][centro_custo_nome]["subcategorias"]:
                            resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome] = {
                                "saida_mes": [0] * 12,
                                "saida_total": 0,
                                "itens": {}
                            }

                        if breve_titulo not in resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["itens"]:
                            resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["itens"][breve_titulo] = {
                                "saida_mes": [0] * 12,
                                "saida_total": 0
                            }

                        valor_abs = abs(valor)
                        # Somar em todos os níveis
                        resultado["categorias_despesas"][centro_custo_nome]["saida_mes"][mes] += valor_abs
                        resultado["categorias_despesas"][centro_custo_nome]["saida_total"] += valor_abs
                        resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["saida_mes"][mes] += valor_abs
                        resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["saida_total"] += valor_abs
                        resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["itens"][breve_titulo]["saida_mes"][mes] += valor_abs
                        resultado["categorias_despesas"][centro_custo_nome]["subcategorias"][categoria_nome]["itens"][breve_titulo]["saida_total"] += valor_abs
                        resultado["totais"]["saida_mes"][mes] += valor_abs
                        resultado["totais"]["saida_total"] += valor_abs

                except Exception as e:
                    continue

        # Processar dados realizados e projeções
        processar_transacoes(transacoes_realizadas, "realizado")
        processar_transacoes(transacoes_projecao, "projecao")

        # Calcular ou usar impostos já calculados
        if empresa:
            # Se filtrado por empresa, buscar alíquotas e recalcular
            aliquotas = obter_aliquotas_impostos(db, empresa)
            print(f"💰 Cash Flow: Alíquotas para empresa {empresa}: PIS={aliquotas['PIS']}%, COFINS={aliquotas['COFINS']}%, ISS={aliquotas['ISS']}%, IRPJ={aliquotas['IRPJ']}%, CSLL={aliquotas['CSLL']}%")
            impostos_calculados = calcular_impostos_mensais(resultado["totais"]["entrada_mes"], aliquotas)
            print(f"📈 Cash Flow: Total de impostos calculados: R$ {impostos_calculados['total_impostos']:,.2f}")
        else:
            # Usar impostos já calculados por transação (consolidado multi-empresa)
            impostos_calculados = {
                "pis_mes": impostos_mes_cf["pis"],
                "cofins_mes": impostos_mes_cf["cofins"],
                "iss_mes": impostos_mes_cf["iss"],
                "irpj_mes": impostos_mes_cf["irpj"],
                "csll_mes": impostos_mes_cf["csll"],
                "pis_total": sum(impostos_mes_cf["pis"]),
                "cofins_total": sum(impostos_mes_cf["cofins"]),
                "iss_total": sum(impostos_mes_cf["iss"]),
                "irpj_total": sum(impostos_mes_cf["irpj"]),
                "csll_total": sum(impostos_mes_cf["csll"]),
                "total_impostos_mes": [impostos_mes_cf["pis"][i] + impostos_mes_cf["cofins"][i] + impostos_mes_cf["iss"][i] + impostos_mes_cf["irpj"][i] + impostos_mes_cf["csll"][i] for i in range(12)],
                "total_impostos": sum(impostos_mes_cf["pis"]) + sum(impostos_mes_cf["cofins"]) + sum(impostos_mes_cf["iss"]) + sum(impostos_mes_cf["irpj"]) + sum(impostos_mes_cf["csll"])
            }
            print(f"📈 Cash Flow: Total de impostos consolidados (multi-empresa): R$ {impostos_calculados['total_impostos']:,.2f}")
            aliquotas = None  # Não há alíquota única no consolidado

        # Estruturar impostos detalhados por tipo
        if empresa and aliquotas:
            label_pis_cf = f"PIS ({aliquotas['PIS']}%)"
            label_cofins_cf = f"COFINS ({aliquotas['COFINS']}%)"
            label_iss_cf = f"ISS ({aliquotas['ISS']}%)"
            label_irpj_cf = f"IRPJ ({aliquotas['IRPJ']}%)"
            label_csll_cf = f"CSLL ({aliquotas['CSLL']}%)"
        else:
            label_pis_cf = "PIS (variável)"
            label_cofins_cf = "COFINS (variável)"
            label_iss_cf = "ISS (variável)"
            label_irpj_cf = "IRPJ (variável)"
            label_csll_cf = "CSLL (variável)"

        resultado["impostos"] = {
            label_pis_cf: {
                "saida_mes": impostos_calculados["pis_mes"],
                "saida_total": impostos_calculados["pis_total"]
            },
            label_cofins_cf: {
                "saida_mes": impostos_calculados["cofins_mes"],
                "saida_total": impostos_calculados["cofins_total"]
            },
            label_iss_cf: {
                "saida_mes": impostos_calculados["iss_mes"],
                "saida_total": impostos_calculados["iss_total"]
            },
            label_irpj_cf: {
                "saida_mes": impostos_calculados["irpj_mes"],
                "saida_total": impostos_calculados["irpj_total"]
            },
            label_csll_cf: {
                "saida_mes": impostos_calculados["csll_mes"],
                "saida_total": impostos_calculados["csll_total"]
            }
        }

        # Adicionar impostos às saídas totais
        for i in range(12):
            resultado["totais"]["saida_mes"][i] += impostos_calculados["total_impostos_mes"][i]
        resultado["totais"]["saida_total"] += impostos_calculados["total_impostos"]

        # Calcular fluxo líquido e saldo acumulado (agora incluindo impostos nas saídas)
        saldo_acumulado = saldo_inicial
        for i in range(12):
            fluxo_mes = resultado["totais"]["entrada_mes"][i] - resultado["totais"]["saida_mes"][i]
            resultado["totais"]["fluxo_liquido_mes"][i] = fluxo_mes
            saldo_acumulado += fluxo_mes
            resultado["totais"]["saldo_acumulado_mes"][i] = saldo_acumulado

        resultado["totais"]["fluxo_liquido_total"] = resultado["totais"]["entrada_total"] - resultado["totais"]["saida_total"]

        print(f"✅ Cash Flow: Gerado com Receitas=R$ {resultado['totais']['entrada_total']:,.2f}, Impostos=R$ {impostos_calculados['total_impostos']:,.2f}, Despesas+Impostos=R$ {resultado['totais']['saida_total']:,.2f}")
        return resultado

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao processar Cashflow Gerencial: {str(e)}")


def obterSaldoInicial(db: Session, empresa_id: int, ano_anterior: int) -> float:
    """
    Obter saldo inicial baseado no resultado do ano anterior
    """
    try:
        # Query para somar todas as transações pagas até o final do ano anterior
        # entra_no_gerencial == True exclui pais desmembrados (evita duplicidade com filhos)
        query = db.query(func.sum(TransacaoFinanceira.valor)).filter(
            TransacaoFinanceira.empresa_id == empresa_id,
            TransacaoFinanceira.data_pagamento.isnot(None),
            func.extract('year', TransacaoFinanceira.data_pagamento) <= ano_anterior,
            TransacaoFinanceira.exibir_no_cash_control == True,
            TransacaoFinanceira.entra_no_gerencial == True
        )

        resultado = query.scalar()
        return float(resultado) if resultado else 0.0

    except Exception as e:
        print(f"Erro ao calcular saldo inicial: {e}")
        return 0.0


@router.get("/relatorios/cash-control")
async def get_cash_control(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    mes: Optional[int] = Query(None),
    empresa: Optional[int] = Query(None)
):
    """
    Cash Control - Extrato bancário real
    Representa todas entradas e saídas da conta
    Diferente de contas a pagar/receber (previsões)
    """
    try:
        print(f"Cash Control: empresa={empresa if empresa else 'TODAS'}, ano={ano}, mes={mes}")

        # Query base para transações efetivamente pagas/recebidas
        # Inclui transações com data_pagamento preenchida OU status = 'pago'
        # entra_no_gerencial == True exclui pais desmembrados (filhos entram individualmente)
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.exibir_no_cash_control == True,
            TransacaoFinanceira.entra_no_gerencial == True,
            or_(
                TransacaoFinanceira.data_pagamento.isnot(None),
                TransacaoFinanceira.status == 'pago'
            )
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Filtros baseados na data de pagamento (ou data_lancamento se data_pagamento for NULL)
        # Usa COALESCE para pegar data_pagamento ou data_lancamento
        data_efetiva = func.coalesce(TransacaoFinanceira.data_pagamento, TransacaoFinanceira.data_lancamento)
        query = query.filter(func.extract('year', data_efetiva) == ano)

        if mes:
            query = query.filter(func.extract('month', data_efetiva) == mes)

        # Ordenar por data efetiva (pagamento ou transação)
        transacoes = query.order_by(data_efetiva).all()

        # Estrutura do cash control
        resultado = {
            "ano": ano,
            "mes": mes,
            "empresa_id": empresa,
            "tipo": "cash_control",
            "movimentacoes": [],
            "resumo": {
                "saldo_inicial": 0,
                "total_entradas": 0,
                "total_saidas": 0,
                "saldo_final": 0,
                "quantidade_movimentacoes": 0
            }
        }

        saldo_corrente = resultado["resumo"]["saldo_inicial"]

        # Processar cada movimentação com campos estendidos
        for transacao in transacoes:
            try:
                valor = float(transacao.valor or 0)
                tipo_movimentacao = "ENTRADA" if transacao.tipo == 'receita' and valor > 0 else "SAÍDA"
                valor_movimento = valor if tipo_movimentacao == "ENTRADA" else abs(valor)

                # Obter descrição da movimentação
                descricao = transacao.nome or transacao.descricao or "Movimentação sem descrição"

                # Cliente/Fornecedor
                contraparte = "Não informado"
                if transacao.cliente_id:
                    try:
                        cliente = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                        if cliente:
                            contraparte = cliente.nome
                    except:
                        pass
                elif transacao.fornecedor_id:
                    try:
                        fornecedor = db.query(Fornecedor).filter(Fornecedor.id == transacao.fornecedor_id).first()
                        if fornecedor:
                            contraparte = fornecedor.nome
                    except:
                        contraparte = f"Fornecedor ID {transacao.fornecedor_id}"

                # Obter categoria gerencial
                categoria = "Sem Categoria"
                if transacao.categoria_gerencial_id:
                    try:
                        cat = db.query(CategoriaGerencial).filter(CategoriaGerencial.id == transacao.categoria_gerencial_id).first()
                        if cat:
                            categoria = cat.nome
                    except:
                        pass

                # Obter centro de custo
                centro_custo = "Sem Centro"
                if transacao.centro_custo_id:
                    try:
                        cc = db.query(CentroCusto).filter(CentroCusto.id == transacao.centro_custo_id).first()
                        if cc:
                            centro_custo = cc.nome
                    except:
                        pass

                # Obter banco/conta (usando getattr pois campo pode não existir no modelo)
                banco = getattr(transacao, 'banco', None) or "Não informado"

                # Competência
                competencia = ""
                if transacao.competencia_mes and transacao.competencia_ano:
                    competencia = f"{transacao.competencia_mes:02d}/{transacao.competencia_ano}"

                # Atualizar saldo
                if tipo_movimentacao == "ENTRADA":
                    saldo_corrente += valor_movimento
                    resultado["resumo"]["total_entradas"] += valor_movimento
                else:
                    saldo_corrente -= valor_movimento
                    resultado["resumo"]["total_saidas"] += valor_movimento

                # Data efetiva: usa data_pagamento se disponível, senão data_lancamento
                data_efetiva_valor = transacao.data_pagamento or transacao.data_lancamento

                # Adicionar movimentação com campos estendidos
                resultado["movimentacoes"].append({
                    "data": data_efetiva_valor.isoformat() if data_efetiva_valor else None,
                    "tipo": tipo_movimentacao,
                    "descricao": descricao,
                    "contraparte": contraparte,
                    "valor": valor_movimento,
                    "saldo": saldo_corrente,
                    "forma_pagamento": transacao.forma_pgto or "",
                    "documento": transacao.numero_nota_fiscal or "",
                    "id_transacao": transacao.id,
                    "categoria": categoria,
                    "centro_custo": centro_custo,
                    "competencia": competencia,
                    "banco": banco,
                    "observacao": ""
                })

                resultado["resumo"]["quantidade_movimentacoes"] += 1

            except Exception as e:
                print(f"Erro ao processar transação ID {transacao.id}: {e}")
                continue

        resultado["resumo"]["saldo_final"] = saldo_corrente

        return resultado

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar Cash Control: {str(e)}")


@router.get("/relatorios/cash-control/export-excel")
async def export_cash_control_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    mes: Optional[int] = Query(None),
    empresa: Optional[int] = Query(None)
):
    """
    Exportar Cash Control para Excel
    """
    try:
        # Reutilizar a lógica do endpoint principal
        data = await get_cash_control(current_user, db, ano, mes, empresa)
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Control"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
        entrada_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        saida_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = 'R$ #,##0.00'
        
        # Cabeçalhos conforme modelo Excel
        headers = [
            "MONTH", "WEEK", "PERIOD", "CATEGORY", "CENTRO CUSTO", "COMPETÊNCIA",
            "PAYMENT DATE", "BANK", "FORMA PAGAMENTO", "DESCRIPTION", "DOCUMENT",
            "REMARK", "INFLOWS", "OUTFLOWS", "SALDO"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Dados
        meses_nomes = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']
        
        row = 2
        for mov in data["movimentacoes"]:
            # Calcular mês/semana da data de pagamento
            if mov["data"]:
                data_pgto = datetime.fromisoformat(mov["data"])
                mes_nome = f"{meses_nomes[data_pgto.month]}/{str(data_pgto.year)[2:]}"
                dia_ano = data_pgto.timetuple().tm_yday
                semana = f"Week {(dia_ano // 7) + 1}"
                
                # Período da semana
                dia_semana = data_pgto.weekday()
                inicio_semana = data_pgto - timedelta(days=dia_semana)
                fim_semana = inicio_semana + timedelta(days=6)
                periodo = f"{inicio_semana.strftime('%d/%m')} A {fim_semana.strftime('%d/%m')}"
                data_formatada = data_pgto.strftime('%d/%m/%Y')
            else:
                mes_nome = "-"
                semana = "-"
                periodo = "-"
                data_formatada = "-"
            
            valores = [
                mes_nome,
                semana,
                periodo,
                mov.get("categoria", "-"),
                mov.get("centro_custo", "-"),
                mov.get("competencia", "-"),
                data_formatada,
                mov.get("banco", "-"),
                mov.get("forma_pagamento", "-"),
                mov.get("descricao", "-"),
                mov.get("documento", "-"),
                mov.get("observacao", "-"),
                mov["valor"] if mov["tipo"] == "ENTRADA" else "",
                mov["valor"] if mov["tipo"] == "SAÍDA" else "",
                mov["saldo"]
            ]
            
            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = border
                
                # Aplicar cor de fundo para entradas/saídas
                if mov["tipo"] == "ENTRADA":
                    cell.fill = entrada_fill
                else:
                    cell.fill = saida_fill
                
                # Formato de moeda para colunas de valor
                if col in [13, 14, 15] and isinstance(valor, (int, float)) and valor != "":
                    cell.number_format = currency_format
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Linha de totais
        ws.cell(row=row, column=12, value="TOTAIS").font = Font(bold=True)
        ws.cell(row=row, column=13, value=data["resumo"]["total_entradas"]).number_format = currency_format
        ws.cell(row=row, column=13).font = Font(bold=True, color="059669")
        ws.cell(row=row, column=14, value=data["resumo"]["total_saidas"]).number_format = currency_format
        ws.cell(row=row, column=14).font = Font(bold=True, color="DC2626")
        ws.cell(row=row, column=15, value=data["resumo"]["saldo_final"]).number_format = currency_format
        ws.cell(row=row, column=15).font = Font(bold=True, color="1E40AF")
        
        # Ajustar largura das colunas
        larguras = [15, 10, 15, 20, 15, 12, 12, 15, 15, 35, 15, 25, 15, 15, 15]
        for i, largura in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = largura
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Salvar em buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nome do arquivo
        mes_nome_pt = ['', 'Janeiro', 'Fevereiro', 'Marco', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        nome_mes = mes_nome_pt[mes] if mes else "Ano"
        filename = f"CashControl_{nome_mes}_{ano}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Cash Control: {str(e)}")


@router.get("/relatorios/top-despesas")
async def get_top_despesas(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    empresa: Optional[int] = Query(None),
    limit: int = Query(10, ge=5, le=50)
):
    """
    Relatório Top Despesas por categoria
    """
    try:
        # Query base para despesas
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'despesa'
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Filtros
        if ano:
            query = query.filter(TransacaoFinanceira.competencia_ano == ano)
        if mes:
            query = query.filter(TransacaoFinanceira.competencia_mes == mes)

        # Join com categorias contábeis
        query = query.outerjoin(
            CategoriaContabil, 
            TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        )

        # Agrupar por categoria
        despesas_categoria = query.with_entities(
            CategoriaContabil.nome.label('categoria_nome'),
            func.sum(func.abs(TransacaoFinanceira.valor)).label('total_despesa'),
            func.count(TransacaoFinanceira.id).label('qtd_transacoes')
        ).group_by(CategoriaContabil.id, CategoriaContabil.nome).all()

        # Processar resultados
        resultado = []
        total_geral = 0

        for item in despesas_categoria:
            categoria_nome = item.categoria_nome or "Sem Categoria Contábil"
            valor = float(item.total_despesa or 0)
            total_geral += valor

            resultado.append({
                "categoria": categoria_nome,
                "valor": valor,
                "transacoes": item.qtd_transacoes,
                "percentual": 0  # Será calculado após ter o total
            })

        # Calcular percentuais e ordenar
        if total_geral > 0:
            for item in resultado:
                item["percentual"] = (item["valor"] / total_geral) * 100

        # Ordenar por valor decrescente
        resultado.sort(key=lambda x: x["valor"], reverse=True)
        resultado = resultado[:limit]

        return {
            "dados": resultado,
            "total_geral": total_geral,
            "filtros": {
                "ano": ano,
                "mes": mes,
                "empresa": empresa,
                "limit": limit
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório de top despesas: {str(e)}")


@router.get("/relatorios/contas-a-pagar")
async def get_contas_a_pagar(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    mes: Optional[int] = Query(None),
    ano: Optional[int] = Query(None),
    tipo_data: Optional[str] = Query("pagamento", description="'pagamento' | 'vencimento' | 'lancamento' | 'contabil' | 'gerencial'"),
    status: Optional[str] = Query(None),
    fornecedor: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    empresa: Optional[int] = Query(None)
):
    """
    Relatório de Contas a Pagar
    Inclui: transações normais e pagamentos ao fornecedor (tipo_filho='split' com nome 'Pagamento')
    Exclui: transações pai desmembradas e retenções de impostos (vão para relatório específico)
    """
    try:
        from sqlalchemy import and_, not_, or_, extract as sql_extract, case as sql_case

        # Query base para despesas:
        # - entra_no_gerencial == True exclui pais desmembrados E retenções de impostos
        # - Inclui pagamentos ao fornecedor (tipo_filho='split', nome 'Pagamento', entra_no_gerencial=True)
        # - Exclui retenções de impostos (tipo_filho='split', nome 'Retenção%', entra_no_gerencial=False)
        # - Exclui pais desmembrados (entra_no_gerencial=False setado pelo desmembramento)
        query = db.query(TransacaoFinanceira).filter(
            and_(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.valor != 0,
                TransacaoFinanceira.entra_no_gerencial == True,
            )
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Filtro de data unificado por tipo_data + mes + ano
        tipo = tipo_data or "pagamento"

        if tipo == "pagamento":
            # Data efetiva de pagamento: usa data_pagamento quando preenchida,
            # caso contrário usa data_vencimento para registros marcados como 'pago'
            # (espelha a lógica de exibição do frontend)
            data_pgto_efetiva = func.coalesce(
                TransacaoFinanceira.data_pagamento,
                sql_case(
                    (TransacaoFinanceira.status == 'pago', TransacaoFinanceira.data_vencimento),
                    else_=None
                )
            )
            if mes:
                query = query.filter(sql_extract('month', data_pgto_efetiva) == mes)
            if ano:
                query = query.filter(sql_extract('year', data_pgto_efetiva) == ano)
        elif tipo == "vencimento":
            if mes:
                query = query.filter(sql_extract('month', TransacaoFinanceira.data_vencimento) == mes)
            if ano:
                query = query.filter(sql_extract('year', TransacaoFinanceira.data_vencimento) == ano)
        elif tipo == "lancamento":
            if mes:
                query = query.filter(sql_extract('month', TransacaoFinanceira.data_lancamento) == mes)
            if ano:
                query = query.filter(sql_extract('year', TransacaoFinanceira.data_lancamento) == ano)
        elif tipo == "gerencial":
            if mes:
                query = query.filter(TransacaoFinanceira.competencia_mes_gerencial == mes)
            if ano:
                query = query.filter(TransacaoFinanceira.competencia_ano_gerencial == ano)
        else:
            # "contabil" — filtro com fallback: usa _contabil se preenchido, senão usa campo legado
            # Isso garante que filhos de desdobramentos sem _contabil próprio ainda apareçam
            if mes:
                query = query.filter(
                    func.coalesce(
                        TransacaoFinanceira.competencia_mes_contabil,
                        TransacaoFinanceira.competencia_mes
                    ) == mes
                )
            if ano:
                query = query.filter(
                    func.coalesce(
                        TransacaoFinanceira.competencia_ano_contabil,
                        TransacaoFinanceira.competencia_ano
                    ) == ano
                )

        # Filtro por status — verifica data_pagamento E campo status
        if status:
            if status == 'pago':
                query = query.filter(
                    or_(
                        TransacaoFinanceira.data_pagamento.isnot(None),
                        TransacaoFinanceira.status == 'pago'
                    )
                )
            elif status == 'pendente':
                query = query.filter(
                    and_(
                        TransacaoFinanceira.data_pagamento.is_(None),
                        or_(
                            TransacaoFinanceira.status.is_(None),
                            TransacaoFinanceira.status != 'pago'
                        )
                    )
                )

        # Buscar transações com joins (Fornecedor + CategoriaContabil + CentroCusto)
        query_with_joins = query.outerjoin(
            Fornecedor, TransacaoFinanceira.fornecedor_id == Fornecedor.id
        ).outerjoin(
            CategoriaContabil, TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        ).outerjoin(
            CentroCusto, TransacaoFinanceira.centro_custo_id == CentroCusto.id
        )

        # Filtro por fornecedor
        if fornecedor:
            query_with_joins = query_with_joins.filter(Fornecedor.nome.ilike(f'%{fornecedor}%'))

        # Filtro por descrição (incluindo título breve)
        if descricao:
            query_with_joins = query_with_joins.filter(
                or_(
                    TransacaoFinanceira.nome.ilike(f'%{descricao}%'),
                    TransacaoFinanceira.descricao.ilike(f'%{descricao}%'),
                    TransacaoFinanceira.titulo_breve.ilike(f'%{descricao}%')
                )
            )

        transacoes = query_with_joins.add_columns(
            Fornecedor.nome.label('fornecedor_nome'),
            CategoriaContabil.nome.label('categoria_contabil_nome'),
            CentroCusto.nome.label('centro_custo_nome')
        ).order_by(
            TransacaoFinanceira.data_pagamento.asc().nullslast(),
            TransacaoFinanceira.data_vencimento.asc().nullslast()
        ).all()

        # Buscar retenções de impostos para todas as transações em lote
        ids_transacoes = [item[0].id for item in transacoes]
        retencoes_map = {}  # {parent_id: {tipo: valor}}
        if ids_transacoes:
            splits_retencoes = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.parent_id.in_(ids_transacoes),
                TransacaoFinanceira.tipo_filho == 'split',
                TransacaoFinanceira.nome.like('Retenção%')
            ).all()
            for split in splits_retencoes:
                pid = split.parent_id
                if pid not in retencoes_map:
                    retencoes_map[pid] = {"inss": 0.0, "irrf": 0.0, "iss": 0.0, "csll_pis_cofins": 0.0}
                nome_upper = (split.nome or '').upper()
                val = abs(float(split.valor or 0))
                if 'INSS' in nome_upper:
                    retencoes_map[pid]["inss"] += val
                elif 'IRRF' in nome_upper or 'IRPJ' in nome_upper:
                    retencoes_map[pid]["irrf"] += val
                elif 'ISS' in nome_upper:
                    retencoes_map[pid]["iss"] += val
                elif 'CSLL' in nome_upper or 'PIS' in nome_upper or 'COFINS' in nome_upper:
                    retencoes_map[pid]["csll_pis_cofins"] += val

        # Processar dados
        contas = []
        total_valor = 0
        total_pendente = 0
        total_pago = 0

        for item in transacoes:
            transacao = item[0]
            fornecedor_nome = item[1] if item[1] else "Fornecedor não informado"
            categoria_contabil_nome = item[2] or None
            centro_custo_nome = item[3] or None

            valor = float(transacao.valor or 0)
            valor_abs = abs(valor)

            # Determinar status — verifica data_pagamento e também campo status do banco
            eh_pago = bool(transacao.data_pagamento) or (transacao.status and transacao.status.lower() == 'pago')
            if eh_pago:
                status_conta = "Pago"
                total_pago += valor_abs
            else:
                status_conta = "Pendente"
                total_pendente += valor_abs

            total_valor += valor_abs

            # Competência contábil e gerencial
            if transacao.competencia_mes_contabil and transacao.competencia_ano_contabil:
                comp_contabil = f"{transacao.competencia_mes_contabil:02d}/{transacao.competencia_ano_contabil}"
            elif transacao.competencia_mes and transacao.competencia_ano:
                comp_contabil = f"{transacao.competencia_mes:02d}/{transacao.competencia_ano}"
            else:
                comp_contabil = None

            if transacao.competencia_mes_gerencial and transacao.competencia_ano_gerencial:
                comp_gerencial = f"{transacao.competencia_mes_gerencial:02d}/{transacao.competencia_ano_gerencial}"
            else:
                comp_gerencial = comp_contabil

            # Retenções de impostos
            ret = retencoes_map.get(transacao.id, {})

            contas.append({
                "id": transacao.id,
                # Campos na nova ordem
                "data_pagamento": transacao.data_pagamento.isoformat() if transacao.data_pagamento else None,
                "data_emissao": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
                "fornecedor": fornecedor_nome,
                "numero_documento": transacao.numero_nota_fiscal,
                "conta_contabil": categoria_contabil_nome,
                "centro_custo": centro_custo_nome,
                "competencia_contabil": comp_contabil,
                "descricao": transacao.nome or transacao.descricao or "Sem descrição",
                "valor_bruto": valor_abs,
                "inss": ret.get("inss", 0.0),
                "irrf": ret.get("irrf", 0.0),
                "iss": ret.get("iss", 0.0),
                "csll_pis_cofins": ret.get("csll_pis_cofins", 0.0),
                "juros_multas": None,  # campo futuro no formulário
                "total_a_pagar": float(transacao.valor_pago) if transacao.valor_pago is not None else valor_abs,
                # Campos de suporte (mantidos para compatibilidade)
                "valor": valor_abs,
                "valor_pago": float(transacao.valor_pago) if transacao.valor_pago is not None else None,
                "status": status_conta,
                "data_vencimento": transacao.data_vencimento.isoformat() if transacao.data_vencimento else None,
                "competencia_gerencial": comp_gerencial,
                "forma_pagamento": transacao.forma_pgto,
            })

        return {
            "contas": contas,
            "resumo": {
                "total_contas": len(contas),
                "valor_total": total_valor,
                "valor_pendente": total_pendente,
                "valor_pago": total_pago
            },
            "filtros": {
                "ano": ano,
                "mes": mes,
                "status": status,
                "empresa": empresa
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório de contas a pagar: {str(e)}")


@router.get("/relatorios/contas-a-receber")
async def get_contas_a_receber(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    cliente: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    empresa: Optional[int] = Query(None),
    tipo_data: Optional[str] = Query("contabil")
):
    """
    Relatório de Contas a Receber
    tipo_data: 'contabil' (padrão) ou 'gerencial'
    """
    try:
        # Query base para receitas
        # entra_no_gerencial == True exclui pais desmembrados (filhos entram individualmente)
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'receita',
            TransacaoFinanceira.entra_no_gerencial == True
        )

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Filtros opcionais — usa competência contábil ou gerencial conforme tipo_data
        usar_gerencial = tipo_data == "gerencial"
        if ano:
            if usar_gerencial:
                query = query.filter(TransacaoFinanceira.competencia_ano_gerencial == ano)
            else:
                query = query.filter(TransacaoFinanceira.competencia_ano_contabil == ano)
        if mes:
            if usar_gerencial:
                query = query.filter(TransacaoFinanceira.competencia_mes_gerencial == mes)
            else:
                query = query.filter(TransacaoFinanceira.competencia_mes_contabil == mes)

        # Filtro por status
        if status:
            if status == 'recebido':
                query = query.filter(TransacaoFinanceira.data_pagamento.isnot(None))
            elif status == 'pendente':
                query = query.filter(TransacaoFinanceira.data_pagamento.is_(None))

        # Buscar transações com joins incluindo filtros de cliente
        query_with_joins = query.outerjoin(
            Cliente, TransacaoFinanceira.cliente_id == Cliente.id
        )

        # Filtro por cliente
        if cliente:
            query_with_joins = query_with_joins.filter(Cliente.nome.ilike(f'%{cliente}%'))

        # Filtro por descrição (incluindo título breve)
        if descricao:
            query_with_joins = query_with_joins.filter(
                or_(
                    TransacaoFinanceira.nome.ilike(f'%{descricao}%'),
                    TransacaoFinanceira.descricao.ilike(f'%{descricao}%'),
                    TransacaoFinanceira.titulo_breve.ilike(f'%{descricao}%')
                )
            )

        nf_col = func.nullif(TransacaoFinanceira.numero_nota_fiscal, '')
        transacoes = query_with_joins.add_columns(
            Cliente.nome.label('cliente_nome')
        ).order_by(
            nf_col.asc().nulls_last(),
            TransacaoFinanceira.data_vencimento.desc()
        ).all()

        # Processar dados
        contas = []
        total_valor = 0
        total_pendente = 0
        total_recebido = 0

        for item in transacoes:
            transacao = item[0]
            cliente_nome = item[1] if item[1] else "Cliente não informado"

            valor_bruto = float(transacao.valor or 0)

            # Buscar dados do cliente (CNPJ)
            cliente_cnpj = "Não informado"
            if transacao.cliente_id:
                try:
                    cliente_obj = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
                    if cliente_obj and cliente_obj.documento:
                        cliente_cnpj = str(cliente_obj.documento)
                except:
                    pass

            # Calcular impostos sobre receita usando alíquotas da empresa
            impostos_detalhados = {"pis": 0, "cofins": 0, "iss": 0, "irpj": 0, "csll": 0}
            aliquotas_usadas = {"pis": 0, "cofins": 0, "iss": 0, "irpj": 0, "csll": 0}
            valor_impostos_total = 0

            if valor_bruto > 0 and transacao.empresa_id:
                try:
                    aliquotas = obter_aliquotas_impostos(db, transacao.empresa_id)
                    impostos_calculados = calcular_impostos_sobre_receita(valor_bruto, aliquotas)
                    impostos_detalhados = {
                        "pis": impostos_calculados["pis"],
                        "cofins": impostos_calculados["cofins"], 
                        "iss": impostos_calculados["iss"],
                        "irpj": impostos_calculados["irpj"],
                        "csll": impostos_calculados["csll"]
                    }
                    aliquotas_usadas = {
                        "pis": aliquotas.get("PIS", 0),
                        "cofins": aliquotas.get("COFINS", 0),
                        "iss": aliquotas.get("ISS", 0),
                        "irpj": aliquotas.get("IRPJ", 0),
                        "csll": aliquotas.get("CSLL", 0)
                    }
                    valor_impostos_total = impostos_calculados["total"]
                except Exception as e:
                    print(f"Erro ao calcular impostos para transação {transacao.id}: {e}")
                    pass

            # Valor líquido = Valor bruto - Impostos
            valor_liquido = valor_bruto - valor_impostos_total

            # Determinar status
            if transacao.data_pagamento:
                status_conta = "Recebido"
                total_recebido += valor_bruto
            else:
                status_conta = "Pendente"
                total_pendente += valor_bruto

            total_valor += valor_bruto

            contas.append({
                "id": transacao.id,
                "cliente": cliente_nome,
                "cliente_cnpj": cliente_cnpj,
                "descricao": transacao.nome or transacao.descricao or "Sem descrição",
                "valor_bruto": valor_bruto,
                "impostos": {**impostos_detalhados, "aliquotas": aliquotas_usadas},
                "valor_impostos_total": valor_impostos_total,
                "valor_liquido": valor_liquido,
                "status": status_conta,
                "competencia_contabil": f"{transacao.competencia_mes_contabil:02d}/{transacao.competencia_ano_contabil}" if transacao.competencia_mes_contabil and transacao.competencia_ano_contabil else (f"{transacao.competencia_mes:02d}/{transacao.competencia_ano}" if transacao.competencia_mes and transacao.competencia_ano else "N/I"),
                "competencia_gerencial": f"{transacao.competencia_mes_gerencial:02d}/{transacao.competencia_ano_gerencial}" if transacao.competencia_mes_gerencial and transacao.competencia_ano_gerencial else "N/I",
                "data_emissao_nf": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
                "data_vencimento": transacao.data_vencimento.isoformat() if transacao.data_vencimento else None,
                "data_pagamento": transacao.data_pagamento.isoformat() if transacao.data_pagamento else None,
                "forma_pagamento": transacao.forma_pgto,
                "numero_nf": transacao.numero_nota_fiscal or "Não informado",
                "link_nota_fiscal": transacao.link_nota_fiscal or None,
                "valor_recebido": float(transacao.valor_recebido) if transacao.valor_recebido is not None else None
            })

        total_impostos = sum(c["valor_impostos_total"] for c in contas)
        total_liquido = total_valor - total_impostos

        aliquotas_resumo = {"pis": 0, "cofins": 0, "iss": 0, "irpj": 0, "csll": 0}
        if contas:
            first_aliq = contas[0].get("impostos", {}).get("aliquotas", {})
            uniform = all(
                c.get("impostos", {}).get("aliquotas", {}) == first_aliq
                for c in contas
            )
            if uniform and first_aliq:
                aliquotas_resumo = first_aliq

        return {
            "contas": contas,
            "resumo": {
                "total_contas": len(contas),
                "valor_total": total_valor,
                "valor_pendente": total_pendente,
                "valor_recebido": total_recebido,
                "total_impostos": round(total_impostos, 2),
                "total_liquido": round(total_liquido, 2),
                "aliquotas": aliquotas_resumo,
                "aliquotas_uniformes": uniform if contas else True
            },
            "filtros": {
                "ano": ano,
                "mes": mes,
                "status": status,
                "empresa": empresa
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório de contas a receber: {str(e)}")


@router.get("/relatorios/pl-contabil/export-excel")
async def export_pl_contabil_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None),
    fornecedor: Optional[int] = Query(None),
    centro_custo: Optional[int] = Query(None)
):
    """
    Exportar P&L Contábil em formato Excel (.xlsx)
    """
    try:
        # Buscar dados do P&L Contábil usando o endpoint existente
        dados_pl = await get_pl_contabil(
            current_user=current_user,
            db=db,
            ano=ano,
            empresa=empresa,
            cliente=cliente,
            projeto=projeto,
            produto_servico=produto_servico,
            fornecedor=fornecedor,
            centro_custo=centro_custo
        )

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"P&L Contábil {ano}"

        # Estilos
        header_fill = PatternFill(start_color="06b6d4", end_color="06b6d4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        subheader_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
        total_font = Font(bold=True, size=10)

        # Cabeçalho principal
        ws.merge_cells('A1:N1')
        ws['A1'] = f"P&L CONTÁBIL - {ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3

        # Cabeçalho da tabela
        meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

        ws[f'A{row}'] = 'DESCRIÇÃO'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill

        for i, mes in enumerate(meses):
            ws.cell(row=row, column=i+2, value=mes)
            ws.cell(row=row, column=i+2).font = header_font
            ws.cell(row=row, column=i+2).fill = header_fill
            ws.cell(row=row, column=i+2).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=14, value='TOTAL')
        ws.cell(row=row, column=14).font = header_font
        ws.cell(row=row, column=14).fill = header_fill
        ws.cell(row=row, column=14).alignment = Alignment(horizontal='center')

        row += 1

        totais = dados_pl['totais']
        
        # Estilos adicionais para hierarquia
        level1_fill = PatternFill(start_color="e0f2fe", end_color="e0f2fe", fill_type="solid")
        level2_fill = PatternFill(start_color="f0f9ff", end_color="f0f9ff", fill_type="solid")
        level3_fill = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

        # RECEITAS - Header
        ws[f'A{row}'] = 'RECEITAS'
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        for i, valor in enumerate(totais['receitas_mes']):
            ws.cell(row=row, column=i+2, value=valor)
            ws.cell(row=row, column=i+2).number_format = '#,##0'
            ws.cell(row=row, column=i+2).fill = subheader_fill
        ws.cell(row=row, column=14, value=totais['receitas_total'])
        ws.cell(row=row, column=14).number_format = '#,##0'
        ws.cell(row=row, column=14).fill = subheader_fill
        row += 1

        # Detalhar receitas por Cliente > Projeto > Produto
        receitas = dados_pl.get('receitas', {})
        for cliente_nome, cliente_data in sorted(receitas.items()):
            ws[f'A{row}'] = f"  {cliente_nome}"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].fill = level1_fill
            for i, valor in enumerate(cliente_data.get('meses', [0]*12)):
                ws.cell(row=row, column=i+2, value=valor)
                ws.cell(row=row, column=i+2).number_format = '#,##0'
                ws.cell(row=row, column=i+2).fill = level1_fill
            ws.cell(row=row, column=14, value=cliente_data.get('total', 0))
            ws.cell(row=row, column=14).number_format = '#,##0'
            ws.cell(row=row, column=14).fill = level1_fill
            row += 1
            
            for projeto_nome, projeto_data in sorted(cliente_data.get('projetos', {}).items()):
                ws[f'A{row}'] = f"    {projeto_nome}"
                ws[f'A{row}'].fill = level2_fill
                for i, valor in enumerate(projeto_data.get('meses', [0]*12)):
                    ws.cell(row=row, column=i+2, value=valor)
                    ws.cell(row=row, column=i+2).number_format = '#,##0'
                    ws.cell(row=row, column=i+2).fill = level2_fill
                ws.cell(row=row, column=14, value=projeto_data.get('total', 0))
                ws.cell(row=row, column=14).number_format = '#,##0'
                ws.cell(row=row, column=14).fill = level2_fill
                row += 1
                
                for produto_nome, produto_data in sorted(projeto_data.get('produtos', {}).items()):
                    ws[f'A{row}'] = f"      {produto_nome}"
                    for i, valor in enumerate(produto_data.get('meses', [0]*12)):
                        ws.cell(row=row, column=i+2, value=valor)
                        ws.cell(row=row, column=i+2).number_format = '#,##0'
                    ws.cell(row=row, column=14, value=produto_data.get('total', 0))
                    ws.cell(row=row, column=14).number_format = '#,##0'
                    row += 1

        # Linha em branco
        row += 1

        # DESPESAS - Header
        ws[f'A{row}'] = 'DESPESAS'
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        for i, valor in enumerate(totais['despesas_mes']):
            ws.cell(row=row, column=i+2, value=abs(valor))
            ws.cell(row=row, column=i+2).number_format = '#,##0'
            ws.cell(row=row, column=i+2).fill = subheader_fill
        ws.cell(row=row, column=14, value=abs(totais['despesas_total']))
        ws.cell(row=row, column=14).number_format = '#,##0'
        ws.cell(row=row, column=14).fill = subheader_fill
        row += 1

        # Detalhar despesas por Centro de Custo > Categoria > Título
        despesas = dados_pl.get('despesas', {})
        for centro_nome, centro_data in sorted(despesas.items()):
            ws[f'A{row}'] = f"  {centro_nome}"
            ws[f'A{row}'].font = Font(bold=True, size=10)
            ws[f'A{row}'].fill = level1_fill
            for i, valor in enumerate(centro_data.get('meses', [0]*12)):
                ws.cell(row=row, column=i+2, value=abs(valor))
                ws.cell(row=row, column=i+2).number_format = '#,##0'
                ws.cell(row=row, column=i+2).fill = level1_fill
            ws.cell(row=row, column=14, value=abs(centro_data.get('total', 0)))
            ws.cell(row=row, column=14).number_format = '#,##0'
            ws.cell(row=row, column=14).fill = level1_fill
            row += 1
            
            for categoria_nome, categoria_data in sorted(centro_data.get('categorias', {}).items()):
                ws[f'A{row}'] = f"    {categoria_nome}"
                ws[f'A{row}'].fill = level2_fill
                for i, valor in enumerate(categoria_data.get('meses', [0]*12)):
                    ws.cell(row=row, column=i+2, value=abs(valor))
                    ws.cell(row=row, column=i+2).number_format = '#,##0'
                    ws.cell(row=row, column=i+2).fill = level2_fill
                ws.cell(row=row, column=14, value=abs(categoria_data.get('total', 0)))
                ws.cell(row=row, column=14).number_format = '#,##0'
                ws.cell(row=row, column=14).fill = level2_fill
                row += 1
                
                for titulo_nome, titulo_data in sorted(categoria_data.get('titulos', {}).items()):
                    ws[f'A{row}'] = f"      {titulo_nome}"
                    for i, valor in enumerate(titulo_data.get('meses', [0]*12)):
                        ws.cell(row=row, column=i+2, value=abs(valor))
                        ws.cell(row=row, column=i+2).number_format = '#,##0'
                    ws.cell(row=row, column=14, value=abs(titulo_data.get('total', 0)))
                    ws.cell(row=row, column=14).number_format = '#,##0'
                    row += 1

        # Linha em branco
        row += 1

        # RESULTADO LÍQUIDO
        ws[f'A{row}'] = 'RESULTADO LÍQUIDO'
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill

        for i, valor in enumerate(totais['resultado_mes']):
            ws.cell(row=row, column=i+2, value=valor)
            ws.cell(row=row, column=i+2).number_format = '#,##0'
            ws.cell(row=row, column=i+2).font = total_font
            ws.cell(row=row, column=i+2).fill = total_fill
        ws.cell(row=row, column=14, value=totais['resultado_total'])
        ws.cell(row=row, column=14).number_format = '#,##0'
        ws.cell(row=row, column=14).font = total_font
        ws.cell(row=row, column=14).fill = total_fill

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"pl_contabil_{ano}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar P&L Contábil: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar P&L Contábil: {str(e)}")


@router.get("/relatorios/cashflow-gerencial/export-excel")
async def export_cashflow_gerencial_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None),
    cliente: Optional[int] = Query(None),
    projeto: Optional[int] = Query(None),
    produto_servico: Optional[int] = Query(None)
):
    """
    Exportar Cash Flow Gerencial em formato Excel (.xlsx)
    Inclui detalhes hierárquicos: Projetos (TIM, VIVO, CLARO, Outros) > Clientes > Produtos
    E Despesas: Centro de Custo > Categoria Contábil > Título
    """
    try:
        dados_cf = await get_cashflow_gerencial(
            current_user=current_user,
            db=db,
            ano=ano,
            empresa=empresa,
            cliente=cliente,
            projeto=projeto,
            produto_servico=produto_servico
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f"Cash Flow {ano}"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=10)
        group_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        group_font = Font(bold=True, size=10)
        subgroup_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        subgroup_font = Font(bold=True, size=9)
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(bold=True, size=10)
        imposto_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        imposto_font = Font(bold=True, size=9)

        ws.merge_cells('A1:N1')
        ws['A1'] = f"CASH FLOW GERENCIAL - {ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

        ws[f'A{row}'] = 'DESCRIÇÃO'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        for i, mes in enumerate(meses):
            ws.cell(row=row, column=i+2, value=mes)
            ws.cell(row=row, column=i+2).font = header_font
            ws.cell(row=row, column=i+2).fill = header_fill
            ws.cell(row=row, column=i+2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=14, value='TOTAL')
        ws.cell(row=row, column=14).font = header_font
        ws.cell(row=row, column=14).fill = header_fill
        ws.cell(row=row, column=14).alignment = Alignment(horizontal='center')
        row += 1

        def write_row(ws, row_num, label, meses_vals, total_val, fill=None, font=None, indent=0):
            label_with_indent = "  " * indent + label
            ws[f'A{row_num}'] = label_with_indent
            if fill:
                ws[f'A{row_num}'].fill = fill
            if font:
                ws[f'A{row_num}'].font = font
            for i, val in enumerate(meses_vals):
                cell = ws.cell(row=row_num, column=i+2, value=val if val != 0 else None)
                cell.number_format = '#,##0'
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
            total_cell = ws.cell(row=row_num, column=14, value=total_val if total_val != 0 else None)
            total_cell.number_format = '#,##0'
            if fill:
                total_cell.fill = fill
            if font:
                total_cell.font = font
            return row_num + 1

        totais = dados_cf['totais']
        row = write_row(ws, row, 'ENTRADAS (RECEITAS)', totais['entrada_mes'], totais['entrada_total'], section_fill, section_font)

        projetos = dados_cf.get('projetos', {})
        ordem_projetos = ['TIM', 'VIVO', 'CLARO']
        projetos_ordenados = []
        outros_projetos = []
        for proj_pref in ordem_projetos:
            for proj_nome in projetos.keys():
                if proj_pref.upper() in proj_nome.upper() and proj_nome not in projetos_ordenados:
                    projetos_ordenados.append(proj_nome)
        for proj_nome in sorted(projetos.keys()):
            if proj_nome not in projetos_ordenados:
                if 'OUTRO' in proj_nome.upper() or 'SEM PROJETO' in proj_nome.upper():
                    outros_projetos.append(proj_nome)
                else:
                    projetos_ordenados.append(proj_nome)
        projetos_ordenados.extend(outros_projetos)

        for projeto_nome in projetos_ordenados:
            projeto_data = projetos[projeto_nome]
            row = write_row(ws, row, projeto_nome, projeto_data.get('entrada_mes', [0]*12), projeto_data.get('entrada_total', 0), group_fill, group_font, indent=1)
            
            clientes = projeto_data.get('clientes', {})
            clientes_ordem = ['TIM', 'VIVO', 'CLARO']
            clientes_ordenados = []
            outros_clientes = []
            for cli_pref in clientes_ordem:
                for cli_nome in clientes.keys():
                    if cli_pref.upper() in cli_nome.upper() and cli_nome not in clientes_ordenados:
                        clientes_ordenados.append(cli_nome)
            for cli_nome in sorted(clientes.keys()):
                if cli_nome not in clientes_ordenados:
                    if 'OUTRO' in cli_nome.upper() or 'SEM CLIENTE' in cli_nome.upper():
                        outros_clientes.append(cli_nome)
                    else:
                        clientes_ordenados.append(cli_nome)
            clientes_ordenados.extend(outros_clientes)
            
            for cliente_nome in clientes_ordenados:
                cliente_data = clientes[cliente_nome]
                row = write_row(ws, row, cliente_nome, cliente_data.get('entrada_mes', [0]*12), cliente_data.get('entrada_total', 0), subgroup_fill, subgroup_font, indent=2)
                
                produtos = cliente_data.get('produtos_servicos', {})
                for produto_nome, produto_data in produtos.items():
                    row = write_row(ws, row, produto_nome, produto_data.get('entrada_mes', [0]*12), produto_data.get('entrada_total', 0), indent=3)

        row += 1
        row = write_row(ws, row, 'SAÍDAS (DESPESAS)', totais['saida_mes'], totais['saida_total'], section_fill, section_font)

        categorias_despesas = dados_cf.get('categorias_despesas', {})
        for centro_custo_nome in sorted(categorias_despesas.keys()):
            centro_data = categorias_despesas[centro_custo_nome]
            row = write_row(ws, row, centro_custo_nome, centro_data.get('saida_mes', [0]*12), centro_data.get('saida_total', 0), group_fill, group_font, indent=1)
            
            subcategorias = centro_data.get('subcategorias', {})
            for categoria_nome, categoria_data in subcategorias.items():
                row = write_row(ws, row, categoria_nome, categoria_data.get('saida_mes', [0]*12), categoria_data.get('saida_total', 0), subgroup_fill, subgroup_font, indent=2)
                
                titulos = categoria_data.get('titulos', {})
                for titulo_nome, titulo_data in titulos.items():
                    row = write_row(ws, row, titulo_nome, titulo_data.get('saida_mes', [0]*12), titulo_data.get('saida_total', 0), indent=3)

        impostos = dados_cf.get('impostos', {})
        if impostos:
            row += 1
            impostos_mes = [0]*12
            impostos_total = 0
            for tipo_imp, imp_data in impostos.items():
                for i, val in enumerate(imp_data.get('mes', [0]*12)):
                    impostos_mes[i] += val
                impostos_total += imp_data.get('total', 0)
            
            row = write_row(ws, row, 'IMPOSTOS SOBRE RECEITAS', impostos_mes, impostos_total, section_fill, section_font)
            impostos_ordem = ['pis', 'cofins', 'iss', 'irpj', 'csll']
            for tipo_imposto in impostos_ordem:
                if tipo_imposto in impostos:
                    imp_data = impostos[tipo_imposto]
                    row = write_row(ws, row, tipo_imposto.upper(), imp_data.get('mes', [0]*12), imp_data.get('total', 0), imposto_fill, imposto_font, indent=1)

        row += 1
        row = write_row(ws, row, 'FLUXO LÍQUIDO', totais['fluxo_liquido_mes'], totais['fluxo_liquido_total'], total_fill, total_font)

        row += 1
        row = write_row(ws, row, 'SALDO ACUMULADO', totais['saldo_acumulado_mes'], totais['saldo_acumulado_mes'][11] if totais['saldo_acumulado_mes'] else 0, total_fill, total_font)

        ws.column_dimensions['A'].width = 50
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"cashflow_{ano}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar Cash Flow: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Cash Flow: {str(e)}")


@router.get("/relatorios/cashflow")
async def get_cashflow(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: int = Query(default_factory=lambda: datetime.now().year),
    empresa: Optional[int] = Query(None)
):
    """
    Relatório de Cash Flow (Fluxo de Caixa) por mês
    """
    try:
        # Query base
        query = db.query(TransacaoFinanceira)

        # Filtro por ano
        query = query.filter(TransacaoFinanceira.competencia_ano == ano)

        # Usar empresa do filtro se fornecido, senão mostrar todas
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)

        # Buscar transações
        transacoes = query.all()

        # Inicializar estrutura
        cashflow = {
            "ano": ano,
            "empresa_id": empresa,
            "atividades": {
                "operacional": {"meses": [0] * 12, "total": 0},
                "investimento": {"meses": [0] * 12, "total": 0},
                "financiamento": {"meses": [0] * 12, "total": 0}
            },
            "fluxo_liquido": {"meses": [0] * 12, "total": 0},
            "saldo_acumulado": {"meses": [0] * 12, "total": 0}
        }

        saldo_anterior = 0

        # Processar transações
        for transacao in transacoes:
            try:
                if not transacao.competencia_mes or transacao.valor is None:
                    continue

                mes = int(transacao.competencia_mes) - 1
                if mes < 0 or mes > 11:
                    continue

                valor = float(transacao.valor)

                # Classificar por atividade (simplificado)
                atividade = "operacional"  # Padrão

                # Lógica lógica para classificar atividades
                if transacao.categoria_contabil_id:
                    try:
                        categoria = db.query(CategoriaContabil).filter(
                            CategoriaContabil.id == transacao.categoria_contabil_id
                        ).first()
                        if categoria and categoria.nome:
                            nome_categoria = str(categoria.nome).lower()
                            if "investimento" in nome_categoria or "ativo" in nome_categoria:
                                atividade = "investimento"
                            elif "empréstimo" in nome_categoria or "financiamento" in nome_categoria:
                                atividade = "financiamento"
                    except:
                        pass

                # Somar aos fluxos
                cashflow["atividades"][atividade]["meses"][mes] += valor
                cashflow["atividades"][atividade]["total"] += valor
            except:
                continue

        # Calcular fluxo líquido e saldo acumulado
        for i in range(12):
            fluxo_mes = (
                cashflow["atividades"]["operacional"]["meses"][i] +
                cashflow["atividades"]["investimento"]["meses"][i] +
                cashflow["atividades"]["financiamento"]["meses"][i]
            )
            cashflow["fluxo_liquido"]["meses"][i] = fluxo_mes
            cashflow["fluxo_liquido"]["total"] += fluxo_mes

            # Saldo acumulado
            saldo_anterior += fluxo_mes
            cashflow["saldo_acumulado"]["meses"][i] = saldo_anterior

        cashflow["saldo_acumulado"]["total"] = saldo_anterior

        return cashflow

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar cashflow: {str(e)}")


@router.get("/relatorios/rateio-impostos")
async def get_rateio_impostos(
    ano: int = Query(2025),
    mes_inicio: int = Query(1, ge=1, le=12),
    mes_fim: int = Query(12, ge=1, le=12),
    empresa: Optional[int] = None,
    fonte: str = Query("realizado", description="Fonte dos dados: 'realizado' para transações reais ou 'previsto' para orçamento"),
    db: Session = Depends(get_db)
):
    """Calcular rateio de impostos por empresa, cliente e produto/serviço
    
    Args:
        fonte: 'realizado' usa transações reais, 'previsto' usa orçamento planejado
    """
    try:
        print(f"\n{'='*80}")
        print(f"🧮 INICIANDO RATEIO DE IMPOSTOS - Ano {ano}, Meses {mes_inicio} a {mes_fim}, Fonte: {fonte}")
        print(f"{'='*80}\n")

        # Validar período
        if mes_inicio > mes_fim:
            raise HTTPException(status_code=400, detail="Mês inicial não pode ser maior que mês final")

        # Buscar todos os impostos ativos
        query_impostos = db.query(Imposto).filter(Imposto.ativo == True)
        if empresa:
            query_impostos = query_impostos.filter(Imposto.empresa_id == empresa)

        impostos = query_impostos.all()
        print(f"📊 Impostos ativos encontrados: {len(impostos)}")

        # Dados para processar
        dados_receitas = []

        if fonte == "realizado":
            # Buscar TRANSAÇÕES REAIS de receita no período
            receitas_query = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.competencia_ano_contabil == ano,
                TransacaoFinanceira.competencia_mes_contabil.between(mes_inicio, mes_fim),
                TransacaoFinanceira.parent_id.is_(None)  # Apenas transações pai
            )
            if empresa:
                receitas_query = receitas_query.filter(TransacaoFinanceira.empresa_id == empresa)
            
            transacoes = receitas_query.all()
            print(f"📊 Transações de receita encontradas: {len(transacoes)}")

            for t in transacoes:
                dados_receitas.append({
                    "empresa_id": t.empresa_id,
                    "cliente_id": t.cliente_id,
                    "produto_servico_id": t.produto_servico_id,
                    "mes": t.competencia_mes_contabil,
                    "valor": float(t.valor) if t.valor else 0.0
                })
        else:
            # Buscar receitas PLANEJADAS no período (orçamento)
            receitas = db.query(LinhaOrcamentaria).join(
                PlanejamentoVersao,
                LinhaOrcamentaria.versao_id == PlanejamentoVersao.id
            ).filter(
                LinhaOrcamentaria.ano == ano,
                LinhaOrcamentaria.categoria == 'receita',
                PlanejamentoVersao.status == 'publicado'
            )
            if empresa:
                receitas = receitas.filter(LinhaOrcamentaria.empresa_id == empresa)

            receitas = receitas.all()
            print(f"📊 Linhas orçamentárias de receita encontradas: {len(receitas)}")

            for r in receitas:
                dados_receitas.append({
                    "empresa_id": r.empresa_id,
                    "cliente_id": r.cliente_id,
                    "produto_servico_id": r.produto_servico_id,
                    "mes": r.mes,
                    "valor": float(r.valor_previsto) if r.valor_previsto else 0.0
                })

        # Estrutura para armazenar o rateio
        # {empresa_nome: {cliente_nome: {produto_nome: {imposto_nome: {meses: [], total: 0}}}}}
        rateio_impostos = {}

        print(f"📊 Processando {len(dados_receitas)} receitas para rateio de impostos...")

        # Processar cada receita (dados normalizados)
        for receita in dados_receitas:
            # Buscar empresa
            empresa_obj = db.query(Empresa).filter(Empresa.id == receita["empresa_id"]).first()
            empresa_nome = empresa_obj.nome_fantasia or empresa_obj.razao_social if empresa_obj else "Sem Empresa"
            
            # Buscar cliente
            cliente_obj = db.query(Cliente).filter(Cliente.id == receita["cliente_id"]).first() if receita["cliente_id"] else None
            cliente_nome = cliente_obj.nome if cliente_obj else "Sem Cliente"
            
            # Buscar produto/serviço
            produto_obj = db.query(ProdutoServico).filter(ProdutoServico.id == receita["produto_servico_id"]).first() if receita["produto_servico_id"] else None
            produto_nome = produto_obj.nome if produto_obj else "Geral"

            # Inicializar estruturas
            if empresa_nome not in rateio_impostos:
                rateio_impostos[empresa_nome] = {}
            if cliente_nome not in rateio_impostos[empresa_nome]:
                rateio_impostos[empresa_nome][cliente_nome] = {}
            if produto_nome not in rateio_impostos[empresa_nome][cliente_nome]:
                rateio_impostos[empresa_nome][cliente_nome][produto_nome] = {}

            # Buscar impostos aplicáveis (específicos do produto ou gerais da empresa)
            impostos_aplicaveis = [
                imp for imp in impostos 
                if imp.empresa_id == receita["empresa_id"] and (
                    imp.produto_servico_id == receita["produto_servico_id"] or 
                    imp.produto_servico_id is None
                )
            ]

            # Usar alíquotas padrão se não houver impostos cadastrados
            if not impostos_aplicaveis:
                # Criar impostos fictícios com alíquotas padrão para cálculo
                aliquotas_padrao = [
                    {"nome": "PIS", "valor": 0.65, "tipo": "federal"},
                    {"nome": "COFINS", "valor": 3.0, "tipo": "federal"},
                    {"nome": "ISS", "valor": 5.0, "tipo": "municipal"},
                    {"nome": "IRPJ", "valor": 7.93, "tipo": "federal"},
                    {"nome": "CSLL", "valor": 2.88, "tipo": "federal"}
                ]
                for aliquota in aliquotas_padrao:
                    imposto_nome = aliquota["nome"]
                    valor_receita = receita["valor"]
                    mes_receita = receita["mes"]
                    
                    if imposto_nome not in rateio_impostos[empresa_nome][cliente_nome][produto_nome]:
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome] = {
                            "meses": [0.0] * (mes_fim - mes_inicio + 1),
                            "total": 0.0,
                            "aliquota": aliquota["valor"],
                            "tipo": aliquota["tipo"],
                            "cumulativo": False
                        }

                    # Calcular valor do imposto
                    valor_imposto = valor_receita * (aliquota["valor"] / 100.0)

                    # Armazenar no mês correto
                    if mes_inicio <= mes_receita <= mes_fim:
                        idx_mes = mes_receita - mes_inicio
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome]["meses"][idx_mes] += valor_imposto
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome]["total"] += valor_imposto
            else:
                # Usar impostos cadastrados
                for imposto in impostos_aplicaveis:
                    imposto_nome = imposto.nome
                    valor_receita = receita["valor"]
                    mes_receita = receita["mes"]
                    aliquota_valor = float(imposto.valor) if imposto.valor else 0.0

                    if imposto_nome not in rateio_impostos[empresa_nome][cliente_nome][produto_nome]:
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome] = {
                            "meses": [0.0] * (mes_fim - mes_inicio + 1),
                            "total": 0.0,
                            "aliquota": aliquota_valor,
                            "tipo": imposto.tipo,
                            "cumulativo": imposto.cumulativo
                        }

                    # Calcular valor do imposto
                    valor_imposto = valor_receita * (aliquota_valor / 100.0)

                    # Armazenar no mês correto
                    if mes_inicio <= mes_receita <= mes_fim:
                        idx_mes = mes_receita - mes_inicio
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome]["meses"][idx_mes] += valor_imposto
                        rateio_impostos[empresa_nome][cliente_nome][produto_nome][imposto_nome]["total"] += valor_imposto

        # Calcular totais consolidados
        total_impostos_geral = 0.0
        totais_por_empresa = {}
        totais_por_tipo = {"federal": 0.0, "estadual": 0.0, "municipal": 0.0, "outros": 0.0}

        for empresa_nome, clientes in rateio_impostos.items():
            totais_por_empresa[empresa_nome] = 0.0
            for cliente_nome, produtos in clientes.items():
                for produto_nome, impostos_dict in produtos.items():
                    for imposto_nome, dados in impostos_dict.items():
                        total_impostos_geral += dados["total"]
                        totais_por_empresa[empresa_nome] += dados["total"]

                        tipo = dados.get("tipo") or "outros"
                        if tipo not in totais_por_tipo:
                            tipo = "outros"
                        totais_por_tipo[tipo] += dados["total"]

        print(f"✅ Rateio de impostos calculado: R$ {total_impostos_geral:,.2f}")

        return {
            "ano": ano,
            "mes_inicio": mes_inicio,
            "mes_fim": mes_fim,
            "total_impostos": total_impostos_geral,
            "totais_por_empresa": totais_por_empresa,
            "totais_por_tipo": totais_por_tipo,
            "rateio": rateio_impostos,
            "quantidade_empresas": len(rateio_impostos)
        }

    except Exception as e:
        print(f"❌ Erro no rateio de impostos: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/relatorios/retencao-fonte")
async def get_retencao_fonte(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    empresa: Optional[int] = Query(None),
    imposto_nome: Optional[str] = Query(None)
):
    """
    Relatório de Retenção na Fonte - Impostos de Terceiros a Recolher
    Lista todas as transações de impostos retidos de fornecedores.
    Identifica retenções pelo nome começando com "Retenção" e parent_id não nulo.
    """
    try:
        from sqlalchemy import and_

        # Query base: transações que são retenções de impostos (filhos com nome "Retenção...")
        query = db.query(TransacaoFinanceira).filter(
            and_(
                TransacaoFinanceira.parent_id.isnot(None),
                TransacaoFinanceira.tipo_filho == 'split',
                TransacaoFinanceira.nome.like('Retenção%'),
                TransacaoFinanceira.entra_no_gerencial == False
            )
        )

        # Filtros
        if empresa:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa)
        if ano:
            query = query.filter(TransacaoFinanceira.competencia_ano == ano)
        if mes:
            query = query.filter(TransacaoFinanceira.competencia_mes == mes)
        if status:
            if status == 'pago':
                query = query.filter(TransacaoFinanceira.data_pagamento.isnot(None))
            elif status == 'pendente':
                query = query.filter(TransacaoFinanceira.data_pagamento.is_(None))
        if imposto_nome:
            query = query.filter(TransacaoFinanceira.nome.ilike(f'%{imposto_nome}%'))

        # Buscar com join para empresa
        transacoes = query.outerjoin(
            Empresa, TransacaoFinanceira.empresa_id == Empresa.id
        ).add_columns(
            Empresa.nome_fantasia.label('empresa_nome')
        ).order_by(TransacaoFinanceira.data_vencimento.desc()).all()

        # Processar dados
        retencoes = []
        total_valor = 0.0
        total_pendente = 0.0
        total_pago = 0.0
        por_imposto = {}  # Agrupamento por tipo de imposto

        for item in transacoes:
            transacao = item[0]
            empresa_nome = item[1] if item[1] else "Empresa não informada"

            valor = abs(float(transacao.valor or 0))
            
            # Extrair nome do imposto (ex: "Retenção ISS" -> "ISS")
            imposto = transacao.nome.replace("Retenção ", "").strip() if transacao.nome else "Imposto"
            
            # Buscar transação pai para obter informações do fornecedor
            transacao_pai = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.id == transacao.parent_id
            ).first()
            
            fornecedor_nome = "Fornecedor não informado"
            if transacao_pai and transacao_pai.fornecedor_id:
                fornecedor_obj = db.query(Fornecedor).filter(Fornecedor.id == transacao_pai.fornecedor_id).first()
                if fornecedor_obj:
                    fornecedor_nome = fornecedor_obj.nome

            # Determinar status
            if transacao.data_pagamento:
                status_conta = "Pago"
                total_pago += valor
            else:
                status_conta = "Pendente"
                total_pendente += valor

            total_valor += valor

            # Agrupar por tipo de imposto
            if imposto not in por_imposto:
                por_imposto[imposto] = {"quantidade": 0, "valor_total": 0.0}
            por_imposto[imposto]["quantidade"] += 1
            por_imposto[imposto]["valor_total"] += valor

            retencoes.append({
                "id": transacao.id,
                "parent_id": transacao.parent_id,
                "imposto": imposto,
                "fornecedor": fornecedor_nome,
                "empresa": empresa_nome,
                "descricao": transacao.descricao or transacao.nome,
                "valor": valor,
                "status": status_conta,
                "competencia": f"{transacao.competencia_mes:02d}/{transacao.competencia_ano}" if transacao.competencia_mes and transacao.competencia_ano else None,
                "data_vencimento": transacao.data_vencimento.isoformat() if transacao.data_vencimento else None,
                "data_pagamento": transacao.data_pagamento.isoformat() if transacao.data_pagamento else None
            })

        # Ordenar resumo por imposto por valor decrescente
        resumo_impostos = sorted(
            [{"nome": k, **v} for k, v in por_imposto.items()],
            key=lambda x: x["valor_total"],
            reverse=True
        )

        return {
            "retencoes": retencoes,
            "resumo": {
                "total_retencoes": len(retencoes),
                "valor_total": total_valor,
                "valor_pendente": total_pendente,
                "valor_pago": total_pago,
                "por_imposto": resumo_impostos
            },
            "filtros": {
                "ano": ano,
                "mes": mes,
                "status": status,
                "empresa": empresa,
                "imposto_nome": imposto_nome
            }
        }

    except Exception as e:
        print(f"❌ Erro no relatório de retenção na fonte: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório de retenção na fonte: {str(e)}")


@router.get("/relatorios/contas-a-pagar/export-excel")
async def export_contas_a_pagar_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    fornecedor: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    empresa: Optional[int] = Query(None),
    tipo_data: Optional[str] = Query("contabil"),
    data_pgto_de: Optional[str] = Query(None),
    data_pgto_ate: Optional[str] = Query(None),
    ano_pgto: Optional[int] = Query(None),
    mes_pgto: Optional[int] = Query(None)
):
    """
    Exportar Contas a Pagar em formato Excel (.xlsx)
    """
    try:
        dados = await get_contas_a_pagar(
            current_user=current_user,
            db=db,
            ano=ano,
            mes=mes,
            status=status,
            fornecedor=fornecedor,
            descricao=descricao,
            empresa=empresa,
            tipo_data=tipo_data,
            data_pgto_de=data_pgto_de,
            data_pgto_ate=data_pgto_ate,
            ano_pgto=ano_pgto,
            mes_pgto=mes_pgto
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Contas a Pagar"

        header_fill = PatternFill(start_color="f59e0b", end_color="f59e0b", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
        total_font = Font(bold=True, size=10)
        pago_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        pendente_fill = PatternFill(start_color="fecaca", end_color="fecaca", fill_type="solid")

        ws.merge_cells('A1:G1')
        titulo = f"CONTAS A PAGAR"
        if ano:
            titulo += f" - {ano}"
        if mes:
            meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                          'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            titulo += f" - {meses_nomes[mes]}"
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        resumo = dados.get('resumo', {})
        ws['A2'] = f"Total: {resumo.get('total_contas', 0)} contas | Valor Total: R$ {resumo.get('valor_total', 0):,.2f} | Pendente: R$ {resumo.get('valor_pendente', 0):,.2f} | Pago: R$ {resumo.get('valor_pago', 0):,.2f}"
        ws['A2'].font = Font(size=10, italic=True)

        row = 4
        headers = ['Fornecedor', 'Descrição', 'Valor', 'Status', 'Vencimento', 'Pagamento', 'Documento']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1

        for conta in dados.get('contas', []):
            ws.cell(row=row, column=1, value=conta.get('fornecedor', ''))
            ws.cell(row=row, column=2, value=conta.get('descricao', ''))
            ws.cell(row=row, column=3, value=conta.get('valor', 0))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=conta.get('status', ''))
            
            if conta.get('status') == 'Pago':
                ws.cell(row=row, column=4).fill = pago_fill
            else:
                ws.cell(row=row, column=4).fill = pendente_fill
            
            vencimento = conta.get('data_vencimento')
            ws.cell(row=row, column=5, value=vencimento if vencimento else '')
            pagamento = conta.get('data_pagamento')
            ws.cell(row=row, column=6, value=pagamento if pagamento else '')
            ws.cell(row=row, column=7, value=conta.get('documento', '') or '')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=3, value=resumo.get('valor_total', 0))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).font = total_font
        ws.cell(row=row, column=3).fill = total_fill

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"contas_a_pagar_{ano or 'todos'}_{mes or 'todos'}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar Contas a Pagar: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Contas a Pagar: {str(e)}")


@router.get("/relatorios/contas-a-receber/export-excel")
async def export_contas_a_receber_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    cliente: Optional[str] = Query(None),
    descricao: Optional[str] = Query(None),
    empresa: Optional[int] = Query(None),
    clientes_com_valores: Optional[str] = Query(None)
):
    """
    Exportar Contas a Receber em formato Excel (.xlsx)
    """
    try:
        dados = await get_contas_a_receber(
            current_user=current_user,
            db=db,
            ano=ano,
            mes=mes,
            status=status,
            cliente=cliente,
            descricao=descricao,
            empresa=empresa
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Contas a Receber"

        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        total_font = Font(bold=True, size=10)
        recebido_fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
        pendente_fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")

        ws.merge_cells('A1:Q1')
        titulo = f"CONTAS A RECEBER"
        if ano:
            titulo += f" - {ano}"
        if mes:
            meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                          'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            titulo += f" - {meses_nomes[mes]}"
        ws['A1'] = titulo
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        resumo = dados.get('resumo', {})
        ws['A2'] = f"Total: {resumo.get('total_contas', 0)} contas | Valor Total: R$ {resumo.get('valor_total', 0):,.2f} | Pendente: R$ {resumo.get('valor_pendente', 0):,.2f} | Recebido: R$ {resumo.get('valor_recebido', 0):,.2f}"
        ws['A2'].font = Font(size=10, italic=True)

        row = 4
        headers = ['Cliente', 'CNPJ', 'Descrição', 'Valor Bruto', 'PIS', 'COFINS', 'ISS', 'IRPJ', 'CSLL', 
                   'Total Impostos', 'Valor Líquido', 'Status', 'Competência', 'Emissão NF', 'Vencimento', 'Recebimento', 'NF']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = header_font
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        row += 1

        for conta in dados.get('contas', []):
            impostos = conta.get('impostos', {})
            ws.cell(row=row, column=1, value=conta.get('cliente', ''))
            ws.cell(row=row, column=2, value=conta.get('cliente_cnpj', ''))
            ws.cell(row=row, column=3, value=conta.get('descricao', ''))
            ws.cell(row=row, column=4, value=conta.get('valor_bruto', 0))
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=impostos.get('pis', 0))
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=impostos.get('cofins', 0))
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=impostos.get('iss', 0))
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=impostos.get('irpj', 0))
            ws.cell(row=row, column=8).number_format = '#,##0.00'
            ws.cell(row=row, column=9, value=impostos.get('csll', 0))
            ws.cell(row=row, column=9).number_format = '#,##0.00'
            ws.cell(row=row, column=10, value=conta.get('valor_impostos_total', 0))
            ws.cell(row=row, column=10).number_format = '#,##0.00'
            ws.cell(row=row, column=11, value=conta.get('valor_liquido', 0))
            ws.cell(row=row, column=11).number_format = '#,##0.00'
            ws.cell(row=row, column=12, value=conta.get('status', ''))
            
            if conta.get('status') == 'Recebido':
                ws.cell(row=row, column=12).fill = recebido_fill
            else:
                ws.cell(row=row, column=12).fill = pendente_fill
            
            ws.cell(row=row, column=13, value=conta.get('competencia', ''))
            emissao = conta.get('data_emissao_nf')
            ws.cell(row=row, column=14, value=emissao if emissao else '')
            vencimento = conta.get('data_vencimento')
            ws.cell(row=row, column=15, value=vencimento if vencimento else '')
            recebimento = conta.get('data_pagamento')
            ws.cell(row=row, column=16, value=recebimento if recebimento else '')
            ws.cell(row=row, column=17, value=conta.get('numero_nf', '') or '')
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=4, value=resumo.get('valor_total', 0))
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=4).font = total_font
        ws.cell(row=row, column=4).fill = total_fill

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 14
        for col in range(5, 12):
            ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"contas_a_receber_{ano or 'todos'}_{mes or 'todos'}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar Contas a Receber: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao exportar Contas a Receber: {str(e)}")