"""
Rotas de transações financeiras - endpoint principal
"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request, Body
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func, select
from typing import List, Optional, Dict
from app.database import get_db
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
from app.models import (
    TransacaoFinanceira, Empresa, Cliente, Fornecedor,
    CategoriaGerencial, CategoriaContabil, ContaBancaria,
    ProdutoServico, Projeto, Imposto
)
from app.models.base import TipoTransacao
from app.auth.oauth import get_current_user
from datetime import date, datetime
from pydantic import BaseModel
from typing import Dict, Any

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

# Função helper para aplicar filtros consistentes entre endpoints
def build_base_transacao_query(db: Session, tipo: str = None, empresa_id: int = None, ano: int = None, mes: int = None):
    """
    Constrói query base para transações com filtros consistentes
    Garante que todos os endpoints usem a mesma fonte de dados
    """
    query = db.query(TransacaoFinanceira)

    if tipo:
        query = query.filter(TransacaoFinanceira.tipo == tipo)

    if empresa_id:
        query = query.filter(TransacaoFinanceira.empresa_id == empresa_id)

    if ano:
        query = query.filter(TransacaoFinanceira.competencia_ano == ano)

    if mes:
        query = query.filter(TransacaoFinanceira.competencia_mes == mes)

    return query


def apply_leaf_nodes_filter(query, db: Session):
    """
    🌳 HIERARQUIA PAI-FILHO: Filtra apenas leaf nodes (nós folha)
    
    Lógica:
    - Leaf nodes = pais sem filhos + todos os filhos
    - Pais com filhos = EXCLUÍDOS (servem apenas como agrupamento)
    
    Isso garante que após desmembramento (1 item → 2 itens):
    - O pai (que agora tem filhos) é excluído dos totais
    - Apenas os 2 filhos aparecem nas contagens
    - Resultado: 2 itens ao invés de 3 (1 pai + 2 filhos)
    """
    from sqlalchemy import select
    
    # Subquery para encontrar todos os IDs que são pais (têm filhos)
    parents_with_children_subquery = select(TransacaoFinanceira.parent_id).where(
        TransacaoFinanceira.parent_id.isnot(None)
    ).distinct()
    
    # Filtrar apenas registros que NÃO estão na lista de pais com filhos
    return query.filter(
        ~TransacaoFinanceira.id.in_(parents_with_children_subquery)
    )

# Schema para criação de transação
class TransacaoCreate(BaseModel):
    empresa_id: int
    tipo: str  # 'receita' ou 'despesa'
    titulo: Optional[str] = None
    descricao: Optional[str] = None
    valor: float
    data_emissao: Optional[str] = None
    data_lancamento: Optional[str] = None
    data_vencimento: Optional[str] = None
    status: Optional[str] = 'pendente'
    forma_recebimento: Optional[str] = None
    cliente_id: Optional[int] = None
    fornecedor_id: Optional[int] = None
    projeto_id: Optional[int] = None
    produto_servico_id: Optional[int] = None
    categoria_gerencial_id: Optional[int] = None
    categoria_contabil_id: Optional[int] = None
    subcategoria_gerencial_id: Optional[int] = None
    subcategoria_contabil_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    conta_contabil_id: Optional[int] = None
    competencia_contabil: Optional[str] = None
    competencia_gerencial: Optional[str] = None
    exibir_no_cash_control: Optional[bool] = True
    entra_no_gerencial: Optional[bool] = True
    numero_nota_fiscal: Optional[str] = None
    link_nota_fiscal: Optional[str] = None
    numero_pedido_compra: Optional[str] = None
    link_pedido_compra: Optional[str] = None
    retencao: Optional[Dict[str, Any]] = None  # Dados de retenção na fonte
    valor_recebido: Optional[float] = None
    valor_pago: Optional[float] = None


# Schema para atualização de transação
class TransacaoUpdate(BaseModel):
    tipo: str
    descricao: str
    valor: float
    data_lancamento: str
    data_vencimento: Optional[str] = None
    data_recebimento: Optional[str] = None
    status: Optional[str] = 'pendente'
    forma_pagamento: Optional[str] = None
    forma_recebimento: Optional[str] = None
    empresa_id: Optional[int] = None
    cliente_id: Optional[int] = None
    fornecedor_id: Optional[int] = None
    categoria_gerencial_id: Optional[int] = None
    subcategoria_gerencial_id: Optional[int] = None
    categoria_contabil_id: Optional[int] = None
    subcategoria_contabil_id: Optional[int] = None
    centro_custo_id: Optional[int] = None
    conta_contabil_id: Optional[int] = None
    projeto_id: Optional[int] = None
    produto_servico_id: Optional[int] = None
    numero_nota_fiscal: Optional[str] = None
    link_nota_fiscal: Optional[str] = None
    numero_pedido_compra: Optional[str] = None
    link_pedido_compra: Optional[str] = None
    competencia: Optional[str] = None
    competencia_contabil: Optional[str] = None
    competencia_gerencial: Optional[str] = None
    exibir_no_cash_control: Optional[bool] = True
    entra_no_gerencial: Optional[bool] = True
    valor_recebido: Optional[float] = None
    valor_pago: Optional[float] = None


# Schema para bulk update
class BulkCategorizationRequest(BaseModel):
    transaction_ids: List[int]
    tipo: str
    updates: Dict[str, Any]


# Rotas para páginas de edição usando formulários completos (MOVIDA PARA main.py)
"""
async def editar_transacao_page(
    request: Request,
    transacao_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # FUNÇÃO MOVIDA PARA main.py para evitar conflito de rota
"""


@router.get("/transacoes/historico-sugestao")
async def get_historico_sugestao(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    tipo: str = Query(..., description="Tipo: 'receita' ou 'despesa'"),
    empresa_id: Optional[int] = Query(None),
    fornecedor_id: Optional[int] = Query(None),
    cliente_id: Optional[int] = Query(None),
    centro_custo_id: Optional[int] = Query(None),
    categoria_gerencial_id: Optional[int] = Query(None),
    categoria_contabil_id: Optional[int] = Query(None)
):
    """
    Busca sugestões de preenchimento baseadas no histórico de transações.
    Retorna a transação mais recente que corresponde aos critérios fornecidos.
    Usado para auto-preenchimento inteligente nos formulários.
    """
    try:
        query = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == tipo
        )
        
        if empresa_id:
            query = query.filter(TransacaoFinanceira.empresa_id == empresa_id)
        if fornecedor_id:
            query = query.filter(TransacaoFinanceira.fornecedor_id == fornecedor_id)
        if cliente_id:
            query = query.filter(TransacaoFinanceira.cliente_id == cliente_id)
        if centro_custo_id:
            query = query.filter(TransacaoFinanceira.centro_custo_id == centro_custo_id)
        if categoria_gerencial_id:
            query = query.filter(TransacaoFinanceira.categoria_gerencial_id == categoria_gerencial_id)
        if categoria_contabil_id:
            query = query.filter(TransacaoFinanceira.categoria_contabil_id == categoria_contabil_id)
        
        transacao = query.order_by(TransacaoFinanceira.data_lancamento.desc()).first()
        
        if not transacao:
            return {"found": False, "sugestao": None}
        
        return {
            "found": True,
            "sugestao": {
                "descricao": transacao.descricao,
                "empresa_id": transacao.empresa_id,
                "fornecedor_id": transacao.fornecedor_id,
                "cliente_id": transacao.cliente_id,
                "centro_custo_id": transacao.centro_custo_id,
                "categoria_gerencial_id": transacao.categoria_gerencial_id,
                "subcategoria_gerencial_id": transacao.subcategoria_gerencial_id,
                "categoria_contabil_id": transacao.categoria_contabil_id,
                "subcategoria_contabil_id": transacao.subcategoria_contabil_id,
                "conta_contabil_id": transacao.conta_contabil_id,
                "projeto_id": transacao.projeto_id,
                "produto_servico_id": transacao.produto_servico_id,
                "valor": abs(transacao.valor) if transacao.valor else None
            }
        }
    except Exception as e:
        print(f"❌ Erro ao buscar histórico de sugestão: {e}")
        return {"found": False, "sugestao": None, "error": str(e)}


@router.get("/transacoes/fornecedores-recentes")
async def get_fornecedores_recentes(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = Query(10, ge=1, le=50)
):
    """
    Retorna os fornecedores mais usados recentemente em despesas.
    Ordenados por frequência de uso nos últimos 90 dias.
    """
    try:
        from sqlalchemy import desc
        from datetime import timedelta
        
        data_limite = date.today() - timedelta(days=90)
        
        fornecedores_frequentes = db.query(
            TransacaoFinanceira.fornecedor_id,
            Fornecedor.nome,
            func.count(TransacaoFinanceira.id).label('frequencia')
        ).join(
            Fornecedor, TransacaoFinanceira.fornecedor_id == Fornecedor.id
        ).filter(
            TransacaoFinanceira.tipo == 'despesa',
            TransacaoFinanceira.fornecedor_id.isnot(None),
            TransacaoFinanceira.data_lancamento >= data_limite
        ).group_by(
            TransacaoFinanceira.fornecedor_id,
            Fornecedor.nome
        ).order_by(
            desc('frequencia')
        ).limit(limit).all()
        
        return [
            {"id": f.fornecedor_id, "nome": f.nome, "frequencia": f.frequencia}
            for f in fornecedores_frequentes
        ]
    except Exception as e:
        print(f"❌ Erro ao buscar fornecedores recentes: {e}")
        return []


@router.get("/transacoes/clientes-recentes")
async def get_clientes_recentes(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    limit: int = Query(10, ge=1, le=50)
):
    """
    Retorna os clientes mais usados recentemente em receitas.
    Ordenados por frequência de uso nos últimos 90 dias.
    """
    try:
        from sqlalchemy import desc
        from datetime import timedelta
        
        data_limite = date.today() - timedelta(days=90)
        
        clientes_frequentes = db.query(
            TransacaoFinanceira.cliente_id,
            Cliente.nome,
            func.count(TransacaoFinanceira.id).label('frequencia')
        ).join(
            Cliente, TransacaoFinanceira.cliente_id == Cliente.id
        ).filter(
            TransacaoFinanceira.tipo == 'receita',
            TransacaoFinanceira.cliente_id.isnot(None),
            TransacaoFinanceira.data_lancamento >= data_limite
        ).group_by(
            TransacaoFinanceira.cliente_id,
            Cliente.nome
        ).order_by(
            desc('frequencia')
        ).limit(limit).all()
        
        return [
            {"id": c.cliente_id, "nome": c.nome, "frequencia": c.frequencia}
            for c in clientes_frequentes
        ]
    except Exception as e:
        print(f"❌ Erro ao buscar clientes recentes: {e}")
        return []


@router.get("/transacoes/summary")
async def get_summary(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    mes: Optional[int] = Query(None, ge=1, le=12),
    ano: Optional[int] = Query(None, ge=2000, le=2100),
    tipo_data: Optional[str] = Query("competencia", description="Tipo de data para filtro: 'competencia', 'competencia_gerencial' ou 'lancamento'")
):
    """
    Resumo consolidado de todas as transações com filtro opcional de mês/ano
    Suporta filtro por competência contábil (padrão), gerencial ou por data de lançamento
    """
    try:
        from sqlalchemy import extract
        
        # 🌳 HIERARQUIA: Calcular totais GERAIS (sem filtro de período) - APENAS LEAF NODES
        total_receitas_geral_query = db.query(func.count(TransacaoFinanceira.id)).filter(
            TransacaoFinanceira.tipo == 'receita'
        )
        total_receitas_geral = apply_leaf_nodes_filter(total_receitas_geral_query, db).scalar() or 0
        
        total_despesas_geral_query = db.query(func.count(TransacaoFinanceira.id)).filter(
            TransacaoFinanceira.tipo == 'despesa'
        )
        total_despesas_geral = apply_leaf_nodes_filter(total_despesas_geral_query, db).scalar() or 0
        
        total_transacoes_geral = total_receitas_geral + total_despesas_geral

        # Query base - TODAS as empresas (comportamento padrão)
        query = db.query(TransacaoFinanceira)

        # Aplicar filtros de período baseado no tipo de data selecionado
        if tipo_data == "lancamento":
            if ano:
                query = query.filter(extract('year', TransacaoFinanceira.data_lancamento) == ano)
            if mes:
                query = query.filter(extract('month', TransacaoFinanceira.data_lancamento) == mes)
            periodo_str = f"mês {mes}/{ano} (por lançamento)" if mes and ano else f"ano {ano} (por lançamento)" if ano else "TODAS as empresas"
        elif tipo_data == "competencia_gerencial":
            if ano:
                query = query.filter(TransacaoFinanceira.competencia_ano_gerencial == ano)
            if mes:
                query = query.filter(TransacaoFinanceira.competencia_mes_gerencial == mes)
            periodo_str = f"mês {mes}/{ano} (por comp. gerencial)" if mes and ano else f"ano {ano} (por comp. gerencial)" if ano else "TODAS as empresas"
        else:
            if ano:
                query = query.filter(TransacaoFinanceira.competencia_ano == ano)
            if mes:
                query = query.filter(TransacaoFinanceira.competencia_mes == mes)
            periodo_str = f"mês {mes}/{ano} (por competência)" if mes and ano else f"ano {ano} (por competência)" if ano else "TODAS as empresas"
        
        print(f"🔍 Summary: Calculando para {periodo_str}")

        # 🌳 HIERARQUIA: Receitas do período - APENAS LEAF NODES
        receitas_query = query.filter(TransacaoFinanceira.tipo == 'receita')
        receitas_query = apply_leaf_nodes_filter(receitas_query, db)
        total_receitas = receitas_query.with_entities(func.sum(TransacaoFinanceira.valor)).scalar() or 0
        qtd_receitas = receitas_query.count()

        # 🌳 HIERARQUIA: Despesas do período - APENAS LEAF NODES
        # 🔧 CORREÇÃO CRÍTICA: Calcular despesas usando ABS para garantir sempre valores positivos
        despesas_query = query.filter(TransacaoFinanceira.tipo == 'despesa')
        despesas_query = apply_leaf_nodes_filter(despesas_query, db)

        # Usar func.abs() diretamente na query para garantir valores sempre positivos
        total_despesas_absoluto = despesas_query.with_entities(func.sum(func.abs(TransacaoFinanceira.valor))).scalar() or 0
        qtd_despesas = despesas_query.count()

        # Garantir que seja sempre um float positivo
        total_despesas_absoluto = float(abs(total_despesas_absoluto))

        # Total de transações do período
        qtd_transacoes = qtd_receitas + qtd_despesas

        # Resultado líquido (receitas - despesas sempre positivas)
        lucro_liquido = float(total_receitas) - total_despesas_absoluto

        print(f"✅ Summary calculado ({periodo_str}): receitas={total_receitas}, despesas_absoluto={total_despesas_absoluto}, lucro_liquido={lucro_liquido}")
        print(f"✅ Contadores: receitas={qtd_receitas}/{total_receitas_geral}, despesas={qtd_despesas}/{total_despesas_geral}, total={qtd_transacoes}/{total_transacoes_geral}")
        print(f"🔍 DEBUG: Verificação de valores - receitas_tipo={type(total_receitas)}, despesas_tipo={type(total_despesas_absoluto)}")

        # Garantir que todos os valores sejam float válidos
        total_receitas_final = float(total_receitas) if total_receitas else 0.0
        total_despesas_final = float(total_despesas_absoluto) if total_despesas_absoluto else 0.0
        lucro_liquido_final = total_receitas_final - total_despesas_final

        return {
            "total_receitas": total_receitas_final,
            "total_despesas": total_despesas_final,  # Sempre valor absoluto positivo
            "lucro_liquido": lucro_liquido_final,
            "resultado": lucro_liquido_final,
            "qtd_receitas": qtd_receitas,
            "qtd_despesas": qtd_despesas,
            "qtd_transacoes": qtd_transacoes,
            "total_receitas_geral": total_receitas_geral,
            "total_despesas_geral": total_despesas_geral,
            "total_transacoes_geral": total_transacoes_geral,
            "modo": "todas_empresas"  # Indicar que são dados consolidados
        }
    except Exception as e:
        print(f"❌ Erro no summary: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao calcular resumo: {str(e)}")


@router.post("/transacoes")
async def criar_transacao(
    transacao_data: TransacaoCreate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Criar nova transação financeira
    Compatível com: POST /api/transacoes
    """
    try:
        # Processar data de lançamento usando fuso horário local
        data_lancamento = date.today()
        if hasattr(transacao_data, 'data_emissao') and transacao_data.data_emissao:
            try:
                # Converter string de data diretamente para date (sem UTC)
                if 'T' in transacao_data.data_emissao:
                    data_lancamento = datetime.fromisoformat(transacao_data.data_emissao.split('T')[0]).date()
                else:
                    data_lancamento = datetime.fromisoformat(transacao_data.data_emissao).date()
            except:
                data_lancamento = date.today()
        elif transacao_data.data_lancamento:
            try:
                # Converter string de data diretamente para date (sem UTC)
                if 'T' in transacao_data.data_lancamento:
                    data_lancamento = datetime.fromisoformat(transacao_data.data_lancamento.split('T')[0]).date()
                else:
                    data_lancamento = datetime.fromisoformat(transacao_data.data_lancamento).date()
            except:
                data_lancamento = date.today()

        # ✅ FLEXIBILIDADE TOTAL: Permitir criação de transações para qualquer empresa
        # Usar empresa_id fornecida no request ou padrão se não especificada
        empresa_id = transacao_data.empresa_id
        if not empresa_id:
            # Se não especificada, usar empresa do usuário como fallback
            user_empresa_id = getattr(current_user, 'empresa_id', None)
            if user_empresa_id:
                empresa_id = user_empresa_id
            else:
                raise HTTPException(status_code=400, detail="empresa_id é obrigatória")

        # Processar valor baseado no tipo
        valor_original = float(transacao_data.valor)
        print(f"🔍 DEBUG: Valor original recebido: {valor_original}, tipo: {transacao_data.tipo}")

        # 🔧 CORREÇÃO: Garantir consistência total no armazenamento
        if transacao_data.tipo == 'despesa':
            # Despesas SEMPRE negativas no banco, independente do valor recebido
            valor = -abs(valor_original)
            print(f"🔍 DEBUG: Despesa - valor convertido para negativo: {valor}")
        elif transacao_data.tipo == 'receita':
            # Receitas SEMPRE positivas no banco
            valor = abs(valor_original)
            print(f"🔍 DEBUG: Receita - valor convertido para positivo: {valor}")
        else:
            valor = valor_original
            print(f"🔍 DEBUG: Tipo neutro - mantendo valor original: {valor}")

        # Validação final antes de salvar
        if transacao_data.tipo == 'despesa' and valor > 0:
            print(f"❌ ERRO: Despesa com valor positivo detectada! Corrigindo...")
            valor = -abs(valor)
        elif transacao_data.tipo == 'receita' and valor < 0:
            print(f"❌ ERRO: Receita com valor negativo detectada! Corrigindo...")
            valor = abs(valor)

        print(f"🔍 DEBUG: Valor FINAL para salvar no banco: {valor}")

        # Mapear forma de pagamento para valores válidos do enum
        forma_pgto_mapping = {
            'transferencia_bancaria': 'transferencia',
            'transferencia': 'transferencia',
            'cartao_credito': 'cartao',
            'cartao': 'cartao',
            'cartao_debito': 'debito',
            'debito': 'debito',
            'dinheiro': 'dinheiro',
            'pix': 'pix',
            'boleto': 'boleto',
            'cheque': 'outros',
            'outros': 'outros',
            'ted': 'transferencia',
            'doc': 'transferencia'
        }

        forma_pgto_original = getattr(transacao_data, 'forma_recebimento', None) or getattr(transacao_data, 'forma_pgto', None)
        forma_pgto_final = None
        if forma_pgto_original:
            forma_pgto_final = forma_pgto_mapping.get(forma_pgto_original, 'outros')

        # VALIDAÇÃO PREVENTIVA: Verificar propriedade e consistência de todas as chaves estrangeiras
        from app.services.validation import validate_foreign_keys_ownership
        validation_result = validate_foreign_keys_ownership(
            db=db,
            empresa_id=empresa_id,
            cliente_id=getattr(transacao_data, 'cliente_id', None),
            fornecedor_id=getattr(transacao_data, 'fornecedor_id', None),
            centro_custo_id=getattr(transacao_data, 'centro_custo_id', None),
            conta_contabil_id=getattr(transacao_data, 'conta_contabil_id', None),
            projeto_id=getattr(transacao_data, 'projeto_id', None),
            produto_servico_id=getattr(transacao_data, 'produto_servico_id', None),
            categoria_contabil_id=transacao_data.categoria_contabil_id,
            subcategoria_contabil_id=getattr(transacao_data, 'subcategoria_contabil_id', None),
            categoria_gerencial_id=transacao_data.categoria_gerencial_id,
            subcategoria_gerencial_id=getattr(transacao_data, 'subcategoria_gerencial_id', None)
        )

        if not validation_result["valid"]:
            raise HTTPException(
                status_code=400,
                detail=f"Erro de validação: {'; '.join(validation_result['errors'])}"
            )

        # Processar competências contábil e gerencial
        comp_ano_contabil = data_lancamento.year
        comp_mes_contabil = data_lancamento.month
        comp_ano_gerencial = data_lancamento.year
        comp_mes_gerencial = data_lancamento.month

        # Se competencia_contabil foi fornecida (formato: YYYY-MM)
        if hasattr(transacao_data, 'competencia_contabil') and transacao_data.competencia_contabil:
            try:
                parts = transacao_data.competencia_contabil.split('-')
                comp_ano_contabil = int(parts[0])
                comp_mes_contabil = int(parts[1])
            except:
                pass  # Usar data_lancamento como fallback

        # Se competencia_gerencial foi fornecida (formato: YYYY-MM)
        if hasattr(transacao_data, 'competencia_gerencial') and transacao_data.competencia_gerencial:
            try:
                parts = transacao_data.competencia_gerencial.split('-')
                comp_ano_gerencial = int(parts[0])
                comp_mes_gerencial = int(parts[1])
            except:
                pass  # Usar data_lancamento como fallback

        # Processar data de vencimento (usar do request ou fallback para data_lancamento)
        data_vencimento = data_lancamento
        if hasattr(transacao_data, 'data_vencimento') and transacao_data.data_vencimento:
            try:
                dv = transacao_data.data_vencimento
                if 'T' in dv:
                    data_vencimento = datetime.fromisoformat(dv.split('T')[0]).date()
                else:
                    data_vencimento = datetime.fromisoformat(dv).date()
            except:
                data_vencimento = data_lancamento

        # Criar transação (usando empresa_id validado, não do cliente)
        transacao = TransacaoFinanceira(
            empresa_id=empresa_id,
            tipo=transacao_data.tipo,
            nome=transacao_data.titulo or transacao_data.descricao or 'Receita',
            descricao=transacao_data.descricao or transacao_data.titulo or 'Receita',
            valor=valor,
            data_lancamento=data_lancamento,
            data_vencimento=data_vencimento,
            competencia_ano=data_lancamento.year,
            competencia_mes=data_lancamento.month,
            competencia_ano_contabil=comp_ano_contabil,
            competencia_mes_contabil=comp_mes_contabil,
            competencia_ano_gerencial=comp_ano_gerencial,
            competencia_mes_gerencial=comp_mes_gerencial,
            status=transacao_data.status or 'pendente',
            forma_pgto=forma_pgto_final,
            cliente_id=transacao_data.cliente_id,
            fornecedor_id=transacao_data.fornecedor_id,
            projeto_id=getattr(transacao_data, 'projeto_id', None),
            produto_servico_id=getattr(transacao_data, 'produto_servico_id', None),
            categoria_gerencial_id=transacao_data.categoria_gerencial_id,
            categoria_contabil_id=transacao_data.categoria_contabil_id,
            subcategoria_gerencial_id=getattr(transacao_data, 'subcategoria_gerencial_id', None),
            subcategoria_contabil_id=getattr(transacao_data, 'subcategoria_contabil_id', None),
            centro_custo_id=transacao_data.centro_custo_id,
            conta_contabil_id=getattr(transacao_data, 'conta_contabil_id', None),
            exibir_no_cash_control=getattr(transacao_data, 'exibir_no_cash_control', True),
            entra_no_gerencial=getattr(transacao_data, 'entra_no_gerencial', True),
            numero_nota_fiscal=getattr(transacao_data, 'numero_nota_fiscal', None),
            link_nota_fiscal=getattr(transacao_data, 'link_nota_fiscal', None),
            numero_pedido_compra=getattr(transacao_data, 'numero_pedido_compra', None),
            link_pedido_compra=getattr(transacao_data, 'link_pedido_compra', None),
            valor_recebido=getattr(transacao_data, 'valor_recebido', None),
            valor_pago=getattr(transacao_data, 'valor_pago', None)
        )

        db.add(transacao)
        db.flush()  # Obter ID sem commit final

        # =========================================================================
        # RETENÇÃO NA FONTE: Criar transações derivadas (Múltiplos Impostos)
        # =========================================================================
        transacoes_criadas = [transacao]
        retencao_info = None
        
        if hasattr(transacao_data, 'retencao') and transacao_data.retencao:
            retencao = transacao_data.retencao
            impostos = retencao.get('impostos', [])
            
            # Compatibilidade: Se vier formato antigo (single imposto), converter para lista
            if not impostos and retencao.get('imposto_id'):
                impostos = [retencao]
            
            if impostos:
                valor_bruto = abs(float(transacao_data.valor))
                total_retido = sum(float(imp.get('valor_retido', 0)) for imp in impostos)
                valor_liquido = valor_bruto - total_retido
                
                if total_retido > 0 and valor_liquido > 0:
                    # Marcar transação original como pai com VALOR BRUTO para P&L consolidado
                    # O pai entra no P&L com valor bruto, os filhos entram em Contas a Pagar e Retenção na Fonte
                    transacao.valor = -abs(valor_bruto)  # Manter valor bruto para P&L consolidado
                    transacao.entra_no_gerencial = True  # Entra no P&L com valor bruto
                    transacao.exibir_no_cash_control = False  # Não exibe no Cash Control (filhos exibem)
                    nomes_impostos = ', '.join([imp.get('imposto_nome', 'Imposto') for imp in impostos])
                    transacao.descricao = f"[DESMEMBRADO] {transacao.descricao or transacao.nome} - Ver transações filhas"
                    
                    fornecedor_id = transacao_data.fornecedor_id
                    
                    # 1. Transação FILHO: Pagamento ao Fornecedor (valor líquido)
                    # NÃO entra no P&L (pai já tem valor bruto), mas entra em Contas a Pagar e Cash Control
                    transacao_fornecedor = TransacaoFinanceira(
                        empresa_id=empresa_id,
                        tipo=transacao.tipo,
                        nome=f"Pagamento - {transacao.nome or transacao.descricao}",
                        descricao=f"Valor líquido após retenção de {nomes_impostos}",
                        valor=-abs(valor_liquido),
                        data_lancamento=data_lancamento,
                        data_vencimento=transacao.data_vencimento,
                        competencia_ano=transacao.competencia_ano,
                        competencia_mes=transacao.competencia_mes,
                        competencia_ano_contabil=comp_ano_contabil,
                        competencia_mes_contabil=comp_mes_contabil,
                        competencia_ano_gerencial=comp_ano_gerencial,
                        competencia_mes_gerencial=comp_mes_gerencial,
                        status=transacao.status,
                        forma_pgto=transacao.forma_pgto,
                        cliente_id=transacao_data.cliente_id,
                        fornecedor_id=fornecedor_id,
                        categoria_gerencial_id=transacao_data.categoria_gerencial_id,
                        categoria_contabil_id=transacao_data.categoria_contabil_id,
                        subcategoria_gerencial_id=getattr(transacao_data, 'subcategoria_gerencial_id', None),
                        subcategoria_contabil_id=getattr(transacao_data, 'subcategoria_contabil_id', None),
                        centro_custo_id=transacao_data.centro_custo_id,
                        conta_contabil_id=getattr(transacao_data, 'conta_contabil_id', None),
                        exibir_no_cash_control=getattr(transacao_data, 'exibir_no_cash_control', True),
                        entra_no_gerencial=False,  # NÃO entra no P&L (pai tem valor bruto consolidado)
                        parent_id=transacao.id,
                        tipo_filho='split'
                    )
                    db.add(transacao_fornecedor)
                    db.flush()
                    transacoes_criadas.append(transacao_fornecedor)
                    
                    # 2. Transações FILHO: Uma para cada imposto retido
                    impostos_detalhes = []
                    for imp in impostos:
                        imposto_nome = imp.get('imposto_nome', 'Imposto')
                        valor_retido = float(imp.get('valor_retido', 0))
                        aliquota = float(imp.get('aliquota', 0))
                        
                        if valor_retido > 0:
                            transacao_imposto = TransacaoFinanceira(
                                empresa_id=empresa_id,
                                tipo=transacao.tipo,
                                nome=f"Retenção {imposto_nome}",
                                descricao=f"Imposto retido na fonte - {transacao.nome or transacao.descricao}",
                                valor=-abs(valor_retido),
                                data_lancamento=data_lancamento,
                                data_vencimento=transacao.data_vencimento,
                                competencia_ano=transacao.competencia_ano,
                                competencia_mes=transacao.competencia_mes,
                                competencia_ano_contabil=comp_ano_contabil,
                                competencia_mes_contabil=comp_mes_contabil,
                                competencia_ano_gerencial=comp_ano_gerencial,
                                competencia_mes_gerencial=comp_mes_gerencial,
                                status='pendente',
                                forma_pgto=None,
                                cliente_id=None,
                                fornecedor_id=None,
                                categoria_gerencial_id=None,
                                categoria_contabil_id=transacao_data.categoria_contabil_id,
                                exibir_no_cash_control=True,
                                entra_no_gerencial=False,
                                parent_id=transacao.id,
                                tipo_filho='split'
                            )
                            db.add(transacao_imposto)
                            db.flush()
                            transacoes_criadas.append(transacao_imposto)
                            
                            impostos_detalhes.append({
                                "imposto_nome": imposto_nome,
                                "aliquota": aliquota,
                                "valor_retido": valor_retido
                            })
                    
                    retencao_info = {
                        "impostos": impostos_detalhes,
                        "valor_bruto": valor_bruto,
                        "total_retido": total_retido,
                        "valor_liquido": valor_liquido
                    }
                    
                    print(f"✅ Retenção na fonte aplicada: Bruto={valor_bruto}, Retido={total_retido}, Líquido={valor_liquido}, Impostos={len(impostos_detalhes)}")

        db.commit()
        for t in transacoes_criadas:
            db.refresh(t)

        response = {
            "id": transacao.id,
            "empresa_id": transacao.empresa_id,
            "tipo": transacao.tipo,
            "nome": transacao.nome,
            "descricao": transacao.descricao,
            "valor": float(transacao.valor),
            "data_lancamento": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
            "status": transacao.status,
            "message": "Transação criada com sucesso!"
        }
        
        if retencao_info:
            response["retencao"] = retencao_info
            response["transacoes_derivadas"] = len(transacoes_criadas) - 1
            num_impostos = len(retencao_info.get('impostos', []))
            nomes = ', '.join([imp['imposto_nome'] for imp in retencao_info.get('impostos', [])])
            response["message"] = f"Despesa criada com retenção de {nomes}! Geradas {len(transacoes_criadas)} transações."
        
        return response

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Erro ao criar transação: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Erro interno do servidor")


@router.get("/transacoes")
async def list_transacoes(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    tipo: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    mes: Optional[int] = Query(None, ge=1, le=12),
    ano: Optional[int] = Query(None, ge=2000, le=2100),
    busca: Optional[str] = Query(None),
    data_inicio: Optional[date] = Query(None),
    data_fim: Optional[date] = Query(None),
    tipo_data: Optional[str] = Query("competencia", description="Tipo de data para filtro: 'competencia', 'competencia_gerencial' ou 'lancamento'")
):
    """
    Lista transações com filtros
    Compatível com: GET /api/transacoes
    Suporta filtro por competência contábil (padrão), gerencial ou por data de lançamento
    """
    try:
        from sqlalchemy import extract
        
        # Query base - TODAS as empresas (comportamento padrão)
        query = db.query(TransacaoFinanceira)

        # Aplicar filtros de período baseado no tipo de data selecionado
        if tipo_data == "lancamento":
            if ano:
                query = query.filter(extract('year', TransacaoFinanceira.data_lancamento) == ano)
            if mes:
                query = query.filter(extract('month', TransacaoFinanceira.data_lancamento) == mes)
            periodo_str = f"mês {mes}/{ano} (por lançamento)" if mes and ano else f"ano {ano} (por lançamento)" if ano else "TODAS as empresas"
        elif tipo_data == "competencia_gerencial":
            if ano:
                query = query.filter(TransacaoFinanceira.competencia_ano_gerencial == ano)
            if mes:
                query = query.filter(TransacaoFinanceira.competencia_mes_gerencial == mes)
            periodo_str = f"mês {mes}/{ano} (por comp. gerencial)" if mes and ano else f"ano {ano} (por comp. gerencial)" if ano else "TODAS as empresas"
        else:
            if ano:
                query = query.filter(TransacaoFinanceira.competencia_ano == ano)
            if mes:
                query = query.filter(TransacaoFinanceira.competencia_mes == mes)
            periodo_str = f"mês {mes}/{ano} (por competência)" if mes and ano else f"ano {ano} (por competência)" if ano else "TODAS as empresas"
        
        print(f"🔍 Listagem: Mostrando transações de {periodo_str}")

        # Aplicar filtros adicionais
        if tipo:
            query = query.filter(TransacaoFinanceira.tipo == tipo)
        if status:
            query = query.filter(TransacaoFinanceira.status == status)
        
        if busca:
            # Busca textual em nome, título breve e descrição
            search_term = f"%{busca}%"
            query = query.filter(
                or_(
                    TransacaoFinanceira.nome.ilike(search_term),
                    TransacaoFinanceira.titulo_breve.ilike(search_term),
                    TransacaoFinanceira.descricao.ilike(search_term)
                )
            )

        if data_inicio:
            query = query.filter(TransacaoFinanceira.data_lancamento >= data_inicio)

        if data_fim:
            query = query.filter(TransacaoFinanceira.data_lancamento <= data_fim)

        # 🌳 HIERARQUIA: Filtrar apenas leaf nodes antes de ordenar
        query = apply_leaf_nodes_filter(query, db)
        
        # Ordenação padrão por data mais recente
        query = query.order_by(TransacaoFinanceira.data_lancamento.desc())

        # Paginação
        page = page or 1
        limit = limit or 10
        offset = (page - 1) * limit

        # Buscar transações com tratamento de erro
        try:
            transacoes_data = query.offset(offset).limit(limit).all()
            # Para o total, usar a query básica sem os JOINs
            total_query = db.query(TransacaoFinanceira)

            # Aplicar filtros de período ao total baseado no tipo de data
            if tipo_data == "lancamento":
                if ano:
                    total_query = total_query.filter(extract('year', TransacaoFinanceira.data_lancamento) == ano)
                if mes:
                    total_query = total_query.filter(extract('month', TransacaoFinanceira.data_lancamento) == mes)
            elif tipo_data == "competencia_gerencial":
                if ano:
                    total_query = total_query.filter(TransacaoFinanceira.competencia_ano_gerencial == ano)
                if mes:
                    total_query = total_query.filter(TransacaoFinanceira.competencia_mes_gerencial == mes)
            else:
                if ano:
                    total_query = total_query.filter(TransacaoFinanceira.competencia_ano == ano)
                if mes:
                    total_query = total_query.filter(TransacaoFinanceira.competencia_mes == mes)
            
            if tipo:
                total_query = total_query.filter(TransacaoFinanceira.tipo == tipo)
            if data_inicio:
                total_query = total_query.filter(TransacaoFinanceira.data_lancamento >= data_inicio)
            if data_fim:
                total_query = total_query.filter(TransacaoFinanceira.data_lancamento <= data_fim)
            
            # 🌳 HIERARQUIA: Aplicar filtro de leaf nodes ao total
            total_query = apply_leaf_nodes_filter(total_query, db)
            total = total_query.count()
            print(f"🔍 DEBUG: Total de transações (leaf nodes) de {periodo_str}: {total}")
        except Exception as db_error:
            print(f"Erro na consulta ao banco: {db_error}")
            transacoes_data = []
            total = 0

        # Formato compatível com frontend
        result = []
        for t in transacoes_data:
            try:
                # Buscar nomes de cliente e fornecedor separadamente quando existirem
                cliente_nome = None
                fornecedor_nome = None

                if t.cliente_id:
                    # ✅ BUSCAR cliente sem filtro de empresa (configuração padrão)
                    cliente = db.query(Cliente).filter(Cliente.id == t.cliente_id).first()
                    if cliente:
                        cliente_nome = cliente.nome

                if t.fornecedor_id:
                    # ✅ BUSCAR fornecedor sem filtro de empresa (configuração padrão)
                    fornecedor = db.query(Fornecedor).filter(Fornecedor.id == t.fornecedor_id).first()
                    if fornecedor:
                        fornecedor_nome = fornecedor.nome

                data_formatada = None
                if t.data_lancamento is not None:
                    data_formatada = t.data_lancamento.isoformat()

                created_formatado = None
                if t.created_at is not None:
                    created_formatado = t.created_at.isoformat()

                valor_numerico = 0.0
                if t.valor is not None:
                    valor_numerico = float(t.valor)

                # Usar o tipo definido no banco de dados
                tipo_real = t.tipo
                if not tipo_real:
                    # Só determinar por valor se o tipo não estiver definido
                    if valor_numerico > 0:
                        tipo_real = 'receita'
                    elif valor_numerico < 0:
                        tipo_real = 'despesa'
                    else:
                        tipo_real = 'neutro'

                # Determinar cliente/fornecedor baseado no tipo (manter compatibilidade)
                cliente_fornecedor = "Não informado"
                if tipo_real == 'receita' and cliente_nome:
                    cliente_fornecedor = cliente_nome
                elif tipo_real == 'despesa' and fornecedor_nome:
                    cliente_fornecedor = fornecedor_nome
                elif cliente_nome:
                    cliente_fornecedor = cliente_nome
                elif fornecedor_nome:
                    cliente_fornecedor = fornecedor_nome

                result.append({
                    "id": t.id,
                    "empresa_id": t.empresa_id,
                    "tipo": tipo_real,
                    "data_lancamento": data_formatada,
                    "data_transacao": data_formatada,
                    "data": data_formatada,
                    "competencia_ano": t.competencia_ano,
                    "competencia_mes": t.competencia_mes,
                    "competencia_ano_contabil": t.competencia_ano_contabil,
                    "competencia_mes_contabil": t.competencia_mes_contabil,
                    "competencia_ano_gerencial": t.competencia_ano_gerencial,
                    "competencia_mes_gerencial": t.competencia_mes_gerencial,
                    "nome": t.nome,
                    "descricao": t.descricao,
                    "valor": valor_numerico,
                    "status": t.status or 'Pendente',
                    "forma_pgto": t.forma_pgto or '-',
                    "categoria": 'Geral',  # TODO: JOIN com categorias quando necessário
                    "cliente_id": t.cliente_id,
                    "cliente_nome": cliente_nome,
                    "fornecedor_id": t.fornecedor_id,
                    "fornecedor_nome": fornecedor_nome,
                    "cliente_fornecedor": cliente_fornecedor,
                    "conta_bancaria": None,
                    "produto_servico": None,
                    "created_at": created_formatado,
                    "parent_id": t.parent_id,
                    "tipo_filho": t.tipo_filho
                })
            except Exception as row_error:
                print(f"Erro ao processar transação {getattr(t, 'id', 'unknown')}: {row_error}")
                continue

        # Calcular páginas totais de forma segura
        total_pages = 1
        if total > 0 and limit > 0:
            total_pages = (total + limit - 1) // limit

        return {
            "transacoes": result,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": total_pages
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao buscar transações: {str(e)}")



@router.get("/transacoes/data-integrity-check")
async def check_data_integrity(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Verificar integridade dos dados - categorias órfãs e subcategorias não mostradas na interface
    """
    try:
        # 1. Verificar transações com categorias que não existem mais (órfãs) - Todas as empresas
        orphan_categories = db.query(
            TransacaoFinanceira.categoria_contabil_id,
            func.count(TransacaoFinanceira.id).label('count')
        ).outerjoin(
            CategoriaContabil,
            TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        ).filter(
            TransacaoFinanceira.categoria_contabil_id.isnot(None),
            CategoriaContabil.id.is_(None)
        ).group_by(TransacaoFinanceira.categoria_contabil_id).all()

        # 2. Verificar categorias que existem mas não aparecem na interface padrão (subcategorias) - Todas as empresas
        hidden_categories = db.query(
            CategoriaContabil.id,
            CategoriaContabil.nome,
            CategoriaContabil.codigo,
            CategoriaContabil.pai_id,
            func.count(TransacaoFinanceira.id).label('transaction_count')
        ).outerjoin(
            TransacaoFinanceira,
            CategoriaContabil.id == TransacaoFinanceira.categoria_contabil_id
        ).filter(
            CategoriaContabil.ativo == True,
            CategoriaContabil.pai_id.isnot(None)  # Subcategorias (não aparecem na listagem padrão)
        ).group_by(
            CategoriaContabil.id,
            CategoriaContabil.nome,
            CategoriaContabil.codigo,
            CategoriaContabil.pai_id
        ).having(func.count(TransacaoFinanceira.id) > 0).all()

        return {
            "orphan_categories": [
                {
                    "categoria_id": item.categoria_contabil_id,
                    "count": item.count,
                    "status": "orphan",
                    "message": f"Categoria ID {item.categoria_contabil_id} não existe mais no banco"
                }
                for item in orphan_categories
            ],
            "hidden_categories": [
                {
                    "categoria_id": item.id,
                    "categoria_nome": item.nome,
                    "categoria_codigo": item.codigo,
                    "pai_id": item.pai_id,
                    "transaction_count": item.transaction_count,
                    "status": "hidden",
                    "message": f"Subcategoria '{item.nome}' existe mas não aparece na listagem padrão da interface"
                }
                for item in hidden_categories
            ]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao verificar integridade: {str(e)}")

@router.get("/transacoes/categorization-summary")
async def get_categorization_summary(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Resumo das transações e categorizações pendentes
    """
    try:
        # ✅ ATUALIZADO: Mostrar dados de TODAS as empresas (não filtrar por empresa)
        # Separação por empresa será apenas opção de relatório
        # 🌳 HIERARQUIA: Contar apenas LEAF NODES

        # Contar totais de TODAS as empresas - APENAS LEAF NODES
        total_transacoes = apply_leaf_nodes_filter(db.query(func.count(TransacaoFinanceira.id)), db).scalar() or 0
        total_receitas = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(TransacaoFinanceira.tipo == 'receita'),
            db
        ).scalar() or 0
        total_despesas = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(TransacaoFinanceira.tipo == 'despesa'),
            db
        ).scalar() or 0

        # 🌳 Contar despesas sem categorização (APENAS LEAF NODES)
        despesas_sem_categoria_contabil = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        despesas_sem_categoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        despesas_sem_centro_custo = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.centro_custo_id.is_(None)
            ),
            db
        ).scalar() or 0

        despesas_sem_conta_contabil = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.conta_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Contar despesas sem subcategorias (APENAS LEAF NODES)
        despesas_sem_subcategoria_contabil = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        despesas_sem_subcategoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Contar receitas sem categorização (APENAS LEAF NODES)
        receitas_sem_cliente = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.cliente_id.is_(None)
            ),
            db
        ).scalar() or 0

        receitas_sem_projeto = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.projeto_id.is_(None)
            ),
            db
        ).scalar() or 0

        receitas_sem_produto_servico = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.produto_servico_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Calcular despesas totalmente categorizadas (APENAS LEAF NODES)
        # 4 atributos: Categoria Gerencial, Subcategoria Gerencial, Centro de Custo, Categoria Contábil
        despesas_completas = apply_leaf_nodes_filter(
            db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_gerencial_id.is_not(None),
                TransacaoFinanceira.subcategoria_gerencial_id.is_not(None),
                TransacaoFinanceira.centro_custo_id.is_not(None),
                TransacaoFinanceira.categoria_contabil_id.is_not(None)
            ),
            db
        ).count()

        # 🌳 Calcular receitas totalmente categorizadas (APENAS LEAF NODES)
        receitas_completas = apply_leaf_nodes_filter(
            db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.cliente_id.is_not(None),
                TransacaoFinanceira.projeto_id.is_not(None),
                TransacaoFinanceira.produto_servico_id.is_not(None)
            ),
            db
        ).count()

        # Total de transações completas e incompletas
        total_completas = despesas_completas + receitas_completas
        total_incompletas = total_transacoes - total_completas

        # 🌳 Calcular valores totais das transações (APENAS LEAF NODES)
        # 🔧 CORREÇÃO: Usar valores absolutos para despesas nos cards de categorização
        valor_total_receitas = apply_leaf_nodes_filter(
            db.query(func.sum(TransacaoFinanceira.valor)).filter(TransacaoFinanceira.tipo == 'receita'),
            db
        ).scalar() or 0

        # Para despesas, usar func.abs() para garantir valores positivos no card
        valor_total_despesas = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(TransacaoFinanceira.tipo == 'despesa'),
            db
        ).scalar() or 0

        # 🌳 Calcular valores das pendências de despesas (APENAS LEAF NODES)
        # 🔧 CORREÇÃO: Garantir que todos os valores de despesas sejam sempre positivos nos cards
        valor_despesas_sem_categoria_contabil = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_despesas_sem_categoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_despesas_sem_centro_custo = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.centro_custo_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_despesas_sem_conta_contabil = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.conta_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Valores das subcategorias em falta (APENAS LEAF NODES)
        valor_despesas_sem_subcategoria_contabil = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_despesas_sem_subcategoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Calcular valores das pendências de receitas (APENAS LEAF NODES)
        valor_receitas_sem_cliente = apply_leaf_nodes_filter(
            db.query(func.sum(TransacaoFinanceira.valor)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.cliente_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_receitas_sem_projeto = apply_leaf_nodes_filter(
            db.query(func.sum(TransacaoFinanceira.valor)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.projeto_id.is_(None)
            ),
            db
        ).scalar() or 0

        valor_receitas_sem_produto_servico = apply_leaf_nodes_filter(
            db.query(func.sum(TransacaoFinanceira.valor)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.produto_servico_id.is_(None)
            ),
            db
        ).scalar() or 0

        # 🌳 Calcular valores das transações completas (APENAS LEAF NODES)
        # 4 atributos: Categoria Gerencial, Subcategoria Gerencial, Centro de Custo, Categoria Contábil
        valor_despesas_completas = apply_leaf_nodes_filter(
            db.query(func.sum(func.abs(TransacaoFinanceira.valor))).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_gerencial_id.is_not(None),
                TransacaoFinanceira.subcategoria_gerencial_id.is_not(None),
                TransacaoFinanceira.centro_custo_id.is_not(None),
                TransacaoFinanceira.categoria_contabil_id.is_not(None)
            ),
            db
        ).scalar() or 0

        valor_receitas_completas = apply_leaf_nodes_filter(
            db.query(func.sum(TransacaoFinanceira.valor)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.cliente_id.is_not(None),
                TransacaoFinanceira.projeto_id.is_not(None),
                TransacaoFinanceira.produto_servico_id.is_not(None)
            ),
            db
        ).scalar() or 0

        return {
            "total_transacoes": total_transacoes,
            "total_receitas": total_receitas,
            "total_despesas": total_despesas,
            "valor_total_receitas": float(valor_total_receitas),
            "valor_total_despesas": float(abs(valor_total_despesas)),  # 🔧 CORREÇÃO: Garantir valor absoluto sempre positivo
            "total_completas": total_completas,
            "total_incompletas": total_incompletas,
            "despesas_completas": despesas_completas,
            "receitas_completas": receitas_completas,
            "valor_despesas_completas": float(valor_despesas_completas),
            "valor_receitas_completas": float(valor_receitas_completas),

            # Pendências de despesas - quantitativos e valores
            "despesas_sem_categoria_contabil": despesas_sem_categoria_contabil,
            "valor_despesas_sem_categoria_contabil": float(valor_despesas_sem_categoria_contabil),
            "despesas_sem_categoria_gerencial": despesas_sem_categoria_gerencial,
            "valor_despesas_sem_categoria_gerencial": float(valor_despesas_sem_categoria_gerencial),
            "despesas_sem_centro_custo": despesas_sem_centro_custo,
            "valor_despesas_sem_centro_custo": float(valor_despesas_sem_centro_custo),
            "despesas_sem_conta_contabil": despesas_sem_conta_contabil,
            "valor_despesas_sem_conta_contabil": float(valor_despesas_sem_conta_contabil),
            "despesas_sem_subcategoria_contabil": despesas_sem_subcategoria_contabil,
            "valor_despesas_sem_subcategoria_contabil": float(valor_despesas_sem_subcategoria_contabil),
            "despesas_sem_subcategoria_gerencial": despesas_sem_subcategoria_gerencial,
            "valor_despesas_sem_subcategoria_gerencial": float(valor_despesas_sem_subcategoria_gerencial),

            # Pendências de receitas - quantitativos e valores
            "receitas_sem_cliente": receitas_sem_cliente,
            "valor_receitas_sem_cliente": float(valor_receitas_sem_cliente),
            "receitas_sem_projeto": receitas_sem_projeto,
            "valor_receitas_sem_projeto": float(valor_receitas_sem_projeto),
            "receitas_sem_produto_servico": receitas_sem_produto_servico,
            "valor_receitas_sem_produto_servico": float(valor_receitas_sem_produto_servico)
        }

    except Exception as e:
        print(f"Erro ao buscar resumo de categorização: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")



@router.get("/transacoes/despesas-categorization")
async def get_despesas_for_categorization(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    filter_type: Optional[str] = Query(None)
):
    """
    Endpoint SIMPLIFICADO para buscar despesas - usa mesma estrutura que summary
    """
    print(f"🔍 DEBUG: Iniciando despesas-categorization, filter_type={filter_type}")
    try:
        # ✅ CORREÇÃO: Incluir nomes das categorias para JavaScript filtrar corretamente
        from app.models.categorias import CategoriaContabil, CategoriaGerencial, CentroCusto
        from app.models.auxiliares import ContaContabil

        # Criar aliases para subcategorias (usando a mesma tabela com pai_id)
        from sqlalchemy.orm import aliased
        SubcategoriaContabil = aliased(CategoriaContabil)
        SubcategoriaGerencial = aliased(CategoriaGerencial)

        # Query com JOINs para obter nomes das categorias
        query = db.query(
            TransacaoFinanceira,
            CategoriaContabil.nome.label('categoria_contabil_nome'),
            CategoriaGerencial.nome.label('categoria_gerencial_nome'),
            SubcategoriaContabil.nome.label('subcategoria_contabil_nome'),
            SubcategoriaGerencial.nome.label('subcategoria_gerencial_nome'),
            CentroCusto.nome.label('centro_custo_nome'),
            ContaContabil.nome.label('conta_contabil_nome')
        ).outerjoin(
            CategoriaContabil, TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        ).outerjoin(
            CategoriaGerencial, TransacaoFinanceira.categoria_gerencial_id == CategoriaGerencial.id
        ).outerjoin(
            SubcategoriaContabil, TransacaoFinanceira.subcategoria_contabil_id == SubcategoriaContabil.id
        ).outerjoin(
            SubcategoriaGerencial, TransacaoFinanceira.subcategoria_gerencial_id == SubcategoriaGerencial.id
        ).outerjoin(
            CentroCusto, TransacaoFinanceira.centro_custo_id == CentroCusto.id
        ).outerjoin(
            ContaContabil, TransacaoFinanceira.conta_contabil_id == ContaContabil.id
        ).filter(
            TransacaoFinanceira.tipo == 'despesa'
        )

        # 🌳 HIERARQUIA: Aplicar filtro de leaf nodes para consistência com summary
        from sqlalchemy import select
        parents_with_children_subquery = select(TransacaoFinanceira.parent_id).where(
            TransacaoFinanceira.parent_id.isnot(None)
        ).distinct()
        query = query.filter(~TransacaoFinanceira.id.in_(parents_with_children_subquery))

        # Executar query com JOINs
        despesas_result = query.all()

        print(f"✅ DEBUG: Encontradas {len(despesas_result)} despesas de todas as empresas")

        # Processar resultados incluindo nomes das categorias
        despesas_data = []
        for result in despesas_result:
            despesa = result[0]  # TransacaoFinanceira object
            despesa_dict = {
                "id": despesa.id,
                "data_lancamento": despesa.data_lancamento.isoformat() if despesa.data_lancamento else None,
                "nome": despesa.nome,
                "descricao": despesa.descricao,
                "valor": float(despesa.valor) if despesa.valor else 0,
                "status": despesa.status,
                "categoria_contabil_id": despesa.categoria_contabil_id,
                "subcategoria_contabil_id": despesa.subcategoria_contabil_id,
                "categoria_gerencial_id": despesa.categoria_gerencial_id,
                "subcategoria_gerencial_id": despesa.subcategoria_gerencial_id,
                "centro_custo_id": despesa.centro_custo_id,
                "conta_contabil_id": despesa.conta_contabil_id,
                # ✅ NOMES necessários para filtros JavaScript
                "categoria_contabil_nome": result.categoria_contabil_nome,
                "categoria_gerencial_nome": result.categoria_gerencial_nome,
                "subcategoria_contabil_nome": result.subcategoria_contabil_nome,
                "subcategoria_gerencial_nome": result.subcategoria_gerencial_nome,
                "centro_custo_nome": result.centro_custo_nome,
                "conta_contabil_nome": result.conta_contabil_nome
            }
            despesas_data.append(despesa_dict)

        print(f"✅ DEBUG: Processadas {len(despesas_data)} despesas com sucesso")
        return despesas_data

    except Exception as e:
        print(f"❌ ERRO ao buscar despesas: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.get("/transacoes/receitas-categorization")
async def get_receitas_for_categorization(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    filter_type: Optional[str] = Query(None),
    ano: Optional[int] = Query(None)
):
    """
    Buscar receitas para categorização
    """
    try:
        from app.models.categorias import CategoriaGerencial

        query = db.query(
            TransacaoFinanceira.id,
            TransacaoFinanceira.data_lancamento,
            TransacaoFinanceira.nome,
            TransacaoFinanceira.descricao,
            TransacaoFinanceira.valor,
            TransacaoFinanceira.status,
            TransacaoFinanceira.cliente_id,
            TransacaoFinanceira.projeto_id,
            TransacaoFinanceira.produto_servico_id,
            TransacaoFinanceira.categoria_gerencial_id,
            TransacaoFinanceira.subcategoria_gerencial_id,
            TransacaoFinanceira.categoria_contabil_id,
            TransacaoFinanceira.subcategoria_contabil_id,
            Cliente.nome.label('cliente_nome'),
            Projeto.nome.label('projeto_nome'),
            ProdutoServico.nome.label('produto_servico_nome'),
            CategoriaGerencial.nome.label('categoria_gerencial_nome')
        ).outerjoin(
            Cliente, TransacaoFinanceira.cliente_id == Cliente.id
        ).outerjoin(
            Projeto, TransacaoFinanceira.projeto_id == Projeto.id
        ).outerjoin(
            ProdutoServico, TransacaoFinanceira.produto_servico_id == ProdutoServico.id
        ).outerjoin(
            CategoriaGerencial, TransacaoFinanceira.categoria_gerencial_id == CategoriaGerencial.id
        ).filter(
            TransacaoFinanceira.tipo == 'receita'
        )

        # 🌳 HIERARQUIA: Aplicar filtro de leaf nodes para consistência com summary
        from sqlalchemy import select
        parents_with_children_subquery = select(TransacaoFinanceira.parent_id).where(
            TransacaoFinanceira.parent_id.isnot(None)
        ).distinct()
        query = query.filter(~TransacaoFinanceira.id.in_(parents_with_children_subquery))

        # Aplicar filtros
        if filter_type:
            if filter_type == 'cliente_null':
                query = query.filter(TransacaoFinanceira.cliente_id.is_(None))
            elif filter_type == 'projeto_null':
                query = query.filter(TransacaoFinanceira.projeto_id.is_(None))
            elif filter_type == 'produto_servico_null':
                query = query.filter(TransacaoFinanceira.produto_servico_id.is_(None))

        if ano:
            query = query.filter(TransacaoFinanceira.competencia_ano == ano)

        receitas = query.order_by(TransacaoFinanceira.data_lancamento.desc()).all()

        return [
            {
                "id": receita.id,
                "data_lancamento": receita.data_lancamento.isoformat() if receita.data_lancamento else None,
                "nome": receita.nome,
                "descricao": receita.descricao,
                "valor": float(receita.valor) if receita.valor else 0,
                "status": receita.status,
                "cliente_id": receita.cliente_id,
                "projeto_id": receita.projeto_id,
                "produto_servico_id": receita.produto_servico_id,
                "categoria_gerencial_id": receita.categoria_gerencial_id,
                "subcategoria_gerencial_id": receita.subcategoria_gerencial_id,
                "categoria_contabil_id": receita.categoria_contabil_id,
                "subcategoria_contabil_id": receita.subcategoria_contabil_id,
                "cliente_nome": receita.cliente_nome,
                "projeto_nome": receita.projeto_nome,
                "produto_servico_nome": receita.produto_servico_nome,
                "categoria_gerencial_nome": receita.categoria_gerencial_nome,
                "observacoes": getattr(receita, 'observacoes', None)
            }
            for receita in receitas
        ]

    except Exception as e:
        print(f"Erro ao buscar receitas para categorização: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.delete("/transacoes/{transacao_id}")
async def delete_transacao(
    transacao_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Deletar uma transação específica - com tratamento de dependências
    """
    try:
        print(f"🗑️ Solicitação de delete para transação {transacao_id}")

        # ✅ PERMISSÃO TOTAL: Usuário pode deletar qualquer transação independente da empresa
        transacao = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.id == transacao_id
        ).first()

        if not transacao:
            raise HTTPException(status_code=404, detail="Transação não encontrada")

        # 🔧 CORREÇÃO: Deletar todas as dependências de forma sistemática
        try:
            total_dependencias = 0

            # 1. Deletar registros de impostos
            from sqlalchemy import text
            impostos_count = db.execute(
                text("DELETE FROM transacao_impostos WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if impostos_count > 0:
                print(f"🔧 Deletados {impostos_count} registros de impostos")
                total_dependencias += impostos_count

            # 2. Deletar categorizações contábeis
            cat_contabil_count = db.execute(
                text("DELETE FROM transacao_categoria_contabil WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if cat_contabil_count > 0:
                print(f"🔧 Deletados {cat_contabil_count} registros de categoria contábil")
                total_dependencias += cat_contabil_count

            # 3. Deletar categorizações gerenciais
            cat_gerencial_count = db.execute(
                text("DELETE FROM transacao_categoria_gerencial WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if cat_gerencial_count > 0:
                print(f"🔧 Deletados {cat_gerencial_count} registros de categoria gerencial")
                total_dependencias += cat_gerencial_count

            # 4. Deletar mensalizações
            mensalizacao_count = db.execute(
                text("DELETE FROM transacao_mensalizacao WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if mensalizacao_count > 0:
                print(f"🔧 Deletados {mensalizacao_count} registros de mensalização")
                total_dependencias += mensalizacao_count

            # 5. Deletar desmembramentos onde é transação derivada
            desmembramento_derivada_count = db.execute(
                text("DELETE FROM desmembramentos_itens WHERE transacao_derivada_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if desmembramento_derivada_count > 0:
                print(f"🔧 Deletados {desmembramento_derivada_count} itens de desmembramento (derivada)")
                total_dependencias += desmembramento_derivada_count

            # 6. Deletar desmembramentos onde é transação origem
            desmembramento_origem_count = db.execute(
                text("DELETE FROM desmembramentos_transacoes WHERE transacao_origem_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if desmembramento_origem_count > 0:
                print(f"🔧 Deletados {desmembramento_origem_count} desmembramentos (origem)")
                total_dependencias += desmembramento_origem_count

            # 7. Verificar e deletar transações filhas (parent_id)
            filhas_count = db.execute(
                text("DELETE FROM transacoes_financeiras WHERE parent_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if filhas_count > 0:
                print(f"🔧 Deletadas {filhas_count} transações filhas")
                total_dependencias += filhas_count

            # 8. Agora deletar a transação principal
            db.delete(transacao)
            db.commit()

            print(f"✅ Transação {transacao_id} deletada com sucesso (incluindo {total_dependencias} dependências)")

            return {
                "success": True,
                "message": f"Transação deletada com sucesso{f' (incluindo {total_dependencias} dependências)' if total_dependencias else ''}",
                "id": transacao_id,
                "dependencias_removidas": total_dependencias
            }

        except Exception as delete_error:
            db.rollback()

            # Tratar especificamente o erro de Foreign Key
            error_str = str(delete_error).lower()
            if "foreign key" in error_str or "constraint" in error_str:
                # Extrair nome da tabela do erro se possível
                table_name = "desconhecida"
                if "desmembramentos_transacoes" in error_str:
                    table_name = "desmembramentos_transacoes"

                raise HTTPException(
                    status_code=409,
                    detail=f"Esta transação não pode ser deletada porque possui dependências na tabela '{table_name}'. "
                          f"Remova as dependências primeiro ou entre em contato com o suporte."
                )
            else:
                # Erro genérico
                raise HTTPException(status_code=500, detail=f"Erro ao deletar transação: {str(delete_error)}")

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico ao deletar transação {transacao_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.get("/transacoes/filter-options")
async def get_filter_options(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    column: str = Query(..., description="Nome da coluna para buscar opções únicas")
):
    """
    Endpoint otimizado para buscar valores únicos de uma coluna específica
    Usado pelos filtros de cabeçalho das tabelas
    """
    try:
        # Mapear nome da coluna para o campo do modelo
        column_mapping = {
            'data_transacao': TransacaoFinanceira.data_lancamento,
            'data_lancamento': TransacaoFinanceira.data_lancamento,
            'descricao': TransacaoFinanceira.descricao,
            'tipo': TransacaoFinanceira.tipo,
            'valor': TransacaoFinanceira.valor,
            'status': TransacaoFinanceira.status,
            'forma_pgto': TransacaoFinanceira.forma_pgto,
            'cliente_nome': Cliente.nome,
            'fornecedor_nome': Fornecedor.nome
        }
        
        if column not in column_mapping:
            raise HTTPException(status_code=400, detail=f"Coluna '{column}' não suportada")
        
        # Query otimizada com DISTINCT
        if column in ['cliente_nome', 'fornecedor_nome']:
            if column == 'cliente_nome':
                query = db.query(Cliente.nome).distinct().order_by(Cliente.nome)
            else:
                query = db.query(Fornecedor.nome).distinct().order_by(Fornecedor.nome)
        else:
            field = column_mapping[column]
            query = db.query(field).distinct().order_by(field)
        
        # Executar query e retornar apenas valores não nulos
        results = query.all()
        unique_values = [str(r[0]) for r in results if r[0] is not None]
        
        print(f"✅ Filtro otimizado '{column}': {len(unique_values)} opções únicas")
        
        return {
            "column": column,
            "options": unique_values,
            "count": len(unique_values)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Erro ao buscar opções de filtro: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")

        if not transacao:
            raise HTTPException(status_code=404, detail="Transação não encontrada")

        # 🔧 CORREÇÃO: Deletar todas as dependências de forma sistemática
        try:
            total_dependencias = 0

            # 1. Deletar registros de impostos
            from sqlalchemy import text
            impostos_count = db.execute(
                text("DELETE FROM transacao_impostos WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if impostos_count > 0:
                print(f"🔧 Deletados {impostos_count} registros de impostos")
                total_dependencias += impostos_count

            # 2. Deletar categorizações contábeis
            cat_contabil_count = db.execute(
                text("DELETE FROM transacao_categoria_contabil WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if cat_contabil_count > 0:
                print(f"🔧 Deletados {cat_contabil_count} registros de categoria contábil")
                total_dependencias += cat_contabil_count

            # 3. Deletar categorizações gerenciais
            cat_gerencial_count = db.execute(
                text("DELETE FROM transacao_categoria_gerencial WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if cat_gerencial_count > 0:
                print(f"🔧 Deletados {cat_gerencial_count} registros de categoria gerencial")
                total_dependencias += cat_gerencial_count

            # 4. Deletar mensalizações
            mensalizacao_count = db.execute(
                text("DELETE FROM transacao_mensalizacao WHERE transacao_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if mensalizacao_count > 0:
                print(f"🔧 Deletados {mensalizacao_count} registros de mensalização")
                total_dependencias += mensalizacao_count

            # 5. Deletar desmembramentos onde é transação derivada
            desmembramento_derivada_count = db.execute(
                text("DELETE FROM desmembramentos_itens WHERE transacao_derivada_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if desmembramento_derivada_count > 0:
                print(f"🔧 Deletados {desmembramento_derivada_count} itens de desmembramento (derivada)")
                total_dependencias += desmembramento_derivada_count

            # 6. Deletar desmembramentos onde é transação origem
            desmembramento_origem_count = db.execute(
                text("DELETE FROM desmembramentos_transacoes WHERE transacao_origem_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if desmembramento_origem_count > 0:
                print(f"🔧 Deletados {desmembramento_origem_count} desmembramentos (origem)")
                total_dependencias += desmembramento_origem_count

            # 7. Verificar e deletar transações filhas (parent_id)
            filhas_count = db.execute(
                text("DELETE FROM transacoes_financeiras WHERE parent_id = :transacao_id"),
                {"transacao_id": transacao_id}
            ).rowcount
            if filhas_count > 0:
                print(f"🔧 Deletadas {filhas_count} transações filhas")
                total_dependencias += filhas_count

            # 8. Agora deletar a transação principal
            db.delete(transacao)
            db.commit()

            print(f"✅ Transação {transacao_id} deletada com sucesso (incluindo {total_dependencias} dependências)")

            return {
                "success": True,
                "message": f"Transação deletada com sucesso{f' (incluindo {total_dependencias} dependências)' if total_dependencias else ''}",
                "id": transacao_id,
                "dependencias_removidas": total_dependencias
            }

        except Exception as delete_error:
            db.rollback()

            # Tratar especificamente o erro de Foreign Key
            error_str = str(delete_error).lower()
            if "foreign key" in error_str or "constraint" in error_str:
                # Extrair nome da tabela do erro se possível
                table_name = "desconhecida"
                if "desmembramentos_transacoes" in error_str:
                    table_name = "desmembramentos_transacoes"

                raise HTTPException(
                    status_code=409,
                    detail=f"Esta transação não pode ser deletada porque possui dependências na tabela '{table_name}'. "
                          f"Remova as dependências primeiro ou entre em contato com o suporte."
                )
            else:
                # Erro genérico
                raise HTTPException(status_code=500, detail=f"Erro ao deletar transação: {str(delete_error)}")

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico ao deletar transação {transacao_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.post("/transacoes/bulk-categorization")
async def bulk_categorization(
    request: BulkCategorizationRequest,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Categorização em massa de transações - ENDPOINT PRINCIPAL
    """
    try:
        print(f"🚀 Bulk categorization iniciada: {len(request.transaction_ids)} transações")
        print(f"🔍 DEBUG: IDs recebidos: {request.transaction_ids}")
        print(f"🔍 DEBUG: Updates: {request.updates}")
        print(f"🔍 DEBUG: Tipo: {request.tipo}")

        if not request.transaction_ids:
            raise HTTPException(status_code=400, detail="IDs de transações são obrigatórios")

        if not request.updates:
            raise HTTPException(status_code=400, detail="Campos para atualizar são obrigatórios")

        # ✅ PERMISSÃO TOTAL: Buscar transações de QUALQUER empresa (alinhado com configuração atual)
        transacoes = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.id.in_(request.transaction_ids),
            TransacaoFinanceira.tipo == request.tipo
        ).all()

        print(f"🔍 DEBUG: Transações encontradas: {len(transacoes)}")

        if not transacoes:
            print(f"❌ DEBUG: Nenhuma transação encontrada para IDs: {request.transaction_ids}")
            # Verificar se as transações existem independente do tipo
            todas_transacoes = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.id.in_(request.transaction_ids)
            ).all()
            print(f"🔍 DEBUG: Total de transações (qualquer tipo): {len(todas_transacoes)}")

            if todas_transacoes:
                tipos_encontrados = [t.tipo for t in todas_transacoes]
                print(f"🔍 DEBUG: Tipos encontrados: {tipos_encontrados}")
                raise HTTPException(
                    status_code=400, 
                    detail=f"Transações encontradas mas com tipos diferentes: {tipos_encontrados}. Esperado: {request.tipo}"
                )
            else:
                raise HTTPException(status_code=404, detail="IDs de transações não encontrados no banco")

        updated_count = 0
        field_updates = {}

        # Processar atualizações
        for transacao in transacoes:
            print(f"🔄 Processando transação ID {transacao.id}")
            for field, value in request.updates.items():
                if hasattr(transacao, field):
                    # Converter valores vazios para None
                    if value == "" or value == "null" or value == "undefined":
                        value = None

                    old_value = getattr(transacao, field)
                    setattr(transacao, field, value)

                    print(f"✅ Campo {field}: {old_value} → {value}")

                    if field not in field_updates:
                        field_updates[field] = 0
                    field_updates[field] += 1
                    updated_count += 1
                else:
                    print(f"⚠️ Campo {field} não existe no modelo TransacaoFinanceira")

        try:
            db.commit()
            print(f"✅ Commit realizado com sucesso!")
        except Exception as commit_error:
            db.rollback()
            print(f"❌ Erro no commit: {commit_error}")
            raise HTTPException(status_code=500, detail=f"Erro ao salvar no banco: {str(commit_error)}")

        result = {
            "success": True,
            "message": "Categorização em massa realizada com sucesso",
            "updated_transactions": len(transacoes),
            "total_field_updates": updated_count,
            "field_updates": field_updates,
            "transaction_ids": request.transaction_ids
        }

        print(f"🎉 Bulk categorization concluída: {result}")
        return result

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico na categorização em massa: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.delete("/transacoes/receitas/delete-all")
async def delete_all_receitas(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    confirm: str = Query(..., description="Digite 'CONFIRMAR' para executar esta operação irreversível")
):
    """
    ⚠️ OPERAÇÃO PERIGOSA: Deletar TODAS as transações de receita do banco de dados
    Esta operação é IRREVERSÍVEL e afeta TODAS as empresas!
    """
    try:
        # Validação de segurança
        if confirm != "CONFIRMAR":
            raise HTTPException(
                status_code=400,
                detail="Para executar esta operação perigosa, você deve passar o parâmetro ?confirm=CONFIRMAR"
            )

        print("🚨 OPERAÇÃO PERIGOSA: Iniciando deleção de TODAS as receitas do banco...")

        # Contar quantas receitas existem antes da deleção
        total_receitas_before = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'receita'
        ).count()

        print(f"🔍 Total de receitas encontradas: {total_receitas_before}")

        if total_receitas_before == 0:
            return {
                "message": "Nenhuma transação de receita encontrada para deletar",
                "deleted_count": 0,
                "status": "nothing_to_delete"
            }

        # Executar a deleção
        try:
            # Deletar todas as transações de receita
            deleted_count = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita'
            ).delete(synchronize_session=False)

            db.commit()

            print(f"✅ {deleted_count} transações de receita deletadas com sucesso!")

            # Verificar se realmente foram deletadas
            remaining_receitas = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita'
            ).count()

            return {
                "message": f"Todas as transações de receita foram deletadas com sucesso!",
                "deleted_count": deleted_count,
                "total_before": total_receitas_before,
                "remaining_receitas": remaining_receitas,
                "status": "success"
            }

        except Exception as delete_error:
            db.rollback()

            # Tratar especificamente erros de Foreign Key
            error_str = str(delete_error).lower()
            if "foreign key" in error_str or "constraint" in error_str:
                raise HTTPException(
                    status_code=409,
                    detail=f"Não é possível deletar as receitas porque existem dependências em outras tabelas. "
                          f"Erro: {str(delete_error)}"
                )
            else:
                raise HTTPException(
                    status_code=500,
                    detail=f"Erro ao deletar receitas: {str(delete_error)}"
                )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico ao deletar receitas: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.delete("/admin/transacoes/receitas/delete-all")
async def admin_delete_all_receitas(
    db: Session = Depends(get_db),
    confirm: str = Query(..., description="Digite 'CONFIRMAR' para executar esta operação irreversível"),
    admin_key: str = Query(..., description="Chave de administrador")
):
    """
    ⚠️ ENDPOINT ADMINISTRATIVO SEM AUTENTICAÇÃO
    OPERAÇÃO PERIGOSA: Deletar TODAS as transações de receita do banco de dados
    Esta operação é IRREVERSÍVEL e afeta TODAS as empresas!
    """
    try:
        # Validação de chave administrativa
        import os
        expected_admin_key = os.getenv("ADMIN_KEY", "admin123")
        if admin_key != expected_admin_key:
            raise HTTPException(status_code=403, detail="Chave de administrador inválida")

        # Validação de segurança
        if confirm != "CONFIRMAR":
            raise HTTPException(
                status_code=400,
                detail="Para executar esta operação perigosa, você deve passar o parâmetro ?confirm=CONFIRMAR"
            )

        print("🚨 OPERAÇÃO PERIGOSA [ADMIN]: Iniciando deleção de TODAS as receitas do banco...")

        # Contar quantas receitas existem antes da deleção
        total_receitas_before = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'receita'
        ).count()

        print(f"🔍 Total de receitas encontradas: {total_receitas_before}")

        if total_receitas_before == 0:
            return {
                "message": "Nenhuma transação de receita encontrada para deletar",
                "deleted_count": 0,
                "status": "nothing_to_delete"
            }

        # Executar a deleção
        try:
            # Deletar todas as transações de receita
            deleted_count = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita'
            ).delete(synchronize_session=False)

            db.commit()

            print(f"✅ [ADMIN] {deleted_count} transações de receita deletadas com sucesso!")

            # Verificar se realmente foram deletadas
            remaining_receitas = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'receita'
            ).count()

            return {
                "message": f"[ADMIN] Todas as transações de receita foram deletadas com sucesso!",
                "deleted_count": deleted_count,
                "total_before": total_receitas_before,
                "remaining_receitas": remaining_receitas,
                "status": "success"
            }

        except Exception as delete_error:
            db.rollback()

            # Tratar especificamente erros de Foreign Key
            error_str = str(delete_error).lower()
            if "foreign key" in error_str or "constraint" in error_str:
                raise HTTPException(
                    status_code=409,
                    detail=f"Não é possível deletar as receitas porque existem dependências em outras tabelas. "
                          f"Erro: {str(delete_error)}"
                )
            else:
                raise HTTPException(
                    status_code=500,
                    detail=f"Erro ao deletar receitas: {str(delete_error)}"
                )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico ao deletar receitas [ADMIN]: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.get("/transacoes/despesas/clear-clientes")
async def clear_clientes_from_despesas(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    confirm: str = Query(..., description="Digite 'CONFIRMAR' para executar esta operação")
):
    """
    Limpar campo cliente_id de TODAS as transações de despesa
    Despesas normalmente não deveriam ter clientes, apenas fornecedores
    """
    try:
        # Validação de segurança
        if confirm != "CONFIRMAR":
            raise HTTPException(
                status_code=400,
                detail="Para executar esta operação, você deve passar o parâmetro ?confirm=CONFIRMAR"
            )

        print("🧹 Iniciando limpeza de campo cliente_id nas despesas...")

        # Contar quantas despesas têm cliente_id preenchido
        despesas_com_cliente = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.tipo == 'despesa',
            TransacaoFinanceira.cliente_id.is_not(None)
        ).count()

        print(f"🔍 Total de despesas com cliente_id preenchido: {despesas_com_cliente}")

        if despesas_com_cliente == 0:
            return {
                "message": "Nenhuma despesa com cliente_id encontrada para limpar",
                "updated_count": 0,
                "status": "nothing_to_update"
            }

        # Executar a limpeza
        try:
            # Atualizar todas as despesas, definindo cliente_id como NULL
            updated_count = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.cliente_id.is_not(None)
            ).update(
                {TransacaoFinanceira.cliente_id: None},
                synchronize_session=False
            )

            db.commit()

            print(f"✅ {updated_count} despesas tiveram o campo cliente_id limpo com sucesso!")

            # Verificar se realmente foram limpas
            remaining_despesas_com_cliente = db.query(TransacaoFinanceira).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.cliente_id.is_not(None)
            ).count()

            return {
                "message": f"Campo cliente_id limpo com sucesso de {updated_count} despesas!",
                "updated_count": updated_count,
                "total_before": despesas_com_cliente,
                "remaining_with_cliente": remaining_despesas_com_cliente,
                "status": "success"
            }

        except Exception as update_error:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Erro ao limpar campo cliente_id das despesas: {str(update_error)}"
            )

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro crítico ao limpar clientes das despesas: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.get("/transacoes/analytics/categorization-charts")
async def get_categorization_charts(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db),
    ano: Optional[int] = Query(None),
    mes: Optional[int] = Query(None)
):
    """
    Endpoint para analytics de categorização - alimenta os 9 gráficos
    ✅ CORREÇÃO: Mostrar dados de TODAS as empresas para categorização
    """
    try:
        # ✅ CORREÇÃO: Remover filtro de empresa - empresa segregará apenas nos relatórios
        result = {
            "despesas": {
                "categoria_contabil": [],
                "subcategoria_contabil": [],
                "categoria_gerencial": [],
                "subcategoria_gerencial": [],
                "centro_custo": [],
                "conta_contabil": []
            },
            "receitas": {
                "cliente": [],
                "projeto": [],
                "produto_servico": []
            }
        }

        # 1. Categorias Contábeis (Despesas) - 🌳 APENAS LEAF NODES
        from sqlalchemy import func
        from app.models.categorias import CategoriaContabil

        # Subquery para transações pai com filhos (excluir dos gráficos)
        parents_with_children = db.query(TransacaoFinanceira.parent_id).filter(
            TransacaoFinanceira.parent_id.isnot(None)
        ).distinct().subquery()

        # Contar por categoria contábil (leaf nodes apenas)
        categoria_contabil_query = db.query(
            CategoriaContabil.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id) &
            (TransacaoFinanceira.tipo == 'despesa')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(CategoriaContabil.id, CategoriaContabil.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Categoria" (APENAS LEAF NODES)
        sem_categoria_contabil = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_categoria_contabil > 0:
            result["despesas"]["categoria_contabil"].append({
                "name": "Sem Categoria",
                "count": sem_categoria_contabil,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'despesa',
                        TransacaoFinanceira.categoria_contabil_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in categoria_contabil_query:
            if int(item.count) > 0:
                result["despesas"]["categoria_contabil"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 2. Categorias Gerenciais (Despesas) - 🌳 APENAS LEAF NODES
        from app.models.categorias import CategoriaGerencial

        categoria_gerencial_query = db.query(
            CategoriaGerencial.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.categoria_gerencial_id == CategoriaGerencial.id) &
            (TransacaoFinanceira.tipo == 'despesa')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(CategoriaGerencial.id, CategoriaGerencial.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Categoria" (APENAS LEAF NODES)
        sem_categoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.categoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_categoria_gerencial > 0:
            result["despesas"]["categoria_gerencial"].append({
                "name": "Sem Categoria",
                "count": sem_categoria_gerencial,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'despesa',
                        TransacaoFinanceira.categoria_gerencial_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in categoria_gerencial_query:
            if int(item.count) > 0:
                result["despesas"]["categoria_gerencial"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 3. Centros de Custo (Despesas) - 🌳 APENAS LEAF NODES
        from app.models.categorias import CentroCusto

        centro_custo_query = db.query(
            CentroCusto.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.centro_custo_id == CentroCusto.id) &
            (TransacaoFinanceira.tipo == 'despesa')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(CentroCusto.id, CentroCusto.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Centro de Custo" (APENAS LEAF NODES)
        sem_centro_custo = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.centro_custo_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_centro_custo > 0:
            result["despesas"]["centro_custo"].append({
                "name": "Sem Centro de Custo",
                "count": sem_centro_custo,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'despesa',
                        TransacaoFinanceira.centro_custo_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in centro_custo_query:
            if int(item.count) > 0:
                result["despesas"]["centro_custo"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 4. Contas Contábeis (Despesas)
        from app.models.auxiliares import ContaContabil

        conta_contabil_query = db.query(
            ContaContabil.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).outerjoin(
            TransacaoFinanceira,
            (TransacaoFinanceira.conta_contabil_id == ContaContabil.id) &
            (TransacaoFinanceira.tipo == 'despesa') &
            (TransacaoFinanceira.categoria_contabil_id.is_not(None)) &
            (TransacaoFinanceira.subcategoria_contabil_id.is_not(None)) &
            (TransacaoFinanceira.categoria_gerencial_id.is_not(None)) &
            (TransacaoFinanceira.subcategoria_gerencial_id.is_not(None)) &
            (TransacaoFinanceira.centro_custo_id.is_not(None))
        ).group_by(ContaContabil.id, ContaContabil.nome).all()

        # Incluir "Sem Conta Contábil" se houver despesas sem conta contábil
        sem_conta_contabil = db.query(func.count(TransacaoFinanceira.id)).filter(
            TransacaoFinanceira.tipo == 'despesa',
            TransacaoFinanceira.conta_contabil_id.is_(None)
        ).scalar() or 0

        if sem_conta_contabil > 0:
            result["despesas"]["conta_contabil"].append({
                "name": "Sem Conta Contábil",
                "count": sem_conta_contabil,
                "total_value": float(db.query(func.sum(TransacaoFinanceira.valor)).filter(
                    TransacaoFinanceira.tipo == 'despesa',
                    TransacaoFinanceira.conta_contabil_id.is_(None)
                ).scalar() or 0)
            })

        for item in conta_contabil_query:
            if int(item.count) > 0:
                result["despesas"]["conta_contabil"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 5. Receitas por Cliente - 🌳 APENAS LEAF NODES
        from app.models.clientes import Cliente

        cliente_query = db.query(
            Cliente.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.cliente_id == Cliente.id) &
            (TransacaoFinanceira.tipo == 'receita')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(Cliente.id, Cliente.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Cliente" (APENAS LEAF NODES)
        sem_cliente = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.cliente_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_cliente > 0:
            result["receitas"]["cliente"].append({
                "name": "Sem Cliente",
                "count": sem_cliente,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'receita',
                        TransacaoFinanceira.cliente_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in cliente_query:
            if int(item.count) > 0:
                result["receitas"]["cliente"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 6. Receitas por Projeto - 🌳 APENAS LEAF NODES
        from app.models.categorias import Projeto

        projeto_query = db.query(
            Projeto.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.projeto_id == Projeto.id) &
            (TransacaoFinanceira.tipo == 'receita')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(Projeto.id, Projeto.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Projeto" (APENAS LEAF NODES)
        sem_projeto = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.projeto_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_projeto > 0:
            result["receitas"]["projeto"].append({
                "name": "Sem Projeto",
                "count": sem_projeto,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'receita',
                        TransacaoFinanceira.projeto_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in projeto_query:
            if int(item.count) > 0:
                result["receitas"]["projeto"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 7. Receitas por Produto/Serviço - 🌳 APENAS LEAF NODES
        from app.models.auxiliares import ProdutoServico

        produto_servico_query = db.query(
            ProdutoServico.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.produto_servico_id == ProdutoServico.id) &
            (TransacaoFinanceira.tipo == 'receita')
        ).filter(
            ~TransacaoFinanceira.id.in_(db.query(parents_with_children))
        ).group_by(ProdutoServico.id, ProdutoServico.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Produto/Serviço" (APENAS LEAF NODES)
        sem_produto_servico = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'receita',
                TransacaoFinanceira.produto_servico_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_produto_servico > 0:
            result["receitas"]["produto_servico"].append({
                "name": "Sem Produto/Serviço",
                "count": sem_produto_servico,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'receita',
                        TransacaoFinanceira.produto_servico_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in produto_servico_query:
            if int(item.count) > 0:
                result["receitas"]["produto_servico"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 2. Subcategorias Contábeis (Despesas) - usar campo subcategoria_contabil_id
        subcategoria_contabil_query = db.query(
            CategoriaContabil.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.subcategoria_contabil_id == CategoriaContabil.id) &
            (TransacaoFinanceira.tipo == 'despesa') &
            (TransacaoFinanceira.categoria_contabil_id.is_not(None)) &
            (TransacaoFinanceira.categoria_gerencial_id.is_not(None)) &
            (TransacaoFinanceira.subcategoria_gerencial_id.is_not(None)) &
            (TransacaoFinanceira.centro_custo_id.is_not(None)) &
            (TransacaoFinanceira.conta_contabil_id.is_not(None))
        ).group_by(CategoriaContabil.id, CategoriaContabil.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Subcategoria" (APENAS LEAF NODES)
        sem_subcategoria_contabil = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_contabil_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_subcategoria_contabil > 0:
            result["despesas"]["subcategoria_contabil"].append({
                "name": "Sem Subcategoria",
                "count": sem_subcategoria_contabil,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'despesa',
                        TransacaoFinanceira.subcategoria_contabil_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in subcategoria_contabil_query:
            if int(item.count) > 0:  # Só incluir se houver transações
                result["despesas"]["subcategoria_contabil"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        # 3. Subcategorias Gerenciais (Despesas) - usar campo subcategoria_gerencial_id
        from app.models.categorias import CategoriaGerencial

        subcategoria_gerencial_query = db.query(
            CategoriaGerencial.nome.label('name'),
            func.count(TransacaoFinanceira.id).label('count'),
            func.sum(TransacaoFinanceira.valor).label('total_value')
        ).join(
            TransacaoFinanceira,
            (TransacaoFinanceira.subcategoria_gerencial_id == CategoriaGerencial.id) &
            (TransacaoFinanceira.tipo == 'despesa') &
            (TransacaoFinanceira.categoria_contabil_id.is_not(None)) &
            (TransacaoFinanceira.subcategoria_contabil_id.is_not(None)) &
            (TransacaoFinanceira.categoria_gerencial_id.is_not(None)) &
            (TransacaoFinanceira.centro_custo_id.is_not(None)) &
            (TransacaoFinanceira.conta_contabil_id.is_not(None))
        ).group_by(CategoriaGerencial.id, CategoriaGerencial.nome).all()

        # 🌳 HIERARQUIA: Incluir "Sem Subcategoria" (APENAS LEAF NODES)
        sem_subcategoria_gerencial = apply_leaf_nodes_filter(
            db.query(func.count(TransacaoFinanceira.id)).filter(
                TransacaoFinanceira.tipo == 'despesa',
                TransacaoFinanceira.subcategoria_gerencial_id.is_(None)
            ),
            db
        ).scalar() or 0

        if sem_subcategoria_gerencial > 0:
            result["despesas"]["subcategoria_gerencial"].append({
                "name": "Sem Subcategoria",
                "count": sem_subcategoria_gerencial,
                "total_value": float(apply_leaf_nodes_filter(
                    db.query(func.sum(TransacaoFinanceira.valor)).filter(
                        TransacaoFinanceira.tipo == 'despesa',
                        TransacaoFinanceira.subcategoria_gerencial_id.is_(None)
                    ),
                    db
                ).scalar() or 0)
            })

        for item in subcategoria_gerencial_query:
            if int(item.count) > 0:  # Só incluir se houver transações
                result["despesas"]["subcategoria_gerencial"].append({
                    "name": item.name,
                    "count": int(item.count),
                    "total_value": float(item.total_value or 0)
                })

        return result

    except Exception as e:
        print(f"Erro ao gerar analytics de categorização: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


# Endpoint para atualizar transação
@router.put("/transacoes/{transacao_id}")
async def update_transacao(
    transacao_id: int,
    transacao_data: TransacaoUpdate,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Atualizar uma transação específica
    """
    try:
        # ✅ PERMISSÃO TOTAL: Usuário pode editar qualquer transação independente da empresa
        transacao = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.id == transacao_id
        ).first()

        if not transacao:
            raise HTTPException(status_code=404, detail="Transação não encontrada")

        # Atualizar campos básicos
        transacao.tipo = transacao_data.tipo
        transacao.descricao = transacao_data.descricao
        transacao.nome = transacao_data.descricao  # Para compatibilidade

        # Processar valor baseado no tipo
        valor = float(transacao_data.valor)
        if transacao_data.tipo == 'despesa':
            valor = -abs(valor)  # Garantir que despesas sejam negativas
        elif transacao_data.tipo == 'receita':
            valor = abs(valor)   # Garantir que receitas sejam positivas
        transacao.valor = valor

        # Processar data de lançamento (usar apenas YYYY-MM-DD, ignorar timezone)
        if transacao_data.data_lancamento:
            try:
                date_str = transacao_data.data_lancamento[:10]
                data_lancamento = datetime.strptime(date_str, '%Y-%m-%d').date()
                transacao.data_lancamento = data_lancamento
                transacao.competencia_ano = data_lancamento.year
                transacao.competencia_mes = data_lancamento.month
            except:
                pass

        # Processar competência geral (campo unificado do formulário)
        # Atualiza TODOS os campos de competência quando apenas 'competencia' é fornecido
        # Campos específicos (competencia_contabil, competencia_gerencial) podem sobrescrever depois
        if hasattr(transacao_data, 'competencia') and transacao_data.competencia:
            try:
                parts = transacao_data.competencia.split('-')
                comp_ano = int(parts[0])
                comp_mes = int(parts[1])
                # Atualizar competência base
                transacao.competencia_ano = comp_ano
                transacao.competencia_mes = comp_mes
                # Sincronizar contábil e gerencial com a competência base
                # (serão sobrescritos se campos específicos forem fornecidos abaixo)
                transacao.competencia_ano_contabil = comp_ano
                transacao.competencia_mes_contabil = comp_mes
                transacao.competencia_ano_gerencial = comp_ano
                transacao.competencia_mes_gerencial = comp_mes
            except:
                pass  # Mantém valores existentes

        # Processar competência contábil (sobrescreve se fornecido explicitamente)
        if hasattr(transacao_data, 'competencia_contabil') and transacao_data.competencia_contabil:
            try:
                parts = transacao_data.competencia_contabil.split('-')
                transacao.competencia_ano_contabil = int(parts[0])
                transacao.competencia_mes_contabil = int(parts[1])
            except:
                pass  # Mantém valores existentes

        # Processar competência gerencial (sobrescreve se fornecido explicitamente)
        if hasattr(transacao_data, 'competencia_gerencial') and transacao_data.competencia_gerencial:
            try:
                parts = transacao_data.competencia_gerencial.split('-')
                transacao.competencia_ano_gerencial = int(parts[0])
                transacao.competencia_mes_gerencial = int(parts[1])
            except:
                pass  # Mantém valores existentes

        # Mapear forma de pagamento para valores válidos do enum
        forma_pgto_mapping = {
            'transferencia_bancaria': 'transferencia',
            'transferencia': 'transferencia',
            'cartao_credito': 'cartao',
            'cartao': 'cartao',
            'cartao_debito': 'debito',
            'debito': 'debito',
            'dinheiro': 'dinheiro',
            'pix': 'pix',
            'boleto': 'boleto',
            'cheque': 'outros',
            'outros': 'outros',
            'ted': 'transferencia',
            'doc': 'transferencia'
        }

        # Atualizar outros campos opcionais
        if transacao_data.forma_pagamento:
            forma_pgto_mapped = forma_pgto_mapping.get(transacao_data.forma_pagamento, 'outros')
            transacao.forma_pgto = forma_pgto_mapped

        if transacao_data.forma_recebimento:
            forma_recebimento_mapped = forma_pgto_mapping.get(transacao_data.forma_recebimento, 'outros')
            transacao.forma_pgto = forma_recebimento_mapped


        if transacao_data.status:
            transacao.status = transacao_data.status

        # Atualizar empresa e relacionamentos
        if transacao_data.empresa_id:
            transacao.empresa_id = transacao_data.empresa_id

        if transacao_data.cliente_id:
            transacao.cliente_id = transacao_data.cliente_id

        if transacao_data.fornecedor_id:
            transacao.fornecedor_id = transacao_data.fornecedor_id

        # Atualizar categorias contábeis
        if transacao_data.categoria_contabil_id:
            transacao.categoria_contabil_id = transacao_data.categoria_contabil_id

        if transacao_data.subcategoria_contabil_id:
            transacao.subcategoria_contabil_id = transacao_data.subcategoria_contabil_id

        # Atualizar categorias gerenciais
        if transacao_data.categoria_gerencial_id:
            transacao.categoria_gerencial_id = transacao_data.categoria_gerencial_id

        if transacao_data.subcategoria_gerencial_id:
            transacao.subcategoria_gerencial_id = transacao_data.subcategoria_gerencial_id

        # Atualizar outros relacionamentos
        if transacao_data.centro_custo_id:
            transacao.centro_custo_id = transacao_data.centro_custo_id

        if transacao_data.conta_contabil_id:
            transacao.conta_contabil_id = transacao_data.conta_contabil_id

        if transacao_data.projeto_id:
            transacao.projeto_id = transacao_data.projeto_id

        if transacao_data.produto_servico_id:
            transacao.produto_servico_id = transacao_data.produto_servico_id

        # Atualizar campos de documentação
        if transacao_data.numero_nota_fiscal:
            transacao.numero_nota_fiscal = transacao_data.numero_nota_fiscal

        if transacao_data.link_nota_fiscal:
            transacao.link_nota_fiscal = transacao_data.link_nota_fiscal

        if transacao_data.numero_pedido_compra:
            transacao.numero_pedido_compra = transacao_data.numero_pedido_compra

        if transacao_data.link_pedido_compra:
            transacao.link_pedido_compra = transacao_data.link_pedido_compra

        # Atualizar flag Cash Control
        if hasattr(transacao_data, 'exibir_no_cash_control') and transacao_data.exibir_no_cash_control is not None:
            transacao.exibir_no_cash_control = transacao_data.exibir_no_cash_control

        # Atualizar flag P&L Gerencial
        if hasattr(transacao_data, 'entra_no_gerencial') and transacao_data.entra_no_gerencial is not None:
            transacao.entra_no_gerencial = transacao_data.entra_no_gerencial

        # Processar datas adicionais (usar apenas YYYY-MM-DD, ignorar timezone)
        if transacao_data.data_vencimento:
            try:
                date_str = transacao_data.data_vencimento[:10]
                transacao.data_vencimento = datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                pass

        if transacao_data.data_recebimento:
            try:
                date_str = transacao_data.data_recebimento[:10]
                transacao.data_pagamento = datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                pass

        if transacao_data.valor_recebido is not None:
            transacao.valor_recebido = transacao_data.valor_recebido
        if transacao_data.valor_pago is not None:
            transacao.valor_pago = transacao_data.valor_pago

        print(f"✅ Salvando transação {transacao_id} com todos os campos atualizados")

        db.commit()
        db.refresh(transacao)

        return {
            "id": transacao.id,
            "tipo": transacao.tipo,
            "descricao": transacao.descricao,
            "valor": float(transacao.valor),
            "data_lancamento": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
            "status": transacao.status,
            "message": "Transação atualizada com sucesso!"
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Erro ao atualizar transação {transacao_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


# 📊 ENDPOINT PARA EXPORTAÇÃO DE TRANSAÇÕES EM EXCEL (DEVE FICAR ANTES DE ROTAS DINÂMICAS)
@router.get("/transacoes/export-excel")
async def export_transacoes_excel(
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Exportar todas as transações para arquivo Excel
    """
    try:
        print(f"📊 Iniciando exportação Excel para usuário {getattr(current_user, 'name', 'N/A')}")

        # Buscar todas as transações com informações detalhadas
        query = db.query(
            TransacaoFinanceira.id,
            TransacaoFinanceira.tipo,
            TransacaoFinanceira.data_lancamento,
            TransacaoFinanceira.nome,
            TransacaoFinanceira.descricao,
            TransacaoFinanceira.valor,
            TransacaoFinanceira.status,
            TransacaoFinanceira.forma_pgto,
            Empresa.nome_fantasia.label('empresa_nome'),
            Cliente.nome.label('cliente_nome'),
            Fornecedor.nome.label('fornecedor_nome'),
            Projeto.nome.label('projeto_nome'),
            ProdutoServico.nome.label('produto_servico_nome'),
            CategoriaContabil.nome.label('categoria_contabil_nome'),
            CategoriaGerencial.nome.label('categoria_gerencial_nome')
        ).outerjoin(
            Empresa, TransacaoFinanceira.empresa_id == Empresa.id
        ).outerjoin(
            Cliente, TransacaoFinanceira.cliente_id == Cliente.id
        ).outerjoin(
            Fornecedor, TransacaoFinanceira.fornecedor_id == Fornecedor.id
        ).outerjoin(
            Projeto, TransacaoFinanceira.projeto_id == Projeto.id
        ).outerjoin(
            ProdutoServico, TransacaoFinanceira.produto_servico_id == ProdutoServico.id
        ).outerjoin(
            CategoriaContabil, TransacaoFinanceira.categoria_contabil_id == CategoriaContabil.id
        ).outerjoin(
            CategoriaGerencial, TransacaoFinanceira.categoria_gerencial_id == CategoriaGerencial.id
        ).order_by(TransacaoFinanceira.data_lancamento.desc())

        transacoes = query.all()
        print(f"📊 Encontradas {len(transacoes)} transações para exportação")

        # Criar workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transações Financeiras"

        # Definir cabeçalhos
        headers = [
            "ID", "Tipo", "Data", "Título", "Descrição", "Valor", "Status", 
            "Forma Pagamento", "Empresa", "Cliente/Fornecedor", "Projeto", 
            "Produto/Serviço", "Categoria Contábil", "Categoria Gerencial"
        ]

        # Aplicar estilo no cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adicionar dados das transações
        for row_num, transacao in enumerate(transacoes, 2):
            ws.cell(row=row_num, column=1, value=transacao.id)
            ws.cell(row=row_num, column=2, value=transacao.tipo.upper())
            ws.cell(row=row_num, column=3, value=transacao.data_lancamento.strftime("%d/%m/%Y") if transacao.data_lancamento else "")
            ws.cell(row=row_num, column=4, value=transacao.nome or "")
            ws.cell(row=row_num, column=5, value=transacao.descricao or "")
            ws.cell(row=row_num, column=6, value=float(transacao.valor) if transacao.valor else 0)
            ws.cell(row=row_num, column=7, value=transacao.status or "")
            ws.cell(row=row_num, column=8, value=transacao.forma_pgto or "")
            ws.cell(row=row_num, column=9, value=transacao.empresa_nome or "")
            ws.cell(row=row_num, column=10, value=transacao.cliente_nome or transacao.fornecedor_nome or "")
            ws.cell(row=row_num, column=11, value=transacao.projeto_nome or "")
            ws.cell(row=row_num, column=12, value=transacao.produto_servico_nome or "")
            ws.cell(row=row_num, column=13, value=transacao.categoria_contabil_nome or "")
            ws.cell(row=row_num, column=14, value=transacao.categoria_gerencial_nome or "")

        # Ajustar largura das colunas
        column_widths = [8, 10, 12, 20, 30, 12, 10, 15, 20, 25, 20, 20, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Salvar em memória
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Gerar nome do arquivo com data atual
        from datetime import datetime
        data_atual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transacoes_bennu_finance_{data_atual}.xlsx"

        print(f"✅ Arquivo Excel gerado: {filename} com {len(transacoes)} transações")

        # Retornar como download
        return StreamingResponse(
            io.BytesIO(excel_buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Erro ao exportar Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao gerar arquivo Excel: {str(e)}")


# 🔧 ENDPOINT DINÂMICO MOVIDO PARA O FINAL - resolve conflito de rotas 422
@router.get("/transacoes/{transacao_id}")
async def get_transacao_by_id(
    transacao_id: int,
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Buscar uma transação específica por ID para edição - DEVE FICAR NO FINAL
    """
    try:
        print(f"🔍 Buscando transação {transacao_id} para edição")

        # Buscar transação sem filtro de empresa (configuração atual do sistema)
        transacao = db.query(TransacaoFinanceira).filter(
            TransacaoFinanceira.id == transacao_id
        ).first()

        if not transacao:
            print(f"❌ Transação {transacao_id} não encontrada")
            raise HTTPException(status_code=404, detail="Transação não encontrada")

        print(f"✅ Transação {transacao_id} encontrada: tipo={transacao.tipo}")

        # Buscar nomes associados (cliente/fornecedor) sem filtro de empresa
        cliente_nome = None
        fornecedor_nome = None

        if transacao.cliente_id:
            cliente = db.query(Cliente).filter(Cliente.id == transacao.cliente_id).first()
            if cliente:
                cliente_nome = cliente.nome

        if transacao.fornecedor_id:
            fornecedor = db.query(Fornecedor).filter(Fornecedor.id == transacao.fornecedor_id).first()
            if fornecedor:
                fornecedor_nome = fornecedor.nome

        # Formatar dados para retorno
        return {
            "id": transacao.id,
            "empresa_id": transacao.empresa_id,
            "tipo": transacao.tipo,
            "data_lancamento": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
            "data_vencimento": transacao.data_vencimento.isoformat() if transacao.data_vencimento else None,
            "data_emissao": transacao.data_lancamento.isoformat() if transacao.data_lancamento else None,
            "nome": transacao.nome,
            "descricao": transacao.descricao,
            "valor": float(transacao.valor) if transacao.valor else 0.0,
            "status": transacao.status or 'pendente',
            "forma_pgto": transacao.forma_pgto,
            "forma_recebimento": transacao.forma_pgto,
            "competencia_ano": transacao.competencia_ano,
            "competencia_mes": transacao.competencia_mes,
            "competencia_ano_contabil": transacao.competencia_ano_contabil,
            "competencia_mes_contabil": transacao.competencia_mes_contabil,
            "competencia_ano_gerencial": transacao.competencia_ano_gerencial,
            "competencia_mes_gerencial": transacao.competencia_mes_gerencial,
            "cliente_id": transacao.cliente_id,
            "cliente_nome": cliente_nome,
            "fornecedor_id": transacao.fornecedor_id,
            "fornecedor_nome": fornecedor_nome,
            "categoria_contabil_id": transacao.categoria_contabil_id,
            "categoria_gerencial_id": transacao.categoria_gerencial_id,
            "centro_custo_id": transacao.centro_custo_id,
            "conta_contabil_id": transacao.conta_contabil_id,
            "subcategoria_contabil_id": transacao.subcategoria_contabil_id,
            "subcategoria_gerencial_id": transacao.subcategoria_gerencial_id,
            "projeto_id": transacao.projeto_id,
            "produto_servico_id": transacao.produto_servico_id,
            "exibir_no_cash_control": getattr(transacao, 'exibir_no_cash_control', True),
            "entra_no_gerencial": getattr(transacao, 'entra_no_gerencial', True),
            "numero_nota_fiscal": transacao.numero_nota_fiscal,
            "link_nota_fiscal": transacao.link_nota_fiscal,
            "numero_pedido_compra": transacao.numero_pedido_compra,
            "link_pedido_compra": transacao.link_pedido_compra,
            "data_pagamento": transacao.data_pagamento.isoformat() if transacao.data_pagamento else None,
            "valor_recebido": float(transacao.valor_recebido) if transacao.valor_recebido is not None else None,
            "valor_pago": float(transacao.valor_pago) if transacao.valor_pago is not None else None,
            "created_at": transacao.created_at.isoformat() if transacao.created_at else None,
            "updated_at": transacao.updated_at.isoformat() if transacao.updated_at else None,
            "parent_id": transacao.parent_id,
            "tipo_filho": transacao.tipo_filho
        }

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Erro ao buscar transação {transacao_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")


@router.post("/impostos/calcular-preview")
async def calcular_impostos_preview(
    data: Dict = Body(...),
    current_user = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Calcula preview dos impostos sobre uma receita
    
    Body:
        - empresa_id: ID da empresa para buscar alíquotas específicas
        - valor: Valor da receita
    
    Returns:
        Dicionário com impostos calculados e alíquotas usadas
    """
    try:
        empresa_id = data.get('empresa_id')
        valor_receita = float(data.get('valor', 0))
        
        if valor_receita <= 0:
            return {
                "impostos": {
                    "pis": 0.0,
                    "cofins": 0.0,
                    "iss": 0.0,
                    "irpj": 0.0,
                    "csll": 0.0,
                    "total": 0.0
                },
                "aliquotas": {
                    "PIS": 0.65,
                    "COFINS": 3.0,
                    "ISS": 5.0,
                    "IRPJ": 7.93,
                    "CSLL": 2.88
                },
                "empresa_nome": None
            }
        
        # Buscar nome da empresa
        empresa_nome = None
        if empresa_id:
            empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
            if empresa:
                empresa_nome = empresa.razao_social
        
        # Buscar alíquotas cadastradas para a empresa
        aliquotas_default = {
            "PIS": 0.65,
            "COFINS": 3.0,
            "ISS": 5.0,
            "IRPJ": 7.93,
            "CSLL": 2.88
        }
        
        aliquotas = {}
        if empresa_id:
            impostos_cadastrados = db.query(Imposto).filter(
                Imposto.empresa_id == empresa_id,
                Imposto.ativo == True
            ).all()
            
            for imposto in impostos_cadastrados:
                codigo = imposto.codigo.upper() if imposto.codigo else imposto.nome.upper()
                for codigo_padrao in ["PIS", "COFINS", "ISS", "IRPJ", "CSLL"]:
                    if codigo_padrao in codigo:
                        aliquotas[codigo_padrao] = float(imposto.valor)
                        break
        
        # Preencher com valores default onde não houver cadastro
        for codigo, valor_default in aliquotas_default.items():
            if codigo not in aliquotas:
                aliquotas[codigo] = valor_default
        
        # Calcular impostos
        aliquota_pis = aliquotas.get("PIS", 0.65) / 100.0
        aliquota_cofins = aliquotas.get("COFINS", 3.0) / 100.0
        aliquota_iss = aliquotas.get("ISS", 5.0) / 100.0
        aliquota_irpj = aliquotas.get("IRPJ", 7.93) / 100.0
        aliquota_csll = aliquotas.get("CSLL", 2.88) / 100.0
        
        pis = valor_receita * aliquota_pis
        cofins = valor_receita * aliquota_cofins
        iss = valor_receita * aliquota_iss
        irpj = valor_receita * aliquota_irpj
        csll = valor_receita * aliquota_csll
        total = pis + cofins + iss + irpj + csll
        
        return {
            "impostos": {
                "pis": round(pis, 2),
                "cofins": round(cofins, 2),
                "iss": round(iss, 2),
                "irpj": round(irpj, 2),
                "csll": round(csll, 2),
                "total": round(total, 2)
            },
            "aliquotas": aliquotas,
            "empresa_nome": empresa_nome,
            "valor_receita": valor_receita
        }
        
    except Exception as e:
        print(f"❌ Erro ao calcular impostos preview: {e}")
        raise HTTPException(status_code=500, detail=f"Erro ao calcular impostos: {str(e)}")